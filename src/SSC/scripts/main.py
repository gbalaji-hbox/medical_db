#!/usr/bin/env python3
"""
SSC Data Consolidator

Consolidates raw CSV files (Patient Details, Diagnosis, Medication) into the template format
(consolidated_view-template - new.xlsx) with:
- Patient data merging by patientid
- Comorbidity mapping from Diagnosis
- Medication consolidation
- Primary/Secondary DX and ICD mapping
"""

import os
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import re
from datetime import datetime
import openpyxl

# Paths
ROOT = r'D:\Work_Folder\medical_db'
PATIENT_FILE = os.path.join(ROOT, 'src', 'SSC', 'Chronic Management Patient Details - 20260403_04-55.csv')
DIAGNOSIS_FILE = os.path.join(ROOT, 'src', 'SSC', 'Patient Diagnosis Code - 20260403_05-19.csv')
MEDICATION_FILE = os.path.join(ROOT, 'src', 'SSC', 'Patient_Medication - 20260403_06-03.csv')
PRESCRIPTION_FILE = os.path.join(ROOT, 'src', 'SSC', 'template', 'api_prescriptioncauselist_202603101243.csv')
TEMPLATE_FILE = os.path.join(ROOT, 'src', 'SSC', 'template', 'consolidated_view-template - new.xlsx')
# Output file with date_time
now = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT_FILE = os.path.join(ROOT, 'src', 'SSC', 'output', f'SSC_consolidated_{now}.xlsx')

os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

# Insurance classification (from XHI scripts)
INSURANCE_TYPE_MAP = [
    (['medicare advantage', 'med adv', 'med adv', 'medicareadvantage', 'medicare adv', 'medicare plus'], 'medicare advantage'),
    (['medicare part', 'medicare/medicare', 'medicare rail', 'medicare'], 'medicare'),
    (['medicaid', 'health link', 'mi health link', 'healthy mi', 'medicaid hmo'], 'medicaid'),
    (['dual', 'd-snp', 'duals', 'dual complete', 'd-snp', 'snp'], 'medicare advantage'),
    (['advantage', 'med adv', 'med adv ppo', 'med adv hmo', 'med adv ppo'], 'medicare advantage'),
    (['hap', 'blue', 'blue cross', 'blue care', 'bcn', 'aetna', 'cigna', 'humana', 'priority health', 'united', 'molina', 'meridian', 'wellcare', 'geha', 'align', 'aco', 'ascension', 'trinity', 'mcclaren', 'gravie', 'meritain', 'umr', 'pa i', 'pai', 'cofinity', 'cofinity', 'core', 'core', 'mutual of omaha', 'etna'], 'commercial'),
    (['self pay', 'self-pay', 'selfpay'], 'self pay'),
    (['workers compensation', 'wc ', 'wc '], 'workers comp'),
    (['motor vehicle', 'auto '], 'auto'),
    (['veterans', 'va', 'triwest', 'veterans administration', 'optum va'], 'veterans'),
    (['tricare'], 'government'),
]

# Comorbidity columns in order (for left-to-right primary/secondary dx)
COMORBIDITY_COLUMNS = [
    'CORONARY ARTERY DISEASE', 'ARRHYTHMIA', 'CONGESTIVE HEART FAILURE', 'PERIPHERAL VASCULAR',
    'VALVULAR HEART', 'CEREBROVASCULAR ACCIDENT', 'HYPERLIPIDEMIA', 'ANGINA PECTORIS', 'HYPOTENSION', 'HYPERTENSION',
    'OBESITY', 'DIABETES', 'CHRONIC KIDNEY DISEASE',
    'COPD', 'RESPIRATORY FAILURE', 'ASTHMA', 'SLEEP APNEA', 'DYSPNEA', 'EMPHYSEMA',
    'BRONCHIECTASIS', 'HYPOXEMIA'
]

# Heart-related comorbidities (prioritize for PRIMARY DX)
HEART_RELATED = [
    'CORONARY ARTERY DISEASE', 'ARRHYTHMIA', 'CONGESTIVE HEART FAILURE', 'PERIPHERAL VASCULAR',
    'VALVULAR HEART', 'CERBOVASCULAR ACCIDENT', 'ANGINA PECTORIS', 'HYPERTENSION'
]

# Cause to comorbidity mapping (from XHI scripts)
CAUSE_TO_COMORBIDITY = {
    'Coronary Artery Disease': 'CORONARY ARTERY DISEASE',
    'Arrhythmia': 'ARRHYTHMIA',
    'CHF (Congestive Heart Failure)': 'CONGESTIVE HEART FAILURE',
    'Peripheral Vascular Disease': 'PERIPHERAL VASCULAR',
    'Valvular Heart Disease': 'VALVULAR HEART',
    'Cerebrovascular Accident': 'CERBOVASCULAR ACCIDENT',
    'Hyperlipidemia': 'HYPERLIPIDEMIA',
    'Angina Pectoris': 'ANGINA PECTORIS',
    'Hypotension': 'HYPOTENSION',
    'Hypertension': 'HYPERTENSION',
    'Obesity': 'OBESITY',
    'Type 2 Diabetes': 'DIABETES',
    'Chronic Kidney': 'CHRONIC KIDNEY DISEASE',
    'COPD': 'COPD',
    'Chronic Bronchitis': 'COPD',
    'Respiratory Failure': 'RESPIRATORY FAILURE',
    'Asthma': 'ASTHMA',
    'Sleep Apnea': 'SLEEP APNEA',
    'Dyspnea': 'DYSPNEA',
    'Emphysema': 'EMPHYSEMA',
    'Emphysema ': 'EMPHYSEMA',
    'Bronchiectasis': 'BRONCHIECTASIS',
    'Hypoxemia': 'HYPOXEMIA',
    'Chronic Hypoxia': 'HYPOXEMIA'
}

COMORBIDITY_TO_CAUSE = {v: k for k, v in CAUSE_TO_COMORBIDITY.items()}

def load_icd_to_cause():
    """Load ICD to cause mapping from prescription file."""
    df = pd.read_csv(PRESCRIPTION_FILE)
    icd_to_cause = {}
    prefix_to_cause = {}
    for _, row in df.iterrows():
        icd = str(row.get('icd_code', '')).strip()
        cause = str(row.get('cause', '')).strip()
        if icd and cause:
            icd_to_cause[icd] = cause
            # Create prefix mapping (first 3 chars)
            if len(icd) >= 3:
                prefix = icd[:3]
                prefix_to_cause[prefix] = cause
            else:
                prefix_to_cause[icd] = cause
    return icd_to_cause, prefix_to_cause

def classify_insurance(ins_text: str) -> str:
    if not ins_text or pd.isna(ins_text):
        return ''
    s = str(ins_text).lower()
    for patterns, label in INSURANCE_TYPE_MAP:
        if any(p in s for p in patterns):
            return label
    return 'other'

def parse_name(full_name: str) -> Dict[str, str]:
    """Parse first, middle, last from full name."""
    if pd.isna(full_name):
        return {'first': '', 'middle': '', 'last': ''}
    parts = str(full_name).strip().split()
    if len(parts) == 0:
        return {'first': '', 'middle': '', 'last': ''}
    elif len(parts) == 1:
        return {'first': parts[0], 'middle': '', 'last': ''}
    elif len(parts) == 2:
        return {'first': parts[0], 'middle': '', 'last': parts[1]}
    else:
        return {'first': parts[0], 'middle': ' '.join(parts[1:-1]), 'last': parts[-1]}

def format_date(date_str: str):
    """Convert date string to datetime object for Excel date format."""
    if pd.isna(date_str) or not date_str:
        return None
    try:
        dt = pd.to_datetime(date_str, format='%m/%d/%Y')
        return dt.to_pydatetime()
    except:
        return None

def map_comorbidities(icds: List[str], prefix_to_cause: Dict[str, str]) -> tuple[Dict[str, str], bool, Dict[str, str]]:
    """Map ICD codes to causes using prefix matching, then to comorbidity flags.
    
    Returns:
    - comorbidity flags dict
    - has_match boolean
    - comorbidity_to_raw_icd dict mapping comorbidity names to raw ICD codes
    """
    causes = set()
    has_match = False
    comorbidity_to_raw_icd = {}
    
    for icd in icds:
        icd = str(icd).strip()
        if not icd:
            continue
        icd_prefix = icd[:3] if len(icd) >= 3 else icd
        if icd_prefix in prefix_to_cause:
            cause = prefix_to_cause[icd_prefix]
            causes.add(cause)
            has_match = True
            # Map cause to comorbidity and store the raw ICD
            if cause in CAUSE_TO_COMORBIDITY:
                comorbidity = CAUSE_TO_COMORBIDITY[cause]
                if comorbidity not in comorbidity_to_raw_icd:
                    comorbidity_to_raw_icd[comorbidity] = icd
    
    flags = {col: 'NO' for col in COMORBIDITY_COLUMNS}
    for cause in causes:
        if cause in CAUSE_TO_COMORBIDITY:
            flags[CAUSE_TO_COMORBIDITY[cause]] = 'YES'
    
    return flags, has_match, comorbidity_to_raw_icd

def consolidate_medications(med_list: List[str]) -> str:
    """Consolidate medication list by removing duplicates."""
    if not med_list:
        return ''
    unique_meds = set()
    for med in med_list:
        if pd.notna(med):
            meds = str(med).split(';')
            for m in meds:
                unique_meds.add(m.strip())
    return '; '.join(sorted(unique_meds))

def main():
    print("Starting SSC consolidation")
    # Load mappings
    icd_to_cause, prefix_to_cause = load_icd_to_cause()
    print("Loaded mappings")

    # Load data
    patient_df = pd.read_csv(PATIENT_FILE, skiprows=1)  # Skip REPORT NAME
    patient_df.columns = patient_df.columns.str.strip()
    # Combine duplicates by patientid, taking first non-empty value for each field
    def combine_first_non_empty(series):
        for val in series:
            if pd.notna(val) and str(val).strip():
                return val
        return ''
    
    patient_df = patient_df.groupby('patientid').agg(combine_first_non_empty).reset_index()
    diag_df = pd.read_csv(DIAGNOSIS_FILE, skiprows=1)
    diag_df.columns = diag_df.columns.str.strip()
    med_df = pd.read_csv(MEDICATION_FILE, skiprows=1)
    med_df.columns = med_df.columns.str.strip()
    med_df.rename(columns={'Patient chart Id': 'patientid'}, inplace=True)
    print("Loaded data")

    # Group diagnosis and medication by patientid
    diag_grouped = diag_df.groupby('patientid').agg({
        'icd10encounterdiagcode': list,
        'icd10encounterdiagdescr': list
    }).reset_index()
    med_grouped = med_df.groupby('patientid').agg({
        'med names': lambda x: '; '.join([str(m) for m in x if pd.notna(m)]),
        'encounter_provider': lambda x: ', '.join(set([str(p) for p in x if pd.notna(p)]))
    }).reset_index()

    # Merge patient with diagnosis and medication
    df = patient_df.merge(diag_grouped, on='patientid', how='left')
    df = df.merge(med_grouped, on='patientid', how='left')
    df = df.fillna('')
    print("Merged, shape:", df.shape)

    # Apply mappings
    output_rows = []
    print("Starting processing", len(df), "rows")
    for i, row in df.iterrows():
        if i % 100 == 0:
            print(f"Processed {i} rows")
        
        # Parse name from patient name
        parsed_name = parse_name(row['patient name'])
        
        icds = row.get('icd10encounterdiagcode', [])
        comorbidities, has_icd_match, comorbidity_to_raw_icd = map_comorbidities(icds, prefix_to_cause)
        consolidated_med = consolidate_medications([row.get('med names', '')])

        # Primary and secondary dx: prioritize heart-related for primary
        true_comorbidities = [col for col in COMORBIDITY_COLUMNS if comorbidities[col] == 'YES']
        # Sort to prioritize heart-related first
        true_comorbidities.sort(key=lambda x: (x not in HEART_RELATED, COMORBIDITY_COLUMNS.index(x)))
        primary_dx = ''
        secondary_dx = ''
        primary_icd = ''
        secondary_icd = ''
        if true_comorbidities:
            primary_comorb = true_comorbidities[0]
            primary_dx = primary_comorb  # Use comorbidity header
            primary_icd = comorbidity_to_raw_icd.get(primary_comorb, '')
            
            if len(true_comorbidities) > 1:
                secondary_comorb = true_comorbidities[1]
                secondary_dx = secondary_comorb  # Use comorbidity header
                secondary_icd = comorbidity_to_raw_icd.get(secondary_comorb, '')

        # Construct full name
        full_name = f"{parsed_name['first']} {parsed_name['middle']} {parsed_name['last']}".strip()
        if not full_name:
            full_name = row['patient name']

        # Address
        address = f"{row.get('patient address2', '')} {row.get('patient address1', '')}".strip()

        output_row = {
            'EMR ID': row['patientid'],
            'PATIENT EMR NAME': row['patient name'],
            'FIRST NAME': parsed_name['first'],
            'MIDDLE NAME': parsed_name['middle'],
            'LAST NAME': parsed_name['last'],
            'PATIENT FULL NAME': full_name,
            'DATE OF BIRTH': format_date(row['patientdob']),
            'GENDER': row['patientsex'],
            'STREET ADDRESS': address,
            'CITY': row['patient city'],
            'STATE': row['patient state'],
            'ZIP': row['patient zip'],
            'HOME PHONE': row['patient homephone'],
            'MOBILE PHONE': row['patient mobile no'],
            'WORK PHONE': row['patient workphone'],
            'EMAIL ADDRESS': row['patient email'],
            'LANGUAGE': row['patient lang'],
            'RACE': row['race'],
            'EMERGENCY CONTACT NAME': row['ptnt emrgncy cntct name'],
            'EMERGENCY RELATIONSHIP': row['ptnt emrgncy cntct rltnshp'],
            'EMERGENCY CONTACT HOME PHONE': row['ptnt emrgncy cntct ph'],
            'EMERGENCY CONTACT MOBILE PHONE': '',
            'MEDICARE ID': row['patient primary policyidnumber'] if classify_insurance(row['patient primary ins pkg name']) == 'medicare' else '',
            'PRIMARY INSURANCE': row['patient primary ins pkg name'],
            'PRIMARY ID': row['patient primary policyidnumber'],
            'PRIMARY GROUP': row['patient primary policygrpnu...'],
            'SECONDARY INSURANCE': row['patient secondary ins pkg name'],
            'SECONDARY ID': row['patient secondary policyidn...'],
            'SECONDARY GROUP': row['patient secondary policygrp...'],
            'TERITARY INSURANCE': row['patient tertiary ins pkg name'],
            'TERITARY ID': row['patient tertiary policyidnu...'],
            'TERITARY GROUP': row['patient tertiary policygrpn...'],
            'INSURANCE TYPE': classify_insurance(row['patient primary ins pkg name']),
            'CO-PAY': row['patient insexpctcopay'],
            **comorbidities,
            'PRIMARY DX': primary_dx,
            'SECONDARY DX': secondary_dx,
            'PRIMARY ICD': primary_icd,
            'SECONDARY ICD': secondary_icd,
            'LAST SEEN DATE': format_date(row['patientlastseend']),
            'NEXT APPT': format_date(row['patientnextappt']),
            'PROVIDER DATA': row['prim prvdrfullnme'],
            'PROVIDER NAME': row['prim prvdrfullnme'],
            'CLINIC FACILITY': row['reg dprtmnt'],
            'PRIMARY CARE PROVIDER': row['patientinspcprvdr'],
            'MEDICATIONS': consolidated_med,
            'ENCOUNTER NOTES': '',
        }
        output_rows.append((output_row, has_icd_match))

    print("Processing done, creating dataframe")
    # Filter: only patients with primary insurance AND primary DX, and not self-pay only
    def is_valid_insurance(row):
        primary_ins = row['PRIMARY INSURANCE']
        secondary_ins = row['SECONDARY INSURANCE']
        tertiary_ins = row['TERITARY INSURANCE']
        primary_type = classify_insurance(primary_ins)
        return primary_type != 'self pay' or secondary_ins or tertiary_ins
    
    filtered_rows = [row for row, has_match in output_rows if row['PRIMARY INSURANCE'] and row['PRIMARY DX'] and is_valid_insurance(row)]
    print(f"Filtered from {len(output_rows)} to {len(filtered_rows)} rows")
    
    # Create output DataFrame
    output_df = pd.DataFrame(filtered_rows)

    # Load template to get column order
    template_df = pd.read_excel(TEMPLATE_FILE, nrows=0)
    template_cols = list(template_df.columns)

    # Reorder columns to match template
    output_df = output_df[template_cols]

    print("Saving")
    # Save to Excel
    output_df.to_excel(OUTPUT_FILE, index=False)
    print(f"Consolidated data saved to {OUTPUT_FILE}")

    # Format Excel file
    format_excel_file(OUTPUT_FILE)
    print(f"Excel formatting applied to {OUTPUT_FILE}")

def format_excel_file(file_path: str):
    """Format Excel file to set EMR ID as text and dates as date format."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find column indices
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    emr_id_col = header_row.index('EMR ID') + 1  # 1-based
    date_cols = []
    for col_name in ['DATE OF BIRTH', 'LAST SEEN DATE', 'NEXT APPT']:
        if col_name in header_row:
            date_cols.append(header_row.index(col_name) + 1)  # 1-based

    # Set EMR ID column to text format
    for row in range(2, ws.max_row + 1):  # Skip header
        cell = ws.cell(row=row, column=emr_id_col)
        cell.number_format = '@'  # Text format

    # Set date columns to date format
    for col in date_cols:
        for row in range(2, ws.max_row + 1):  # Skip header
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.number_format = 'mm/dd/yyyy'

    wb.save(file_path)

if __name__ == '__main__':
    main()