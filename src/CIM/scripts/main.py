#!/usr/bin/env python3
"""
CIM Data Cleaner and Template Formatter

Transforms the CIM raw Excel (Final_Hbox_3_19_26.xlsx) to match the consolidated view template
(consolidated_view-template - new.xlsx) with:
- Patient name parsing (First/Middle/Last Name)
- Comorbidity mapping from Registries and Problem List
- Insurance type classification
- Provider name parsing
- Missing column additions
- Generates unique diseases dictionary for future use
"""

import os
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

# Insurance type classification constants (copied from MCA/convert_to_consolidate.py)
INSURANCE_TYPE_MAP = [
    (['medicare advantage', 'med adv', 'med adv', 'medicareadvantage', 'medicare adv', 'medicare plus'], 'medicare advantage'),
    (['medicare part', 'medicare/medicare', 'medicare rail', 'medicare'], 'medicare'),
    (['medicaid', 'health link', 'mi health link', 'healthy mi', 'medicaid hmo'], 'medicaid'),
    (['dual', 'd-snp', 'duals', 'dual complete', 'd-snp', 'snp'], 'medicare advantage'),
    (['advantage', 'med adv', 'med adv ppo', 'med adv hmo', 'med adv ppo'], 'medicare advantage'),
    (['hap', 'blue', 'blue cross', 'blue care', 'bcn', 'aetna', 'cigna', 'humana', 'priority health', 'united', 'molina', 'meridian', 'wellcare', 'geha', 'align', 'aco', 'ascension', 'trinity', 'mcclaren', 'gravie', 'meritain', 'umr', 'pa i', 'pai', 'cofinity', 'cofinity', 'core', 'core', 'mutual of omaha', 'etna'], 'commercial'),
    (['self pay', 'self-pay', 'selfpay'], 'self pay'),
    (['workers compensation', 'wc ', 'wc '], 'workers comp'),
    (['motor vehicle', 'auto '], 'auto'),
    (['veterans', 'va', 'triwest', 'veterans administration', 'optum va'], 'veterans'),
    (['tricare'], 'government'),
]


def classify_insurance(ins_text: str) -> str:
    """Classify insurance type based on observed Primary Cvg/Primary Payer values."""
    if not ins_text or not str(ins_text).strip() or str(ins_text).lower().strip() in ('nan', 'none', 'null', ''):
        return ''
    s = str(ins_text).lower()
    for patterns, label in INSURANCE_TYPE_MAP:
        for p in patterns:
            if p and p in s:
                return label
    return 'other'


class CIMTemplateFormatter:
    """Class to format CIM raw data to match consolidated view template."""

    def __init__(self, input_excel: str, output_excel: str):
        self.input_excel = input_excel
        self.output_excel = output_excel

        # Load cause to ICD mapping
        self.cause_to_icd = self._load_cause_mapping()

        # Comorbidity columns
        self.comorbidity_columns = [
            'CORONARY ARTERY DISEASE', 'ARRHYTHMIA', 'CONGESTIVE HEART FAILURE',
            'PERIPHERAL VASCULAR', 'VALVULAR HEART', 'CEREBROVASCULAR ACCIDENT',
            'HYPERLIPIDEMIA', 'ANGINA PECTORIS', 'HYPOTENSION', 'HYPERTENSION',
            'OBESITY', 'DIABETES', 'CHRONIC KIDNEY DISEASE', 'COPD',
            'RESPIRATORY FAILURE', 'ASTHMA', 'SLEEP APNEA', 'DYSPNEA',
            'EMPHYSEMA', 'BRONCHIECTASIS', 'HYPOXEMIA'
        ]

        # Comorbidity to ICD mapping
        self.comorbidity_to_icd = {
            'CONGESTIVE HEART FAILURE': self.cause_to_icd.get('CHF (Congestive Heart Failure)', ''),
            'HYPERTENSION': self.cause_to_icd.get('Hypertension', ''),
            'OBESITY': self.cause_to_icd.get('Obesity', ''),
            'COPD': self.cause_to_icd.get('COPD', ''),
            'ASTHMA': self.cause_to_icd.get('Asthma', ''),
            'SLEEP APNEA': self.cause_to_icd.get('Sleep Apnea', ''),
            'CORONARY ARTERY DISEASE': self.cause_to_icd.get('Coronary Artery Disease', ''),
            'ARRHYTHMIA': self.cause_to_icd.get('Arrhythmia', ''),
            'PERIPHERAL VASCULAR': self.cause_to_icd.get('Peripheral Vascular Disease', ''),
            'DIABETES': self.cause_to_icd.get('Type 2 Diabetes', ''),
            'CHRONIC KIDNEY DISEASE': self.cause_to_icd.get('Chronic Kidney', ''),
            'CEREBROVASCULAR ACCIDENT': self.cause_to_icd.get('Cerebrovascular Accident', ''),
            'VALVULAR HEART': self.cause_to_icd.get('Valvular Heart Disease', ''),
            'ANGINA PECTORIS': self.cause_to_icd.get('Angina Pectoris', ''),
            'HYPERLIPIDEMIA': self.cause_to_icd.get('Hyperlipidemia', ''),
            'HYPOTENSION': self.cause_to_icd.get('Hypotension', ''),
            'DYSPNEA': self.cause_to_icd.get('Dyspnea', ''),
            'EMPHYSEMA': self.cause_to_icd.get('Emphysema ', ''),
            'HYPOXEMIA': self.cause_to_icd.get('Hypoxemia', ''),
            'RESPIRATORY FAILURE': self.cause_to_icd.get('Respiratory Failure', ''),
            'BRONCHIECTASIS': self.cause_to_icd.get('Bronchiectasis', ''),
        }

        # Registry to comorbidity mapping
        self.registry_to_comorbidity = {
            'Hypertension Registry': 'HYPERTENSION',
            'Coronary Artery Disease Registry': 'CORONARY ARTERY DISEASE',
            'Obesity Registry': 'OBESITY',
            'COPD Registry': 'COPD',
            'Asthma Registry': 'ASTHMA',
            'Peripheral Vascular Disease Registry': 'PERIPHERAL VASCULAR',
            'Arrhythmia Registry': 'ARRHYTHMIA',
            'Sleep Apnea': 'SLEEP APNEA',
            'Diabetes Registry': 'DIABETES',
            'Chronic Kidney Disease Registry': 'CHRONIC KIDNEY DISEASE',
            'Congestive Heart Failure Registry': 'CONGESTIVE HEART FAILURE',
            'Depression Registry': 'DEPRESSION',  # Not in template, but for completeness
            'Cancer Population Registry': 'CANCER',  # Not in template
            'ICU Stay Registry': 'ICU_STAY',  # Not in template
            # Add more as needed
        }

        # Problem list keywords to comorbidity mapping
        self.problem_keywords = {
            'HYPERTENSION': ['hypertension', 'primary hypertension', 'essential hypertension'],
            'CORONARY ARTERY DISEASE': ['coronary artery disease', 'coronary artery calcification', 'coronary atherosclerotic disease', 'ashd', 'arteriosclerotic heart disease'],
            'ARRHYTHMIA': ['arrhythmia', 'atrial fibrillation', 'atrial flutter', 'cardiac arrhythmia', 'a-fib', 'a fib'],
            'CONGESTIVE HEART FAILURE': ['heart failure', 'congestive heart failure', 'chf', 'chronic diastolic heart failure'],
            'PERIPHERAL VASCULAR': ['peripheral vascular disease', 'peripheral vascular'],
            'VALVULAR HEART': ['valvular heart disease', 'valvular disorder', 'mitral valve', 'aortic valve', 'tricuspid valve'],
            'CEREBROVASCULAR ACCIDENT': ['stroke', 'cva', 'cerebrovascular accident', 'transient ischemic attack', 'tia'],
            'HYPERLIPIDEMIA': ['hyperlipidemia', 'hyperlipemia', 'mixed hyperlipidemia', 'other hyperlipidemia'],
            'ANGINA PECTORIS': ['angina', 'angina pectoris', 'stable angina'],
            'HYPOTENSION': ['hypotension'],
            'OBESITY': ['obesity', 'class 3 severe obesity', 'severe obesity'],
            'DIABETES': ['diabetes', 'type 2 diabetes', 'diabetes mellitus'],
            'CHRONIC KIDNEY DISEASE': ['chronic kidney disease', 'ckd', 'stage 3b chronic kidney disease'],
            'COPD': ['copd', 'chronic obstructive pulmonary disease'],
            'RESPIRATORY FAILURE': ['respiratory failure'],
            'ASTHMA': ['asthma'],
            'SLEEP APNEA': ['sleep apnea', 'osa', 'obstructive sleep apnea'],
            'DYSPNEA': ['dyspnea', 'sob', 'shortness of breath'],
            'EMPHYSEMA': ['emphysema'],
            'BRONCHIECTASIS': ['bronchiectasis'],
            'HYPOXEMIA': ['hypoxemia', 'hypoxia', 'chronic hypoxia'],
        }

        # Template column order
        self.template_columns = [
            'EMR ID', 'PATIENT EMR NAME', 'FIRST NAME', 'MIDDLE NAME', 'LAST NAME',
            'PATIENT FULL NAME', 'DATE OF BIRTH', 'GENDER', 'STREET ADDRESS', 'CITY',
            'STATE', 'ZIP', 'HOME PHONE', 'MOBILE PHONE', 'WORK PHONE', 'EMAIL ADDRESS',
            'LANGUAGE', 'RACE', 'EMERGENCY CONTACT NAME', 'EMERGENCY RELATIONSHIP',
            'EMERGENCY CONTACT HOME PHONE', 'EMERGENCY CONTACT MOBILE PHONE', 'MEDICARE ID',
            'PRIMARY INSURANCE', 'PRIMARY ID', 'PRIMARY GROUP', 'SECONDARY INSURANCE',
            'SECONDARY ID', 'SECONDARY GROUP', 'TERITARY INSURANCE', 'TERITARY ID',
            'TERITARY GROUP', 'INSURANCE TYPE', 'CO-PAY',
            # Comorbidities
            'CORONARY ARTERY DISEASE', 'ARRHYTHMIA', 'CONGESTIVE HEART FAILURE',
            'PERIPHERAL VASCULAR', 'VALVULAR HEART', 'CEREBROVASCULAR ACCIDENT',
            'HYPERLIPIDEMIA', 'ANGINA PECTORIS', 'HYPOTENSION', 'HYPERTENSION',
            'OBESITY', 'DIABETES', 'CHRONIC KIDNEY DISEASE', 'COPD',
            'RESPIRATORY FAILURE', 'ASTHMA', 'SLEEP APNEA', 'DYSPNEA',
            'EMPHYSEMA', 'BRONCHIECTASIS', 'HYPOXEMIA',
            # Diagnosis fields
            'PRIMARY DX', 'SECONDARY DX', 'PRIMARY ICD', 'SECONDARY ICD',
            'LAST SEEN DATE', 'NEXT APPT', 'PROVIDER DATA', 'PROVIDER NAME',
            'CLINIC FACILITY', 'PRIMARY CARE PROVIDER', 'MEDICATIONS', 'ENCOUNTER NOTES'
        ]

        # Unique diseases dictionary
        self.unique_diseases = {}

    def _load_cause_mapping(self) -> Dict[str, str]:
        """Load cause to ICD mapping from api_prescriptioncauselist.csv."""
        cause_file = Path(__file__).parent.parent / 'template' / 'api_prescriptioncauselist_202603101243.csv'
        cause_df = pd.read_csv(cause_file)
        # Create mapping from cause name to ICD code
        cause_to_icd = dict(zip(cause_df['cause'], cause_df['icd_code']))
        return cause_to_icd

    def _parse_patient_name(self, full_name: str) -> Tuple[str, str, str, str]:
        """Parse full name into components."""
        if not full_name:
            return '', '', '', ''
        full_name = full_name.strip()
        if ',' not in full_name:
            # Assume single word is first name
            return full_name, '', '', full_name
        # Assuming format: Last, First Middle
        last_name, first_part = full_name.split(',', 1)
        last_name = last_name.strip()
        first_part = first_part.strip()

        name_parts = first_part.split()
        first_name = name_parts[0] if name_parts else ''
        middle_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''

        full_name_formatted = f"{first_name} {middle_name} {last_name}".strip()

        return last_name, first_name, middle_name, full_name_formatted

    def _parse_date(self, date_str):
        """Parse date string to datetime."""
        if not date_str or pd.isna(date_str) or str(date_str).strip() == '':
            return None
        try:
            return pd.to_datetime(date_str)
        except:
            return None

    def _parse_provider_name(self, pcp: str) -> str:
        """Parse PCP to First Last."""
        if not pcp:
            return ''
        parts = [p.strip() for p in pcp.split(',')]
        if len(parts) >= 2:
            first = parts[1].split()[0] if parts[1] else ''
            last = parts[0]
            return f"{first} {last}".strip()
        return pcp

    def _get_comorbidities(self, registries: str, problems: str) -> Dict[str, str]:
        """Get comorbidity flags from registries and problems."""
        comorbidities = {col: 'NO' for col in self.comorbidity_columns}

        # From registries
        if registries:
            reg_list = [r.strip() for r in registries.split('\n') if r.strip()]
            for reg in reg_list:
                if reg in self.registry_to_comorbidity:
                    comorb = self.registry_to_comorbidity[reg]
                    if comorb in comorbidities:
                        comorbidities[comorb] = 'YES'
                # Add to unique diseases
                if reg not in self.unique_diseases:
                    self.unique_diseases[reg] = self.registry_to_comorbidity.get(reg, 'UNKNOWN')

        # From problems
        if problems:
            prob_list = [p.strip() for p in problems.split(';') if p.strip()]
            for prob in prob_list:
                prob_lower = prob.lower()
                for comorb, keywords in self.problem_keywords.items():
                    if any(kw in prob_lower for kw in keywords):
                        if comorb in comorbidities:
                            comorbidities[comorb] = 'YES'
                        # Add to unique diseases
                        if prob not in self.unique_diseases:
                            self.unique_diseases[prob] = comorb
                        break

        return comorbidities

    def _get_primary_secondary_dx(self, comorbidities: Dict[str, str]) -> Tuple[str, str, str, str]:
        """Get primary and secondary DX and ICD."""
        yes_comorbs = [col for col, val in comorbidities.items() if val == 'YES']
        primary_dx = yes_comorbs[0] if yes_comorbs else ''
        secondary_dx = yes_comorbs[1] if len(yes_comorbs) > 1 else ''
        primary_icd = self.comorbidity_to_icd.get(primary_dx, '')
        secondary_icd = self.comorbidity_to_icd.get(secondary_dx, '')
        return primary_dx, secondary_dx, primary_icd, secondary_icd

    def process_data(self):
        """Process the raw data and save to template format."""
        # Load raw data
        df = pd.read_excel(self.input_excel)

        # Prepare output data
        output_data = []

        for _, row in df.iterrows():
            # Parse patient name
            last_name, first_name, middle_name, full_name_formatted = self._parse_patient_name(str(row.get('Patient', '')))

            # Combine street address
            street_addr1 = str(row.get('Street Address', '')).strip()
            street_addr2 = str(row.get('Street Address.1', '')).strip()
            if street_addr2 and street_addr2.lower() not in ('', 'nan'):
                street_address = f"{street_addr2}, {street_addr1}".strip()
            else:
                street_address = street_addr1

            # Get comorbidities
            comorbidities = self._get_comorbidities(str(row.get('Registries', '')), str(row.get('Problem List', '')))

            # Get primary/secondary DX
            primary_dx, secondary_dx, primary_icd, secondary_icd = self._get_primary_secondary_dx(comorbidities)

            # Handle copay
            co_pay_val = row.get('Copay Due', '')
            if not co_pay_val or str(co_pay_val).strip() == '':
                co_pay = 0.0
            else:
                try:
                    co_pay = float(co_pay_val)
                except:
                    co_pay = 0.0

            # Clean phone numbers
            def clean_phone(phone):
                phone = str(phone).strip()
                if phone in ('0000000000', '9999999999', '') or phone.lower() == 'nan':
                    return ''
                return phone

            home_phone = clean_phone(row.get('Patient Home Phone', ''))
            mobile_phone = clean_phone(row.get('Patient Cell Phone', ''))
            emerg_home_phone = clean_phone(row.get('Emerg Contact Ph', ''))
            emerg_mobile_phone = ''  # Not in raw

            # Clean clinic facility
            clinic_facility = str(row.get('Dept/Loc', '')).strip()
            clinic_facility = re.sub(r'\s*\[.*?\]', '', clinic_facility).strip()

            # Parse provider name from Encounter Provider
            encounter_provider = str(row.get('Encounter Provider', ''))
            provider_name = self._parse_provider_name(encounter_provider) if encounter_provider else ''

            # Build row
            row_data = {
                'EMR ID': str(row.get('MRN', '')),
                'PATIENT EMR NAME': row.get('Patient', ''),
                'FIRST NAME': first_name,
                'MIDDLE NAME': middle_name,
                'LAST NAME': last_name,
                'PATIENT FULL NAME': full_name_formatted,
                'DATE OF BIRTH': self._parse_date(str(row.get('DOB', ''))),
                'GENDER': row.get('Sex', ''),
                'STREET ADDRESS': street_address,
                'CITY': row.get('Pt City', ''),
                'STATE': row.get('Patient State', ''),
                'ZIP': row.get('ZIP Code', ''),
                'HOME PHONE': home_phone,
                'MOBILE PHONE': mobile_phone,
                'WORK PHONE': '',
                'EMAIL ADDRESS': row.get('Pt. E-mail Address', ''),
                'LANGUAGE': row.get('Language', ''),
                'RACE': row.get('Race', ''),
                'EMERGENCY CONTACT NAME': self._parse_provider_name(str(row.get('Primary Emer Cont Name', ''))) if row.get('Primary Emer Cont Name') and str(row.get('Primary Emer Cont Name')).lower() != 'nan' else '',
                'EMERGENCY RELATIONSHIP': row.get('Primary Emer Cont Rel', ''),
                'EMERGENCY CONTACT HOME PHONE': emerg_home_phone,
                'EMERGENCY CONTACT MOBILE PHONE': emerg_mobile_phone,
                'MEDICARE ID': row.get('Medicare Sub ID', ''),
                'PRIMARY INSURANCE': row.get('Primary Payer', ''),
                'PRIMARY ID': row.get('Primary Mem ID', ''),
                'PRIMARY GROUP': row.get('Pat Primary CVG Payer ID', ''),
                'SECONDARY INSURANCE': row.get('Secondary Payer', ''),
                'SECONDARY ID': row.get('Secondary Mem ID', ''),
                'SECONDARY GROUP': '',
                'TERITARY INSURANCE': row.get('Tertiary Payer', ''),
                'TERITARY ID': row.get('Tertiary Mem ID', ''),
                'TERITARY GROUP': '',
                'INSURANCE TYPE': classify_insurance(str(row.get('Primary Payer', ''))),
                'CO-PAY': co_pay,
                **comorbidities,
                'PRIMARY DX': primary_dx,
                'SECONDARY DX': secondary_dx,
                'PRIMARY ICD': primary_icd,
                'SECONDARY ICD': secondary_icd,
                'LAST SEEN DATE': self._parse_date(str(row.get('Last Visit Date', ''))),
                'NEXT APPT': self._parse_date(str(row.get('Next Appt', ''))),
                'PROVIDER DATA': row.get('Encounter Provider', ''),
                'PROVIDER NAME': provider_name,
                'CLINIC FACILITY': clinic_facility,
                'PRIMARY CARE PROVIDER': row.get('PCP', ''),
                'MEDICATIONS': row.get('Current Medications', ''),
                'ENCOUNTER NOTES': ''
            }
            output_data.append(row_data)

        # Save using openpyxl with formatting
        wb = Workbook()
        ws = wb.active

        # Write headers
        for col_num, header in enumerate(self.template_columns, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, row_data in enumerate(output_data, 2):
            for col_num, col_name in enumerate(self.template_columns, 1):
                value = row_data.get(col_name, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_name == 'EMR ID':
                    cell.data_type = 'str'
                elif col_name in ['DATE OF BIRTH', 'LAST SEEN DATE', 'NEXT APPT']:
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.date()
                        cell.number_format = 'mm-dd-yyyy'

        # Save
        wb.save(self.output_excel)

        # Save unique diseases dictionary
        diseases_file = Path(self.output_excel).parent / 'unique_diseases_mapping.json'
        import json
        with open(diseases_file, 'w') as f:
            json.dump(self.unique_diseases, f, indent=4)

        print(f"Processed {len(output_data)} rows. Output saved to {self.output_excel}")
        print(f"Unique diseases mapping saved to {diseases_file}")
        import json
        with open(diseases_file, 'w') as f:
            json.dump(self.unique_diseases, f, indent=4)

        print(f"Processed {len(output_data)} rows. Output saved to {self.output_excel}")
        print(f"Unique diseases mapping saved to {diseases_file}")


if __name__ == '__main__':
    # Paths
    input_excel = Path(__file__).parent.parent / 'Final_Hbox_3_19_26.xlsx'
    output_dir = Path(__file__).parent.parent / 'output'
    output_excel = output_dir / f'CIM_consolidated_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    # Create output directory if not exists
    output_dir.mkdir(exist_ok=True)

    # Process
    formatter = CIMTemplateFormatter(str(input_excel), str(output_excel))
    formatter.process_data()