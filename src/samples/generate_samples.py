#!/usr/bin/env python3
"""
Generate sample raw input files for all six ETL modules.

EHR system per module:
  MCA  → CGM APRIMA   (main cardiology EMR)
  HCT  → NextGen      (location-wise ICD / demographics reports)
  SSC  → Athena       (chronic care management CSV exports)
  CAM  → Epic (Henry Ford) — registry / problem list export
  CIM  → Epic (Henry Ford) — intensive care management export
  XHI  → DrChrono     (external EMR final report CSVs)

Each file is a faithful replica of a real export from that system:
  - Authentic FILTERS / [Report Name] metadata rows in row 1
  - Exact column headers the pipeline cleaners expect
  - 5 randomly generated (but realistically formatted) patient records
  - Same row-offset structure so the pipeline could actually parse them

Run from project root:
    python src/samples/generate_samples.py
"""

from pathlib import Path
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Shared fake-data pools — realistic cardiology demographics
# ---------------------------------------------------------------------------
LAST_NAMES  = ["Martinez", "Johnson", "Patel", "Williams", "Chen"]
FIRST_NAMES = ["Robert",   "Mary",    "James", "Patricia", "Michael"]
MID_INI     = ["A",        "B",       "C",     "D",        "E"]
STREETS  = ["123 Oak St", "456 Maple Ave", "789 Pine Rd", "321 Elm Blvd", "654 Cedar Ln"]
CITIES   = ["Fort Worth", "Fort Worth", "Granbury", "Weatherford", "Fort Worth"]
STATES   = ["TX", "TX", "TX", "TX", "TX"]
ZIPS     = ["76104", "76107", "76049", "76086", "76109"]
DOBS     = ["01/15/1952", "03/22/1948", "07/04/1955", "11/30/1960", "09/12/1943"]
DOBS_ISO = ["1952-01-15", "1948-03-22", "1955-07-04", "1960-11-30", "1943-09-12"]
AGES     = [74, 78, 71, 66, 83]
SEXES    = ["M", "F", "M", "F", "M"]
GENDERS  = ["Male", "Female", "Male", "Female", "Male"]
PHONES_H = ["(817) 555-1001", "(817) 555-2002", "(817) 555-3003", "(817) 555-4004", "(817) 555-5005"]
PHONES_C = ["(817) 555-6006", "(817) 555-7007", "(817) 555-8008", "(817) 555-9009", "(817) 555-0010"]
EMAILS   = ["r.martinez@gmail.com", "m.johnson@yahoo.com",
            "j.patel@gmail.com",    "p.williams@hotmail.com", "m.chen@gmail.com"]
RACES    = ["White", "Black or African American", "Asian", "White", "Asian"]
LANGS    = ["English", "English", "English", "Spanish", "English"]

PATIENT_IDS  = ["109001", "109002", "109003", "109004", "109005"]
MEDICARE_IDS = ["1EG4-TE5-MK72", "2AB3-CD4-EF56", "3GH7-IJ8-KL90", "4MN1-OP2-QR34", "5ST6-UV7-WX89"]

PRIM_INS  = ["Medicare",         "BlueCross BlueShield of TX", "Aetna Medicare Advantage",
             "United Healthcare", "Humana Gold Plus"]
PRIM_IDS  = ["MCR10000001", "BCB20000002", "AET30000003", "UHC40000004", "HUM50000005"]
PRIM_GRP  = ["",            "GRP-BC001",   "GRP-AET002",  "GRP-UHC003",  "GRP-HUM004"]

SEC_INS   = ["",              "Medicaid",       "Cigna",          "Molina Healthcare", "Priority Health"]
SEC_IDS   = ["",              "MCD70000007",    "CGN80000008",    "MOL90000009",       "PH10000010"]
SEC_GRP   = ["",              "",               "GRP-CGN001",     "",                  "GRP-PH002"]

PROVIDERS = [
    "KHAMBATTA, SHEREZADE J", "MESIHA, NANCY", "EDWARDS, MATTHEW",
    "KHAMBATTA, SHEREZADE J", "MESIHA, NANCY"
]
FACILITIES = [
    "Heart Ctr Of N TX Fort Worth", "Heart Ctr Of N TX Fort Worth",
    "Heart Ctr Of N TX Granbury",   "Heart Ctr Of N TX Weatherford",
    "Heart Ctr Of N TX Fort Worth"
]
COPAYS   = [20.00, 30.00, 15.00, 25.00, 10.00]
MEDS     = [
    "Lisinopril 10 MG Tablet",
    "Metformin HCl 500 MG Tablet",
    "Atorvastatin 40 MG Tablet",
    "Amlodipine Besylate 5 MG Tablet",
    "Furosemide 20 MG Tablet"
]

# ICD-10 codes for five patients (primary + secondary each)
DIAG_CODES = [
    ("I50.9",  "I10"),
    ("I25.10", "E11.65"),
    ("J44.1",  "I10"),
    ("I49.9",  "G47.33"),
    ("N18.3",  "I10"),
]
DIAG_DESCS = [
    ("Heart failure, unspecified",                          "Essential (primary) hypertension"),
    ("Atherosclerotic heart disease of native coronary artery", "Type 2 diabetes mellitus with hyperglycemia"),
    ("Chronic obstructive pulmonary disease with acute exacerbation", "Essential (primary) hypertension"),
    ("Unspecified cardiac arrhythmia",                      "Obstructive sleep apnea (adult) (pediatric)"),
    ("Chronic kidney disease, stage 3b",                    "Essential (primary) hypertension"),
]

EMERG_NAMES  = ["Susan Martinez",  "Thomas Johnson", "Priya Patel",    "Kevin Williams", "Wei Chen"]
EMERG_RELS   = ["Spouse",          "Son",            "Daughter",       "Wife",           "Daughter"]
EMERG_PHONES = ["(817) 555-1111",  "(817) 555-2222", "(817) 555-3333", "(817) 555-4444", "(817) 555-5555"]

REGISTRIES = [
    "Congestive Heart Failure Registry\nHypertension Registry",
    "Coronary Artery Disease Registry\nDiabetes Registry",
    "COPD Registry\nHypertension Registry",
    "Arrhythmia Registry\nSleep Apnea",
    "Chronic Kidney Disease Registry\nHypertension Registry",
]
PROBLEM_LISTS = [
    "Heart failure, unspecified; Essential hypertension",
    "Coronary artery disease; Type 2 diabetes mellitus",
    "COPD; Essential hypertension",
    "Atrial fibrillation; Sleep apnea; Hypertension",
    "Stage 3b chronic kidney disease; Hypertension",
]

DATES_LAST      = ["01/10/2026", "02/14/2026", "03/05/2026", "01/28/2026", "02/20/2026"]
DATES_NEXT      = ["04/15/2026", "05/10/2026", "06/02/2026", "04/30/2026", "05/22/2026"]
DATETIMES_LAST  = ["01/10/2026 09:00 AM", "02/14/2026 10:30 AM", "03/05/2026 11:00 AM",
                   "01/28/2026 02:00 PM",  "02/20/2026 03:30 PM"]
DATETIMES_NEXT  = ["04/15/2026 09:00 AM", "05/10/2026 10:30 AM", "06/02/2026 11:00 AM",
                   "04/30/2026 02:00 PM",  "05/22/2026 03:30 PM"]

# Location filter string shared by all HCT reports
HCT_LOCATIONS = (
    "Heart Ctr Of N TX Fort Worth, Heart Ctr Of N TX Granbury, Heart Ctr Of N TX Weatherford"
)
DATE_RANGE = "01/01/2025 To 03/31/2026"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def bold_header(ws, row_num: int = 1):
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fnt  = Font(bold=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = fnt
        cell.alignment = Alignment(horizontal="center")


def write_xlsx(path: Path, rows: list[list], header_row: int = 1):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    bold_header(ws, header_row)
    wb.save(path)


def write_csv(path: Path, rows: list[list]):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)


def empty_row(n: int) -> list:
    return [""] * n


# ---------------------------------------------------------------------------
# MCA — patients_by_insurance_sample.xlsx
#
# Real APRIMA "Patients by Insurance" report:
#   Rows 1-6  : report title, blank, filters, blank, blank, blank
#   Rows 7-19 : payer summary block (totals etc.)
#   Row  20   : column header row (13 columns, read by index in cleaner)
#   Row  21+  : data rows (payer-group rows + patient rows)
#
# Column layout (index 0-12 as used by InsuranceDataCleaner):
#   0  Account Name  (payer group header — blank on patient rows)
#   1  Claim Type
#   2  Patient ID
#   3  Chart Name
#   4  Policy / Member ID
#   5  Patient Name
#   6  Address
#   7  Phone Numbers
#   8  Account (e.g. "Medicare (P)")
#   9  Eff Date
#   10 Term Date
#   11 Sex
#   12 Birth Date
# ---------------------------------------------------------------------------
def gen_mca_patients_by_insurance():
    path = BASE / "MCA" / "patients_by_insurance_sample.xlsx"

    col_list = ("Account Name, Claim Type, Patient ID, Chart Name, Policy/Member ID, "
                "Patient Name, Address, Phone Numbers, Account, Eff Date, Term Date, Sex, Birth Date")
    filter_row = (
        f"FILTERS - [Account]: All Accounts, "
        f"[Filter 1]: (Effective Date): {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Patients by Insurance"
    )

    headers = [
        "Account Name", "Claim Type", "Patient ID", "Chart Name", "Policy/Member ID",
        "Patient Name", "Address", "Phone Numbers", "Account",
        "Eff Date", "Term Date", "Sex", "Birth Date"
    ]

    rows = [[filter_row] + empty_row(12)]  # row 1
    for _ in range(2, 20):                 # rows 2-19 — blank metadata block
        rows.append(empty_row(13))
    rows.append(headers)                   # row 20 — column headers

    # Data — payer group row then 5 patient rows
    rows.append(["Medicare"] + empty_row(12))
    for i in range(5):
        phone_combined = f"{PHONES_H[i]}\n{PHONES_C[i]}"
        full_addr = f"{STREETS[i]}, {CITIES[i]}, {STATES[i]} {ZIPS[i]}"
        account_str = f"{PRIM_INS[i]} (P)"
        rows.append([
            "",                                       # 0  Account Name (blank on patient rows)
            "P",                                      # 1  Claim Type
            PATIENT_IDS[i],                           # 2  Patient ID
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]}",     # 3  Chart Name
            PRIM_IDS[i],                              # 4  Policy/Member ID
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]} {MID_INI[i]}",  # 5  Patient Name
            full_addr,                                # 6  Address
            phone_combined,                           # 7  Phone Numbers
            account_str,                              # 8  Account
            "01/01/2025",                             # 9  Eff Date
            "",                                       # 10 Term Date
            SEXES[i],                                 # 11 Sex
            DOBS[i],                                  # 12 Birth Date
        ])

    write_xlsx(path, rows, header_row=20)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# MCA — patients_with_visits_sample.xlsx
#
# Real APRIMA "Patients With Visits By Insurance" report:
#   Rows 1-13 : report header / filter block
#   Row  14   : column headers
#   Row  15+  : payer-group rows + patient rows
#
# Column layout (index 0-7 as used by VisitsDataCleaner):
#   0  Payer Name       (group header row)
#   1  Plan Type
#   2  Member ID
#   3  Patient (format "ID/LastName, FirstName")
#   4  Address
#   5  Phone (format "(xxx) xxx-xxxx cell" or "home")
#   6  Last Visit
#   7  Visit Count
# ---------------------------------------------------------------------------
def gen_mca_patients_with_visits():
    path = BASE / "MCA" / "patients_with_visits_sample.xlsx"

    col_list = "Payer Name, Plan Type, Member ID, Patient, Address, Phone, Last Visit, Visit Count"
    filter_row = (
        f"FILTERS - [Visit Dates]: {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Patients With Visits By Insurance"
    )

    headers = ["Payer Name", "Plan Type", "Member ID", "Patient",
               "Address", "Phone", "Last Visit", "Visit Count"]

    rows = [[filter_row] + empty_row(7)]
    for _ in range(2, 14):
        rows.append(empty_row(8))
    rows.append(headers)  # row 14

    rows.append(["Medicare"] + empty_row(7))
    for i in range(5):
        pid_name   = f"{PATIENT_IDS[i]}/{LAST_NAMES[i]}, {FIRST_NAMES[i]}"
        full_addr  = f"{STREETS[i]}, {CITIES[i]}, {STATES[i]} {ZIPS[i]}"
        phone_type = f"{PHONES_C[i]} cell"
        rows.append([
            "",            # 0  Payer Name (blank on patient rows)
            "Medicare",    # 1  Plan Type
            PRIM_IDS[i],   # 2  Member ID
            pid_name,      # 3  Patient
            full_addr,     # 4  Address
            phone_type,    # 5  Phone
            DATES_LAST[i], # 6  Last Visit
            str(i + 3),    # 7  Visit Count
        ])

    write_xlsx(path, rows, header_row=14)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# MCA — patients_by_diagnosis_sample.xlsx
#
# Real APRIMA "Patient by Diagnosis" report — multi-row per patient:
#   Rows 1-14 : report header / filter block
#   Row  15   : column header row
#   Row  16+  : patient header row, then sub-rows for diagnoses and meds
#
# Column layout (index 0-6 as used by PatientsDataCleaner):
#   0  Patient (format "ID/LastName, FirstName")   — OR blank on sub-rows
#   1  MRN / "Code - Description" / medication text — sub-row data
#   2  Address
#   3  Phone
#   4  DOB
#   5  Age
#   6  Sex
# ---------------------------------------------------------------------------
def gen_mca_patients_by_diagnosis():
    path = BASE / "MCA" / "patients_by_diagnosis_sample.xlsx"

    col_list = "Patient, MRN, Address, Phone, DOB, Age, Sex"
    filter_row = (
        f"FILTERS - [Visit Span]: {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Patients by Diagnosis"
    )

    headers = ["Patient", "MRN / Diagnosis / Medication", "Address", "Phone", "DOB", "Age", "Sex"]

    rows = [[filter_row] + empty_row(6)]
    for _ in range(2, 15):
        rows.append(empty_row(7))
    rows.append(headers)  # row 15

    for i in range(5):
        pid_name  = f"{PATIENT_IDS[i]}/{LAST_NAMES[i]}, {FIRST_NAMES[i]}"
        full_addr = f"{STREETS[i]}, {CITIES[i]}, {STATES[i]} {ZIPS[i]}"
        icd1, icd2 = DIAG_CODES[i]
        dsc1, dsc2 = DIAG_DESCS[i]
        # Patient header row
        rows.append([pid_name, PATIENT_IDS[i], full_addr, PHONES_H[i], DOBS[i], str(AGES[i]), SEXES[i]])
        # Diagnosis sub-rows
        rows.append(["", f"{icd1}-{dsc1}", "", "", "", "", ""])
        rows.append(["", f"{icd2}-{dsc2}", "", "", "", "", ""])
        # Medication sub-row  (format: "MedName; [StartDate - EndDate]")
        rows.append(["", f"{MEDS[i]}; [01/01/2025 - ]", "", "", "", "", ""])

    write_xlsx(path, rows, header_row=15)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# MCA — patient_list_sample.xlsx  (Patient Center export)
#
# This is a simple Excel export from Manage Patients > Patient Center.
# Row 1 = column headers (no filter metadata).
# 25+ columns — cleaner reads specific indices:
#   [2]=Patient ID, [3]=Patient Name, [7]=DOB,
#   [14]=Addr1, [15]=Addr2, [16]=City, [17]=State, [18]=ZIP, [20]=Notes
# ---------------------------------------------------------------------------
def gen_mca_patient_list():
    path = BASE / "MCA" / "patient_list_sample.xlsx"

    # 25-column layout matching the indices the cleaner reads
    headers = [
        "Status",       "Category",           # 0, 1
        "Patient ID",   "Patient Name",        # 2, 3
        "Phone 1",      "Phone 2",   "Email",  # 4, 5, 6
        "Birth Date",                          # 7
        "Insurance",    "Ins ID",    "Ins Grp", "PCP", "Provider", "Sex",  # 8-13
        "Address 1",    "Address 2",           # 14, 15
        "City",         "State",     "ZIP",    # 16, 17, 18
        "Email 2",                             # 19
        "Notes",                               # 20
        "Language",     "Race",      "Status2", "MRN"  # 21-24
    ]

    rows = [headers]
    for i in range(5):
        emerg = f"ER {EMERG_NAMES[i]} ({EMERG_RELS[i]}) {EMERG_PHONES[i]}"
        rows.append([
            "Active", "Regular",
            PATIENT_IDS[i], f"{LAST_NAMES[i]}, {FIRST_NAMES[i]}",
            PHONES_H[i], PHONES_C[i], EMAILS[i],
            DOBS[i],
            PRIM_INS[i], PRIM_IDS[i], PRIM_GRP[i],
            PROVIDERS[i], PROVIDERS[i], SEXES[i],
            STREETS[i], "",
            CITIES[i], STATES[i], ZIPS[i],
            EMAILS[i],
            emerg,
            LANGS[i], RACES[i], "Active", PATIENT_IDS[i],
        ])

    write_xlsx(path, rows, header_row=1)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# MCA — appointment_report_sample.xlsx
#
# APRIMA "Appointment Report" — scheduled appointments:
#   Rows 1-12 : filter / metadata block
#   Row  13   : column headers
#   Row  14+  : appointment rows + "Reason: ..." sub-rows (skipped by cleaner)
#
# Column layout (index 0-2):
#   0  Appointment DateTime (MM/DD/YYYY HH:MM AM/PM)
#   1  Patient / Phone (format "LastName, FirstName/(xxx) xxx-xxxx")
#   2  Date of Birth
# ---------------------------------------------------------------------------
def gen_mca_appointment_report():
    path = BASE / "MCA" / "appointment_report_sample.xlsx"

    col_list = "Appointment DateTime, Patient / Phone, Date of Birth"
    filter_row = (
        f"FILTERS - [Date Range]: {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Appointment Report"
    )

    headers = ["Appointment DateTime", "Patient / Phone", "Date of Birth"]

    rows = [[filter_row] + empty_row(2)]
    for _ in range(2, 13):
        rows.append(empty_row(3))
    rows.append(headers)  # row 13

    for i in range(5):
        pat_phone = f"{LAST_NAMES[i]}, {FIRST_NAMES[i]}/{PHONES_C[i]}"
        rows.append([DATETIMES_NEXT[i], pat_phone, DOBS[i]])
        rows.append(["Reason: Follow-up cardiology visit", "", ""])

    write_xlsx(path, rows, header_row=13)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# MCA — copay_report_sample.xlsx
#
# APRIMA "Copay Report":
#   Rows 1-12 : filter / metadata block
#   Row  13   : column headers (10 columns)
#   Row  14+  : data rows
#
# Column layout (index 0-9 as used by CopayDataCleaner):
#   0  Date    1  Patient Name    2  DOB    3  ID
#   4  Phone   5  Provider        6  Primary Insurance
#   7  Secondary Insurance        8  Group  9  Actual Copay
# ---------------------------------------------------------------------------
def gen_mca_copay_report():
    path = BASE / "MCA" / "copay_report_sample.xlsx"

    col_list = "Date, Patient Name, DOB, Patient ID, Phone, Provider, Primary Insurance, Secondary Insurance, Group, Actual Copay"
    filter_row = (
        f"FILTERS - [Date Range]: {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Copay Report"
    )

    headers = ["Date", "Patient Name", "DOB", "Patient ID", "Phone",
               "Provider", "Primary Insurance", "Secondary Insurance", "Group", "Actual Copay"]

    rows = [[filter_row] + empty_row(9)]
    for _ in range(2, 13):
        rows.append(empty_row(10))
    rows.append(headers)  # row 13

    for i in range(5):
        rows.append([
            DATES_LAST[i],
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]}",
            DOBS[i],
            PATIENT_IDS[i],
            PHONES_H[i],
            PROVIDERS[i],
            PRIM_INS[i],
            SEC_INS[i] if SEC_INS[i] else "",
            PRIM_GRP[i],
            COPAYS[i],
        ])

    write_xlsx(path, rows, header_row=13)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# MCA — services_by_provider_sample.xlsx  (optional input)
#
# APRIMA "Services By Provider Summary":
#   Standard header at row 1.
#   Columns: Patient, Birthdate, Billing Provider  (plus extras)
# ---------------------------------------------------------------------------
def gen_mca_services_by_provider():
    path = BASE / "MCA" / "services_by_provider_sample.xlsx"

    col_list = "Patient, Birthdate, Billing Provider, Service Date, CPT Code, Description"
    filter_row = (
        f"FILTERS - [Timespan]: {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Services By Provider Summary"
    )

    headers = ["Patient", "Birthdate", "Billing Provider", "Service Date", "CPT Code", "Description"]

    rows = [[filter_row] + empty_row(5)]
    for _ in range(2, 5):
        rows.append(empty_row(6))
    rows.append(headers)  # row 5 (header=4 not needed here — cleaner dynamic-finds headers)

    for i in range(5):
        rows.append([
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]}",
            DOBS[i],
            PROVIDERS[i],
            DATES_LAST[i],
            "99213",
            "Office/outpatient visit, established patient",
        ])

    write_xlsx(path, rows, header_row=5)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# HCT — patient-demographics_sample.xlsx
#
# [Report Name]: HBOX-Demographics-New
# Rows 1-4 : metadata (row 1 = FILTERS cell)
# Row 5    : column headers
# Row 6+   : data
# ---------------------------------------------------------------------------
def gen_hct_demographics():
    path = BASE / "HCT" / "patient-demographics_sample.xlsx"

    col_list = ("Per Nbr, Pat Name, Birth Dt, Hm Phone, Cell Phone, Day Phone, Email Addr, "
                "Preferred Language, Race, Lst Enc Loc Name, Lst Enc Dt, Nxt Appt Dt, "
                "Rendering, Prim Care Phys, Addr 2, Addr 1, City, State, Zip, Sex")
    # NextGen export metadata format
    filter_row = (
        f"FILTERS - [Locations]: {HCT_LOCATIONS}, "
        f"[Filter 1]: (Service Date): {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: HBOX-Demographics-New"
    )

    headers = [
        "Per Nbr", "Pat Name", "Birth Dt", "Hm Phone", "Cell Phone",
        "Day Phone", "Email Addr", "Preferred Language", "Race",
        "Lst Enc Loc Name", "Lst Enc Dt", "Nxt Appt Dt",
        "Rendering", "Prim Care Phys",
        "Addr 2", "Addr 1", "City", "State", "Zip", "Sex"
    ]

    n = len(headers)
    rows = [[filter_row] + empty_row(n - 1)]  # row 1
    for _ in range(2, 5):                      # rows 2-4 empty
        rows.append(empty_row(n))
    rows.append(headers)                       # row 5

    for i in range(5):
        rows.append([
            PATIENT_IDS[i],
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]} {MID_INI[i]}",
            DOBS[i],
            PHONES_H[i], PHONES_C[i], "",
            EMAILS[i], LANGS[i], RACES[i],
            FACILITIES[i], DATES_LAST[i], DATES_NEXT[i],
            PROVIDERS[i], PROVIDERS[i],
            "", STREETS[i], CITIES[i], STATES[i], ZIPS[i], SEXES[i],
        ])

    write_xlsx(path, rows, header_row=5)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# HCT — patient-insurance_sample.xlsx
#
# [Report Name]: Patient Insurance Report
# Rows 1-4 : metadata
# Row 5    : column headers
# Row 6+   : data (one row per COB level per patient)
#
# Columns: Per Nbr, Enc Cob, Payer Name, Pol Nbr, Group Name, Ins Type, Co Amt
# ---------------------------------------------------------------------------
def gen_hct_insurance():
    path = BASE / "HCT" / "patient-insurance_sample.xlsx"

    col_list = "Per Nbr, Enc Cob, Payer Name, Pol Nbr, Group Name, Ins Type, Co Amt"
    filter_row = (
        f"FILTERS - [Locations]: {HCT_LOCATIONS}, "
        f"[Filter 1]: (Service Date): {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: Patient Insurance Report"
    )

    headers = ["Per Nbr", "Enc Cob", "Payer Name", "Pol Nbr", "Group Name", "Ins Type", "Co Amt"]

    n = len(headers)
    rows = [[filter_row] + empty_row(n - 1)]
    for _ in range(2, 5):
        rows.append(empty_row(n))
    rows.append(headers)  # row 5

    for i in range(5):
        # Primary (COB 1)
        rows.append([PATIENT_IDS[i], "1", PRIM_INS[i], PRIM_IDS[i],
                     PRIM_GRP[i], "Medicare", COPAYS[i]])
        # Secondary (COB 2) — only if patient has one
        if SEC_INS[i]:
            rows.append([PATIENT_IDS[i], "2", SEC_INS[i], SEC_IDS[i],
                         SEC_GRP[i], "Medicare Advantage", 0.00])

    write_xlsx(path, rows, header_row=5)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# HCT — HCT ICD Codes sample
#
# This is the exact format shown by the user.
# [Report Name]: EHR Manual Charge Processing History
# Rows 1-4 : metadata (row 1 = full FILTERS string)
# Row 5    : ALL column headers (Prac Name, Loc Name, Encounter, Dt of Svc,
#            Pat Name, Per Nbr, Diag 1-12, Diag 1-12 Desc, Diag 1-12 Type,
#            Rendering, Diag 1-12 Risk Ind, Diag1-12 EOC)
# Row 6+   : data rows (one encounter row per patient)
# ---------------------------------------------------------------------------
def gen_hct_icd_codes():
    path = BASE / "HCT" / "HCT_Location_Wise_Provider_Wise_Patient_Wise_ICD_Codes_sample.xlsx"

    diag_cols  = ", ".join(f"Diag {j}" for j in range(1, 13))
    desc_cols  = ", ".join(f"Diag {j} Desc" for j in range(1, 13))
    type_cols  = ", ".join(f"Diag {j} Type" for j in range(1, 13))
    risk_cols  = ", ".join(f"Diag {j} Risk Ind" for j in range(1, 13))
    eoc_cols   = ", ".join(f"Diag{j} EOC " for j in range(1, 13))

    col_list = (
        f"Prac Name, Loc Name, Encounter, Dt of Svc, Pat Name, Per Nbr, "
        f"{diag_cols}, {desc_cols}, {type_cols}, Rendering, "
        f"{risk_cols}, {eoc_cols}"
    )
    filter_row = (
        f"FILTERS - [Locations]: {HCT_LOCATIONS}, "
        f"[Filter 1]: (Service Date): {DATE_RANGE}, "
        f"[Columns ]: {col_list}, "
        f"[Report Name]: EHR Manual Charge Processing History"
    )

    # Build the exact header row matching the [Columns] list
    headers = ["Prac Name", "Loc Name", "Encounter", "Dt of Svc", "Pat Name", "Per Nbr"]
    headers += [f"Diag {j}"        for j in range(1, 13)]
    headers += [f"Diag {j} Desc"   for j in range(1, 13)]
    headers += [f"Diag {j} Type"   for j in range(1, 13)]
    headers += ["Rendering"]
    headers += [f"Diag {j} Risk Ind" for j in range(1, 13)]
    headers += [f"Diag{j} EOC "    for j in range(1, 13)]  # note trailing space — matches real export

    n = len(headers)
    rows = [[filter_row] + empty_row(n - 1)]  # row 1
    for _ in range(2, 5):                      # rows 2-4 empty
        rows.append(empty_row(n))
    rows.append(headers)                       # row 5

    for i in range(5):
        icd1, icd2 = DIAG_CODES[i]
        dsc1, dsc2 = DIAG_DESCS[i]
        row = [
            "Heart Ctr Of N TX",          # Prac Name
            FACILITIES[i],                # Loc Name
            f"ENC{10000 + i}",            # Encounter
            DATES_LAST[i],                # Dt of Svc
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]} {MID_INI[i]}",  # Pat Name
            PATIENT_IDS[i],               # Per Nbr
        ]
        # Diag 1-12 (fill 1 and 2, rest blank)
        row += [icd1, icd2] + [""] * 10
        # Diag 1-12 Desc
        row += [dsc1, dsc2] + [""] * 10
        # Diag 1-12 Type
        row += ["W", "W"] + [""] * 10     # W = Working diagnosis
        # Rendering
        row += [PROVIDERS[i]]
        # Diag 1-12 Risk Ind
        row += ["Y", "Y"] + [""] * 10
        # Diag1-12 EOC
        row += ["1", "2"] + [""] * 10
        rows.append(row)

    write_xlsx(path, rows, header_row=5)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# SSC — three CSV files
#
# All three have skiprows=1 in pandas, meaning:
#   Row 1 : "REPORT NAME: <report name>" metadata
#   Row 2 : column headers
#   Row 3+: data
# ---------------------------------------------------------------------------
def gen_ssc_patient_details():
    # Athena Health CSV export — first row is report name, row 2 is headers
    path = BASE / "SSC" / "Chronic_Management_Patient_Details_sample.csv"
    headers = [
        "patientid", "patient name", "patientdob", "patientsex",
        "patient address2", "patient address1", "patient city", "patient state",
        "patient zip", "patient homephone", "patient mobile no", "patient workphone",
        "patient email", "patient lang", "race",
        "ptnt emrgncy cntct name", "ptnt emrgncy cntct rltnshp", "ptnt emrgncy cntct ph",
        "patient primary ins pkg name", "patient primary policyidnumber",
        "patient primary policygrpnu...",
        "patient secondary ins pkg name", "patient secondary policyidn...",
        "patient secondary policygrp...",
        "patient tertiary ins pkg name",  "patient tertiary policyidnu...",
        "patient tertiary policygrpn...",
        "patient insexpctcopay",
        "patientlastseend", "patientnextappt",
        "prim prvdrfullnme", "reg dprtmnt", "patientinspcprvdr",
    ]
    rows = [
        ["REPORT NAME: Chronic Management Patient Details"],
        headers,
    ]
    for i in range(5):
        rows.append([
            PATIENT_IDS[i],
            f"{FIRST_NAMES[i]} {MID_INI[i]} {LAST_NAMES[i]}",
            DOBS[i], GENDERS[i],
            "",  STREETS[i], CITIES[i], STATES[i], ZIPS[i],
            PHONES_H[i], PHONES_C[i], "", EMAILS[i],
            LANGS[i], RACES[i],
            EMERG_NAMES[i], EMERG_RELS[i], EMERG_PHONES[i],
            PRIM_INS[i],  PRIM_IDS[i],  PRIM_GRP[i],
            SEC_INS[i],   SEC_IDS[i],   SEC_GRP[i],
            "",           "",            "",
            COPAYS[i],
            DATES_LAST[i], DATES_NEXT[i],
            PROVIDERS[i], FACILITIES[i], PROVIDERS[i],
        ])
    write_csv(path, rows)
    print(f"  {path.name}")


def gen_ssc_diagnosis():
    path = BASE / "SSC" / "Patient_Diagnosis_Code_sample.csv"
    rows = [
        ["REPORT NAME: Patient Diagnosis Code"],
        ["patientid", "icd10encounterdiagcode", "icd10encounterdiagdescr"],
    ]
    for i in range(5):
        icd1, icd2 = DIAG_CODES[i]
        dsc1, dsc2 = DIAG_DESCS[i]
        rows.append([PATIENT_IDS[i], icd1, dsc1])
        rows.append([PATIENT_IDS[i], icd2, dsc2])
    write_csv(path, rows)
    print(f"  {path.name}")


def gen_ssc_medication():
    path = BASE / "SSC" / "Patient_Medication_sample.csv"
    rows = [
        ["REPORT NAME: Patient Medication"],
        ["Patient chart Id", "med names", "encounter_provider"],
    ]
    for i in range(5):
        rows.append([PATIENT_IDS[i], MEDS[i], PROVIDERS[i]])
    write_csv(path, rows)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# CAM — data_new_sample.xlsx
#
# CGM APRIMA "Registry/Problem List" export — standard header at row 1.
# Column names come from row.get(...) calls in CAM main.py.
# ---------------------------------------------------------------------------
def gen_cam_data_new():
    # Epic (Henry Ford) export — standard header at row 1, no filter metadata row
    path = BASE / "CAM" / "data_new_sample.xlsx"
    headers = [
        "MRN", "Patient", "DOB", "Sex", "Race",
        "Mailing Address", "City", "State",
        "Phone", "Mobile #", "E-mail", "Lang",
        "Primary Emer Cont Name", "Primary Emer Cont Rel", "Primary Emer Cont Home Phone",
        "Medicare Sub ID", "Primary Cvg", "Primary Mem ID",
        "Secondary Payer", "Secondary Mem ID",
        "Tertiary Payer", "Tertiary Mem ID",
        "Registries", "Problem List",
        "Copay Due", "Encounter Provider", "Dept/Loc",
        "PCP", "Current Medications",
        "Last Visit Date", "Next Appt Date",
    ]
    rows = [headers]
    for i in range(5):
        full_addr = f"{STREETS[i]}, {CITIES[i]}, {STATES[i]} {ZIPS[i]}"
        rows.append([
            PATIENT_IDS[i],
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]} {MID_INI[i]}",
            DOBS[i], SEXES[i], RACES[i],
            full_addr, CITIES[i], STATES[i],
            PHONES_H[i], PHONES_C[i], EMAILS[i], LANGS[i],
            EMERG_NAMES[i], EMERG_RELS[i], EMERG_PHONES[i],
            MEDICARE_IDS[i], PRIM_INS[i], PRIM_IDS[i],
            SEC_INS[i],  SEC_IDS[i],
            "",          "",
            REGISTRIES[i], PROBLEM_LISTS[i],
            COPAYS[i], PROVIDERS[i], FACILITIES[i],
            PROVIDERS[i], MEDS[i],
            DATES_LAST[i], DATES_NEXT[i],
        ])
    write_xlsx(path, rows, header_row=1)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# CIM — final_hbox_sample.xlsx
#
# Intensive care management export — standard header at row 1.
# Column names from row.get(...) calls in CIM main.py.
# ---------------------------------------------------------------------------
def gen_cim_final_hbox():
    # Epic (Henry Ford) export — standard header at row 1, no filter metadata row
    path = BASE / "CIM" / "final_hbox_sample.xlsx"
    headers = [
        "MRN", "Patient", "DOB", "Sex", "Race",
        "Street Address", "Street Address.1", "Pt City", "Patient State", "ZIP Code",
        "Patient Home Phone", "Patient Cell Phone", "Pt. E-mail Address", "Language",
        "Primary Emer Cont Name", "Primary Emer Cont Rel", "Emerg Contact Ph",
        "Medicare Sub ID", "Primary Payer", "Primary Mem ID", "Pat Primary CVG Payer ID",
        "Secondary Payer", "Secondary Mem ID",
        "Tertiary Payer", "Tertiary Mem ID",
        "Registries", "Problem List",
        "Copay Due", "Encounter Provider", "Dept/Loc",
        "PCP", "Current Medications",
        "Last Visit Date", "Next Appt",
    ]
    rows = [headers]
    for i in range(5):
        rows.append([
            PATIENT_IDS[i],
            f"{LAST_NAMES[i]}, {FIRST_NAMES[i]} {MID_INI[i]}",
            DOBS[i], SEXES[i], RACES[i],
            STREETS[i], "", CITIES[i], STATES[i], ZIPS[i],
            PHONES_H[i], PHONES_C[i], EMAILS[i], LANGS[i],
            EMERG_NAMES[i], EMERG_RELS[i], EMERG_PHONES[i],
            MEDICARE_IDS[i], PRIM_INS[i], PRIM_IDS[i], PRIM_GRP[i],
            SEC_INS[i],  SEC_IDS[i],
            "",          "",
            REGISTRIES[i], PROBLEM_LISTS[i],
            COPAYS[i], PROVIDERS[i], FACILITIES[i],
            PROVIDERS[i], MEDS[i],
            DATES_LAST[i], DATES_NEXT[i],
        ])
    write_xlsx(path, rows, header_row=1)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
# XHI — three CSV files (no report-name metadata row — standard CSV)
# ---------------------------------------------------------------------------
def gen_xhi_emr():
    # DrChrono export — standard CSV, no report metadata row
    path = BASE / "XHI" / "EMR_Final_Report_sample.csv"
    headers = [
        "Chart ID", "Full Name", "Date of Birth", "Gender",
        "Address", "City", "State", "Zip Code",
        "Home Phone", "Cell Phone", "Office Phone", "Email", "Race",
        "Emerg Contact Name", "Emerg Contact Relation", "Emerg Contact Phone",
        "Practice Official Name",
        "Date of Last Appointment", "Date of Next Appointment",
        "Provider", "Primary Provider", "Primary Care Physician",
        "Primary Ins Payer", "Primary Member ID", "Primary Ins Group #",
        "Secondary Ins Payer", "Secondary Member ID", "Secondary Ins Group #",
        "Expected Copay", "Appointment Notes",
    ]
    rows = [headers]
    for i in range(5):
        rows.append([
            PATIENT_IDS[i],
            f"{FIRST_NAMES[i]} {MID_INI[i]} {LAST_NAMES[i]}",
            DOBS_ISO[i], GENDERS[i],
            STREETS[i], CITIES[i], STATES[i], ZIPS[i],
            PHONES_H[i], PHONES_C[i], "", EMAILS[i], RACES[i],
            EMERG_NAMES[i], EMERG_RELS[i], EMERG_PHONES[i],
            "Midwest Cardiology, PC",
            DATES_LAST[i], DATES_NEXT[i],
            PROVIDERS[i], PROVIDERS[i], PROVIDERS[i],
            PRIM_INS[i],  PRIM_IDS[i],  PRIM_GRP[i],
            SEC_INS[i],   SEC_IDS[i],   SEC_GRP[i],
            COPAYS[i], "Follow-up cardiology visit",
        ])
    write_csv(path, rows)
    print(f"  {path.name}")


def gen_xhi_medication():
    path = BASE / "XHI" / "medication_report_sample.csv"
    rows = [["Chart ID", "Medication", "Status", "Prescribed Datetime", "Start Taking Datetime"]]
    for i in range(5):
        rows.append([PATIENT_IDS[i], MEDS[i], "Active",
                     DATETIMES_LAST[i], DATETIMES_LAST[i]])
        rows.append([PATIENT_IDS[i], MEDS[(i + 1) % 5], "Inactive",
                     "06/01/2025 09:00 AM", "06/01/2025 09:00 AM"])
    write_csv(path, rows)
    print(f"  {path.name}")


def gen_xhi_problem():
    path = BASE / "XHI" / "problem_report_sample.csv"
    rows = [["Chart ID", "Problem", "ICD10 Code"]]
    for i in range(5):
        icd1, icd2 = DIAG_CODES[i]
        dsc1, dsc2 = DIAG_DESCS[i]
        rows.append([PATIENT_IDS[i], dsc1, icd1])
        rows.append([PATIENT_IDS[i], dsc2, icd2])
    write_csv(path, rows)
    print(f"  {path.name}")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("MCA")
    gen_mca_patients_by_insurance()
    gen_mca_patients_with_visits()
    gen_mca_patients_by_diagnosis()
    gen_mca_patient_list()
    gen_mca_appointment_report()
    gen_mca_copay_report()
    gen_mca_services_by_provider()

    print("HCT")
    gen_hct_demographics()
    gen_hct_insurance()
    gen_hct_icd_codes()

    print("SSC")
    gen_ssc_patient_details()
    gen_ssc_diagnosis()
    gen_ssc_medication()

    print("CAM")
    gen_cam_data_new()

    print("CIM")
    gen_cim_final_hbox()

    print("XHI")
    gen_xhi_emr()
    gen_xhi_medication()
    gen_xhi_problem()

    print("\nDone — 19 sample files in src/samples/")
