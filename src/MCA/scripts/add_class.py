# Add DemographicScheduleCleaner class to data_cleaner.py

new_class = '''
class DemographicScheduleCleaner:
    """Class for cleaning demographic data from multi-sheet Excel workbooks."""

    def __init__(self, input_path: str, output_path: str):
        self.input_path = input_path
        self.output_path = output_path
        self.text_cleaner = TextCleaner()

    def clean_data(self) -> int:
        """Clean demographic data from all sheets and return number of records processed."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel file reading. Install with: pip install openpyxl")

        # Load the workbook
        workbook = openpyxl.load_workbook(self.input_path, data_only=True)
        records = []

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Skip if sheet is empty
            if sheet.max_row < 10:
                continue

            try:
                record = self._extract_patient_data(sheet)
                if record:
                    records.append(record)
            except Exception as e:
                print(f"Warning: Error processing sheet {sheet_name}: {e}")
                continue

        # Write to CSV
        self._write_csv(records)
        return len(records)

    def _extract_patient_data(self, sheet) -> Optional[Dict]:
        """Extract patient data from a single sheet."""
        record = {}

        try:
            # Patient Name (Row 3, Column 1 - 1-indexed)
            patient_name = str(sheet.cell(row=3, column=1).value or "").strip()
            if not patient_name or patient_name.lower() in ["nan", "none", ""]:
                return None
            record["patient_name"] = self.text_cleaner.clean_text(patient_name)

            # External ID (Row 3, Column 6)
            external_id = str(sheet.cell(row=3, column=6).value or "").strip()
            record["external_id"] = external_id

            # DOB (Row 7, Column 4)
            dob = str(sheet.cell(row=7, column=4).value or "").strip()
            record["dob"] = dob

            # Phone # - combine cell and home (Row 8, Columns 4 and 12)
            cell_phone = str(sheet.cell(row=8, column=4).value or "").strip()
            home_phone = str(sheet.cell(row=8, column=12).value or "").strip()
            # Clean phone numbers (remove prefixes like "cell:" and "home:")
            cell_clean = cell_phone.replace("cell:", "").replace("home:", "").strip()
            home_clean = home_phone.replace("cell:", "").replace("home:", "").strip()
            combined_phone = f"{cell_clean}; {home_clean}".strip("; ")
            record["phone"] = combined_phone

            # Email (Row 10, Column 4)
            email = str(sheet.cell(row=10, column=4).value or "").strip()
            record["email"] = email

            # Language (Row 11, Column 4)
            language = str(sheet.cell(row=11, column=4).value or "").strip()
            record["language"] = language

            # Race (Row 12, Column 4)
            race = str(sheet.cell(row=12, column=4).value or "").strip()
            record["race"] = race

            # Sex (Row 7, Column 17)
            sex = str(sheet.cell(row=7, column=17).value or "").strip()
            record["sex"] = sex

            # Primary Address (Row 14-15, Column 4)
            addr_line1 = str(sheet.cell(row=14, column=4).value or "").strip()
            addr_line2 = str(sheet.cell(row=15, column=4).value or "").strip()
            primary_address = f"{addr_line1} {addr_line2}".strip()
            record["primary_address"] = primary_address

            # Notes (Row 16, Column 4)
            notes = str(sheet.cell(row=16, column=4).value or "").strip()
            record["notes"] = notes

            # Primary Payer/ Plan (Row 23, Column 4)
            primary_payer = str(sheet.cell(row=23, column=4).value or "").strip()
            # Clean up the payer info (remove extra formatting)
            primary_payer = primary_payer.replace("_x000d_", " ").replace("\\n", " ")
            record["primary_payer_plan"] = primary_payer

            # Sec Payer/ Plan (Row 23, Column 19)
            sec_payer = str(sheet.cell(row=23, column=19).value or "").strip()
            sec_payer = sec_payer.replace("_x000d_", " ").replace("\\n", " ")
            record["sec_payer_plan"] = sec_payer

            # Primary Ins Copay - sum of Primary, Specialist, and Other (Rows 25-27, Column 4)
            primary_copay = self._extract_copay_amount(str(sheet.cell(row=25, column=4).value or ""))
            specialist_copay = self._extract_copay_amount(str(sheet.cell(row=26, column=4).value or ""))
            other_copay = self._extract_copay_amount(str(sheet.cell(row=27, column=4).value or ""))

            total_copay = primary_copay + specialist_copay + other_copay
            record["primary_ins_copay"] = f"{total_copay:.2f}"

            # Medication Name - collect all medications starting from Row 36, Column 1
            medications = []
            row = 36  # Start from row 36 (1-indexed)
            while row <= sheet.max_row:
                med_name = str(sheet.cell(row=row, column=1).value or "").strip()
                if med_name and med_name.lower() not in ["nan", "none", ""]:
                    medications.append(med_name)
                row += 2  # Medications seem to be every other row

            record["medication_name"] = "; ".join(medications)

        except Exception as e:
            print(f"Error extracting data: {e}")
            return None

        return record

    def _extract_copay_amount(self, copay_text: str) -> float:
        """Extract numeric amount from copay text like 'Primary: $0.00'."""
        if not copay_text:
            return 0.0

        # Find dollar amount
        import re
        match = re.search(r'\\$?(\\d+\\.?\\d*)', copay_text.replace(',', ''))
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                return 0.0
        return 0.0

    def _write_csv(self, records: List[Dict]) -> None:
        """Write records to CSV file."""
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)

        if not records:
            return

        fieldnames = [
            "patient_name", "external_id", "dob", "phone", "email", "language",
            "race", "sex", "primary_address", "notes", "primary_payer_plan",
            "sec_payer_plan", "primary_ins_copay", "medication_name"
        ]

        with open(self.output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
'''

# Read the current file
with open('d:/Work_Folder/medical_db/src/MCA/scripts/data_cleaner.py', 'r', encoding='utf-8') as f:
    content = f.read()

# Write back the file with the new class
with open('d:/Work_Folder/medical_db/src/MCA/scripts/data_cleaner.py', 'w', encoding='utf-8') as f:
    f.write(content + new_class)

print('DemographicScheduleCleaner class added successfully!')