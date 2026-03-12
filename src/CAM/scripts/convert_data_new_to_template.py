import os
import re
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
TEMPLATE_PATH = os.path.join(ROOT, 'src', 'CIM', 'template', 'consolidated_view-template.xlsx')
DATA_PATH = os.path.join(ROOT, 'src', 'CIM', 'data_new.xlsx')
OUTPUT_DIR = os.path.join(ROOT, 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

ZIP_RE = re.compile(r"(\b\d{5}(?:-\d{4})?\b)")
PHONE_LABEL_RE = re.compile(r"\b(hm|home|wk|work|cell|mobile|m|mb)\b", re.IGNORECASE)

# make `constants` importable
_CIM_ROOT = os.path.join(os.path.dirname(__file__), '..')
if _CIM_ROOT not in sys.path:
    sys.path.insert(0, os.path.abspath(_CIM_ROOT))
try:
    import constants
except Exception:
    constants = None


CAUSE_SEVERITY_PATH = os.path.join('src', 'CIM', 'mappings', 'cause_severity.csv')
DISEASE_CSV = os.path.join(ROOT, 'src', 'CIM', 'disease', 'api_prescriptioncauselist_202603101243.csv')


def clean_insurance_name(s):
    if s is None:
        return ''
    s = str(s).strip()
    if not s:
        return ''
    # remove leading numeric code before first dash
    if '-' in s:
        parts = s.split('-', 1)
        # if leading part contains digits, drop it
        if re.search(r"\d", parts[0]):
            s = parts[1].strip()
    # remove trailing group/id tokens that look like codes
    # remove tokens that include digits (likely codes), but keep words like MEDICARE
    toks = s.split()
    toks = [t for t in toks if not re.search(r"\d", t)]
    s = ' '.join(toks)
    # collapse multiple spaces
    s = re.sub(r"\s{2,}", ' ', s)
    return s.strip(' ,')


def validate_and_pick_email(s):
    if s is None:
        return ''
    txt = str(s)
    # split on common separators
    parts = re.split(r'[;,|\\/\n]+', txt)
    email_re = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    for p in parts:
        p = p.strip()
        m = email_re.search(p)
        if m:
            return m.group(0)
    # fallback: return first token-looking part
    for p in parts:
        p = p.strip()
        if '@' in p:
            return p
    return ''


def format_person_name(name):
    if name is None:
        return ''
    s = str(name).strip()
    if not s:
        return ''
    # if 'Last, First' format
    if ',' in s:
        parts = [p.strip() for p in s.split(',', 1)]
        if len(parts) >= 2:
            last = parts[0]
            first = parts[1].split()[0] if parts[1] else ''
            return (first + ' ' + last).strip()
    toks = s.split()
    if len(toks) >= 2:
        return toks[0] + ' ' + toks[-1]
    return s


def load_severity_map(path):
    sev = {}
    try:
        import pandas as _pd
        if os.path.exists(path):
            df = _pd.read_csv(path)
            for _, r in df.iterrows():
                sev[str(r['cause']).strip().lower()] = float(r.get('severity') or 0)
    except Exception:
        pass
    return sev


def find_problem_matches(problem_text, disease_causes=None):
    matches = []
    if not problem_text:
        return matches
    def normalize(s):
        return re.sub(r'[^a-z0-9 ]+', ' ', s.lower())

    tokens = re.split(r'[;,|/\\\n]+', str(problem_text))
    for t in tokens:
        t = str(t).strip()
        if not t:
            continue
        key = t.lower()
        # direct mapping from constants
        if constants and key in constants.PROBLEM_TO_ICD and constants.PROBLEM_TO_ICD[key].get('cause'):
            c = constants.PROBLEM_TO_ICD[key]['cause']
            icd = constants.PROBLEM_TO_ICD[key].get('icd','')
            matches.append({'cause': c, 'icd': icd, 'token': t})
            continue
        # synonyms
        if constants:
            for syn, cause_name in constants.SYNONYM_TO_CAUSE.items():
                if syn in key:
                    matches.append({'cause': cause_name, 'icd': '', 'token': t})
                    break
        # try matching against disease causes (substring, normalized)
        if disease_causes:
            nk = normalize(key)
            for cause in disease_causes:
                if not cause:
                    continue
                nc = normalize(cause)
                if nc and nc in nk:
                    matches.append({'cause': cause, 'icd': '', 'token': t})
                    break
    return matches

def format_provider_person(s):
    if s is None:
        return ''
    s = str(s).strip()
    if not s:
        return ''
    # remove degree tokens
    s = re.sub(r"\b(MD|DO|PA|NP|PhD|M\.D\.|D\.O\.)\b", "", s, flags=re.IGNORECASE)
    s = s.replace('  ', ' ').strip(' ,')
    # if 'Last, First ...'
    if ',' in s:
        parts = [p.strip() for p in s.split(',', 1) if p.strip()]
        if len(parts) >= 2:
            last = parts[0]
            first_part = parts[1].strip()
            # preserve middle initials (do not drop tokens from the given name)
            return f"{first_part} {last}".strip()
    # otherwise keep the given form (preserves middle initials)
    return s

def split_phone_labels(cell):
    out = {'HOME_PHONE': None, 'WORK_PHONE': None, 'MOBILE_PHONE': None}
    if pd.isna(cell) or not str(cell).strip():
        return out
    parts = re.split(r'[;,|\\/\n]+', str(cell))
    for p in parts:
        p = p.strip()
        if not p:
            continue
        key = None
        lower = p.lower()
        if re.search(r'\b(hm|home)\b', lower):
            key = 'HOME_PHONE'
        elif re.search(r'\b(wk|work|wrk)\b', lower):
            key = 'WORK_PHONE'
        elif re.search(r'\b(cell|mobile|m|mb)\b', lower):
            key = 'MOBILE_PHONE'
        # extract number-like
        m = re.search(r"(\+?\d[\d\-\s\(\)]+\d)", p)
        num = m.group(1).strip() if m else p
        if key:
            out[key] = num
        else:
            # unlabeled heuristics: long number -> mobile, else home
            digits = re.sub(r'\D', '', num)
            if len(digits) >= 10:
                out['MOBILE_PHONE'] = out['MOBILE_PHONE'] or num
            else:
                out['HOME_PHONE'] = out['HOME_PHONE'] or num
    return out

def clean_clinic(cf):
    if cf is None:
        return None
    s = str(cf).strip()
    # remove bracketed numeric ids
    s = re.sub(r"\s*\[\s*\d+\s*\]\s*$", "", s)
    # if there's a hyphen indicating a location suffix, prefer the right-most part
    if '-' in s:
        parts = [p.strip() for p in s.split('-')]
        for part in reversed(parts):
            if part:
                part = re.sub(r'\b(P\.C\.|PC|LLC|INC|CORP)\b', '', part, flags=re.IGNORECASE).strip(' ,')
                return part
    # fallback: remove trailing ' - ...' or comma suffix
    s = re.sub(r"\s*[-,]\s*[^-,\[]+$", "", s).strip()
    return s

def main():
    tpl = pd.read_excel(TEMPLATE_PATH, sheet_name=0, nrows=0, engine='openpyxl')
    tpl_cols = list(tpl.columns)
    data = pd.read_excel(DATA_PATH, sheet_name=0, engine='openpyxl')

    out_rows = []
    severity_map = load_severity_map(CAUSE_SEVERITY_PATH)
    # load disease causes for substring matching
    disease_causes = []
    if os.path.exists(DISEASE_CSV):
        try:
            ddf = pd.read_csv(DISEASE_CSV)
            if 'cause' in ddf.columns:
                disease_causes = [str(x).strip().lower() for x in ddf['cause'].fillna('') if str(x).strip()]
        except Exception:
            disease_causes = []
    for _, r in data.iterrows():
        out = {c: None for c in tpl_cols}
        out['EMR_ID'] = r.get('MRN')
        out['PATIENT_EMR_NAME'] = r.get('Patient')
        first = r.get('Patient First Name')
        last = r.get('Patient Last Name')
        if pd.notna(first) and pd.notna(last) and str(first).strip() and str(last).strip():
            out['FIRST_NAME'] = str(first).strip()
            out['LAST_NAME'] = str(last).strip()
            out['PATIENT_FULL_NAME'] = f"{out['FIRST_NAME']} {out['LAST_NAME']}"
        else:
            # parse Patient field
            p = r.get('Patient')
            if isinstance(p, str) and ',' in p:
                last, rest = [x.strip() for x in p.split(',',1)]
                first_token = rest.split()[0] if rest else ''
                out['FIRST_NAME'] = first_token
                out['LAST_NAME'] = last
                out['PATIENT_FULL_NAME'] = f"{first_token} {last}".strip()
            elif isinstance(p, str):
                toks = p.split()
                if len(toks) >= 2:
                    out['FIRST_NAME'] = toks[0]
                    out['LAST_NAME'] = ' '.join(toks[1:])
                    out['PATIENT_FULL_NAME'] = f"{out['FIRST_NAME']} {out['LAST_NAME']}"

        # Mailing Address and ZIP
        mailing = r.get('Mailing Address')
        if pd.notna(mailing):
            mailing_s = str(mailing).strip()
            z = ZIP_RE.search(mailing_s)
            if z:
                out['ZIP'] = z.group(1)
                street = re.sub(re.escape(z.group(0)) + r"\s*$", "", mailing_s).strip(' ,;')
                out['STREET_ADDRESS'] = street
            else:
                out['STREET_ADDRESS'] = mailing_s
        out['CITY'] = r.get('City')
        out['STATE'] = r.get('State')

        # phones
        ph = split_phone_labels(r.get('Phone'))
        out['HOME_PHONE'] = ph.get('HOME_PHONE')
        out['WORK_PHONE'] = ph.get('WORK_PHONE')
        out['MOBILE_PHONE'] = r.get('Mobile #') or ph.get('MOBILE_PHONE')

        # email: pick first valid email
        out['EMAIL_ADDRESS'] = validate_and_pick_email(r.get('E-mail') or r.get('Pt. E-mail Address'))

        out['LANGUAGE'] = r.get('Lang')
        out['EMERGENCY_CONTACT_NAME'] = format_person_name(r.get('Primary Emer Cont Name'))
        out['EMERGENCY_RELATIONSHIP'] = r.get('Primary Emer Cont Rel')
        out['EMERGENCY_CONTACT_HOME_PHONE'] = r.get('Primary Emer Cont Home Phone')

        # insurances
        # insurance names: use Payor columns (Primary Payer / Secondary Payer / Tertiary Payer)
        out['PRIMARY_INSURANCE'] = r.get('Primary Payer')
        out['PRIMARY_ID'] = r.get('Primary Mem ID') or r.get('Medicare Sub ID')
        raw_sec = r.get('Secondary Payer')
        out['SECONDARY_INSURANCE'] = clean_insurance_name(raw_sec)
        out['SECONDARY_ID'] = r.get('Secondary Mem ID')
        raw_ter = r.get('Tertiary Payer')
        out['TERITARY_INSURANCE'] = clean_insurance_name(raw_ter)
        out['TERITARY_ID'] = r.get('Tertiary Mem ID')

        out['CO-PAY'] = r.get('Copay Due')
        # default disease flags to 'NO'
        disease_flags = [c for c in tpl_cols if c.isupper() and c not in ('EMR_ID','PATIENT_EMR_NAME')]
        for f in disease_flags:
            if f not in out:
                out[f] = 'NO'

        out['LAST_SEEN_DATE'] = r.get('Last Visit Date')
        out['NEXT_APPT_DATE'] = r.get('Next Appt Date')

        # provider from Encounter Provider column preferred, fallback to PCP/Provider
        pcp = r.get('Encounter Provider') or r.get('PCP') or r.get('PCP Name') or r.get('Provider')
        # strip degree tokens from provider data
        pcp_data = ''
        if pcp is not None:
            pcp_data = re.sub(r"\b(MD|DO|PA|NP|PhD|M\.D\.|D\.O\.)\b", "", str(pcp), flags=re.IGNORECASE).strip(' ,')
        out['PROVIDER_DATA'] = pcp_data
        out['PROVIDER_NAME'] = format_provider_person(pcp_data)

        # DATE_OF_BIRTH: accept common column names and coerce to date string
        dob_val = None
        # direct common names
        for key in ('DOB', 'Date of Birth', 'Birth Date'):
            if key in data.columns and pd.notna(r.get(key)) and str(r.get(key)).strip():
                dob_val = r.get(key)
                break
        # fallback: any column containing 'dob'
        if dob_val is None:
            for col in data.columns:
                if isinstance(col, str) and 'dob' in col.lower():
                    v = r.get(col)
                    if pd.notna(v) and str(v).strip():
                        dob_val = v
                        break
        try:
            out['DATE_OF_BIRTH'] = pd.to_datetime(dob_val).date() if dob_val is not None else None
            # GENDER: prefer 'Sex' or 'Gender' columns
            gen = None
            for key in ('Sex', 'sex', 'Gender', 'gender'):
                if key in data.columns and pd.notna(r.get(key)) and str(r.get(key)).strip():
                    gen = r.get(key)
                    break
            out['GENDER'] = gen
        except Exception:
            out['DATE_OF_BIRTH'] = dob_val
        out['CLINIC_FACILITY'] = clean_clinic(r.get('Dept/Loc'))

        # -- Problem List -> medical cause matching (populate PRIMARY/SECONDARY and flags)
        pl_text = r.get('Problem List') or r.get('ProblemList') or r.get('Problems')
        matches = find_problem_matches(pl_text, disease_causes=disease_causes)
        primary = None
        secondary = None
        if matches:
            # prefer explicit labels if present in token text
            for m in matches:
                if 'primary' in str(m.get('token','')).lower():
                    primary = m['cause']
                    break
            if not primary:
                # choose highest severity
                uniq = {}
                for m in matches:
                    c = str(m.get('cause','')).strip()
                    if c and c.lower() not in uniq:
                        uniq[c.lower()] = severity_map.get(c.lower(), 0)
                if uniq:
                    sorted_causes = sorted(uniq.items(), key=lambda x: x[1], reverse=True)
                    primary = sorted_causes[0][0]
                    if len(sorted_causes) > 1:
                        secondary = sorted_causes[1][0]
        out['PRIMARY_DX'] = primary
        out['SECONDARY_DX'] = secondary

        # set boolean disease flags similar to convert_to_consolidated.py mapping
        cause_to_flag = {
            'coronary artery disease': 'CORONARY_ARTERY_DISEASE',
            'arrhythmia': 'ARRHYTHMIA',
            'chf (congestive heart failure)': 'CONGESTIVE_HEART_FAILURE',
            'peripheral vascular disease': 'PERIPHERAL_VASCULAR',
            'valvular heart disease': 'VALVULAR_HEART',
            'cerebrovascular accident': 'CERBOVASCULAR_ACCIDENT',
            'hyperlipidemia': 'HYPERLIPIDEMIA',
            'angina pectoris': 'ANGINA_PECTORIS',
            'hypotension': 'HYPOTENSION',
            'hypertension or pre hypertensive': 'HYPERTENSION',
            'obesity': 'OBESITY',
            'type 2 diabetes': 'DIABETES',
            'chronic kidney': 'CHRONIC_KIDNEY_DISEASE',
            'copd': 'COPD',
            'asthma': 'ASTHMA',
            'sleep apnea': 'SLEEP_APNEA',
            'dyspnea': 'DYSPNEA'
        }
        # include additional pulmonary-related flags
        cause_to_flag.update({
            'emphysema': 'EMPHYSEMA',
            'bronchiectasis': 'BRONCHIECTASIS',
            'hypoxemia': 'HYPOXEMIA',
            'pulmonary hypertension': 'PULMONARY_HYPERTENSION',
            'respiratory failure': 'RESPIRATORY_FAILURE'
        })
        for flag in cause_to_flag.values():
            if flag in out:
                out[flag] = 'NO'
        for m in matches:
            cn = str(m.get('cause','')).strip().lower()
            if cn in cause_to_flag and cause_to_flag[cn] in out:
                out[cause_to_flag[cn]] = 'YES'

        # insurance type classification (use constants if available)
        try:
            if constants:
                out['INSURANCE_TYPE'] = constants.classify_insurance(str(out.get('PRIMARY_INSURANCE') or ''))
            else:
                out['INSURANCE_TYPE'] = ''
        except Exception:
            out['INSURANCE_TYPE'] = ''

        out_rows.append(out)

    out_df = pd.DataFrame(out_rows, columns=tpl_cols)

    # ensure DATE_OF_BIRTH are timestamps (preserve time if present)
    try:
        out_df['DATE_OF_BIRTH'] = pd.to_datetime(out_df['DATE_OF_BIRTH'], errors='coerce')
    except Exception:
        pass

    # read date number format from template sheet (try to find a sample format)
    template_date_format = None
    try:
        wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
        ws = wb.active
        # find header column index for DATE_OF_BIRTH
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        if 'DATE_OF_BIRTH' in header:
            col_idx = header.index('DATE_OF_BIRTH') + 1
            # inspect first few rows for a cell number_format
            for r in range(2, min(10, ws.max_row) + 1):
                cell = ws.cell(row=r, column=col_idx)
                nf = getattr(cell, 'number_format', None)
                if nf and nf not in (None, 'General'):
                    template_date_format = nf
                    break
    except Exception:
        template_date_format = None

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_xlsx = os.path.join(OUTPUT_DIR, f'consolidated_data_new_{ts}.xlsx')

    # normalize NaN and literal 'nan' strings to empty
    try:
        out_df = out_df.where(pd.notnull(out_df), '')
        out_df = out_df.replace(r'^\s*nan\s*$', '', regex=True)
    except Exception:
        pass

    # write Excel using openpyxl engine so we can set cell formats exactly
    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='Sheet1')
        ws = writer.sheets['Sheet1']
        # set DATE_OF_BIRTH column number format to match template if found
        if template_date_format and 'DATE_OF_BIRTH' in out_df.columns:
            col_idx = out_df.columns.get_loc('DATE_OF_BIRTH') + 1
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, len(out_df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell.number_format = template_date_format
                except Exception:
                    pass

    print('Wrote preview and full output to', out_xlsx)
    print('\nPreview (first 10 rows):')
    print(out_df.head(10).to_string(index=False))

if __name__ == '__main__':
    main()
