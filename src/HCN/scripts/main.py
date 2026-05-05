"""
HCN ETL Pipeline — Heart Center of Nevada
Parses PatientReport PDF (Intergy / NextGen export) using PyMuPDF.

Memory-safe strategy:
  1. PyMuPDF streams one page at a time (no full-PDF RAM load).
  2. Parsed records written directly to CSV as we go.
  3. CSV -> Excel at the end using openpyxl write-only mode.
  4. Appointment PDF and Drugs PDF are each extracted to cleaned/ CSV,
     then merged into the consolidated Excel in a second pass.
"""

import re
import sys
import csv
from datetime import datetime, date
from pathlib import Path

import fitz  # PyMuPDF — pip install pymupdf

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
MODULE_DIR  = Path(__file__).resolve().parents[1]
OUTPUT_DIR  = MODULE_DIR / "output"
CLEANED_DIR = MODULE_DIR / "cleaned"
OUTPUT_DIR.mkdir(exist_ok=True)
CLEANED_DIR.mkdir(exist_ok=True)

PDF_PATH = next(
    (f for f in sorted(MODULE_DIR.glob("PatientReport*.PDF"))),
    next((f for f in sorted(MODULE_DIR.glob("PatientReport*.pdf"))), None),
)
APPT_PDF_PATH = next(
    (f for f in sorted(MODULE_DIR.glob("Appt*.PDF"))),
    next((f for f in sorted(MODULE_DIR.glob("Appt*.pdf"))), None),
)
DRUGS_PDF_PATH = next(
    (f for f in sorted(MODULE_DIR.glob("PatientMedication*.PDF"))),
    next((f for f in sorted(MODULE_DIR.glob("PatientMedication*.pdf"))), None),
)
DIAG_PDF_PATH = next(
    (f for f in sorted(MODULE_DIR.glob("Detailed*.PDF"))),
    next((f for f in sorted(MODULE_DIR.glob("Detailed*.pdf"))), None),
)
API_CAUSE_CSV = MODULE_DIR / "template" / "api_prescriptioncauselist_202603101243.csv"

# ---------------------------------------------------------------------------
# Column order (matches template)
# ---------------------------------------------------------------------------
TEMPLATE_COLS = [
    "EMR ID", "PATIENT EMR NAME", "FIRST NAME", "MIDDLE NAME", "LAST NAME",
    "PATIENT FULL NAME", "DATE OF BIRTH", "GENDER", "STREET ADDRESS", "CITY",
    "STATE", "ZIP", "HOME PHONE", "MOBILE PHONE", "WORK PHONE", "EMAIL ADDRESS",
    "LANGUAGE", "RACE", "EMERGENCY CONTACT NAME", "EMERGENCY RELATIONSHIP",
    "EMERGENCY CONTACT HOME PHONE", "EMERGENCY CONTACT MOBILE PHONE",
    "MEDICARE ID", "PRIMARY INSURANCE", "PRIMARY ID", "PRIMARY GROUP",
    "SECONDARY INSURANCE", "SECONDARY ID", "SECONDARY GROUP",
    "TERITARY INSURANCE", "TERITARY ID", "TERITARY GROUP",
    "INSURANCE TYPE", "CO-PAY",
    "CORONARY ARTERY DISEASE", "ARRHYTHMIA", "CONGESTIVE HEART FAILURE",
    "PERIPHERAL VASCULAR", "VALVULAR HEART", "CERBOVASCULAR ACCIDENT",
    "HYPERLIPIDEMIA", "ANGINA PECTORIS", "HYPOTENSION", "HYPERTENSION",
    "OBESITY", "DIABETES", "CHRONIC KIDNEY DISEASE", "COPD",
    "RESPIRATORY FAILURE", "ASTHMA", "SLEEP APNEA", "DYSPNEA",
    "EMPHYSEMA", "BRONCHIECTASIS", "HYPOXEMIA",
    "PRIMARY DX", "SECONDARY DX", "PRIMARY ICD", "SECONDARY ICD",
    "LAST SEEN DATE", "NEXT APPT", "PROVIDER DATA", "PROVIDER NAME",
    "CLINIC FACILITY", "PRIMARY CARE PROVIDER", "MEDICATIONS", "ENCOUNTER NOTES",
]

EMPTY_COMORBIDITIES = {c: "" for c in TEMPLATE_COLS[34:55]}  # cols 35-55

_DATE_COLS = {"DATE OF BIRTH", "LAST SEEN DATE", "NEXT APPT"}


def _to_date(s: str):
    if not s:
        return ''
    try:
        return datetime.strptime(s.strip(), '%m/%d/%Y').date()
    except ValueError:
        return s


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def parse_name(raw: str):
    """'DEMELAS, TIMOTHY W' -> (last, first, middle)"""
    raw = raw.strip()
    if "," in raw:
        last, rest = raw.split(",", 1)
        parts = rest.strip().split()
        return last.strip(), (parts[0] if parts else ""), (parts[1] if len(parts) > 1 else "")
    parts = raw.split()
    return (parts[-1] if parts else raw), (parts[0] if len(parts) > 1 else ""), (parts[1] if len(parts) > 2 else "")


def parse_city_state_zip(line: str):
    """'Las Vegas, NV  89121-1632' -> (city, state, zip)"""
    m = re.match(r"^(.+?),\s+([A-Z]{2})\s+([\d\-]+)", line.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
    return "", "", ""


def extract_copay(notes_text: str):
    """
    - notes_text is None  -> None (blank in output)
    - 'NO COPAY' present  -> 0
    - 'CP$<n>' present    -> float(n)
    - otherwise           -> None (blank)
    """
    if notes_text is None:
        return None
    upper = notes_text.upper()
    if re.search(r'\bNO\s+CO[-\s]?PAY\b', upper):
        return 0
    m = re.search(r'CP\$(\d+(?:\.\d+)?)', upper)
    if m:
        return float(m.group(1))
    return None


def infer_insurance_type(plan_name: str) -> str:
    up = plan_name.upper()
    if "PPO"      in up: return "PPO"
    if "HMO"      in up: return "HMO"
    if "MEDICARE" in up or "MCRE" in up: return "Medicare"
    if "MEDICAID" in up or "MEDI-CAL" in up: return "Medicaid"
    if re.search(r'\bVA\b', up) or "VETERANS" in up: return "VA"
    return ""


# ---------------------------------------------------------------------------
# Insurance section parser
# ---------------------------------------------------------------------------

# Known keywords that signal non-plan-data lines
_SKIP_INS = {"active", "inactive", "termed", "self", "spouse", "child", "other",
             "y", "n", "yes", "no"}


def parse_insurance_section(ins_lines: list) -> list:
    """
    PyMuPDF splits each insurance row across multiple lines:
      '1 PLANCODE - PLAN NAME'  <- plan line
      'MEMBER_ID'
      'Active'
      'SUBSCRIBER NAME'
      'Self'
      'Y'
    A new policy starts when we see a line matching a plan-code pattern.
    """
    policies = []
    cur = None

    def _new_policy(line):
        # Returns (plan_code, plan_display_name)
        # Display name = the human-readable part AFTER " - " only
        # Format A: "1 PLANCODE - Name"
        m = re.match(r'^(\d+)\s+([A-Z][A-Z0-9]+)\s+-\s*(.*)', line)
        if m:
            return m.group(2), m.group(3).strip()
        # Format B: "PLANCODE - Name"
        m = re.match(r'^([A-Z][A-Z0-9]{1,})\s+-\s*(.*)', line)
        if m:
            return m.group(1), m.group(2).strip()
        # Format C: "3794 - Name" (numeric code only)
        m = re.match(r'^(\d+)\s+-\s*(.*)', line)
        if m:
            return m.group(1), m.group(2).strip()
        return None, None

    def _resolve_ids(p):
        """If Active was never seen (page cut off), still assign _ids to member/group."""
        if not p["member_id"] and p.get("_ids"):
            ids = p["_ids"]
            p["member_id"] = ids[-1]
            p["group"]     = ids[-2] if len(ids) >= 2 else ""
        return p

    for line in ins_lines:
        line = line.strip()
        if not line:
            continue

        code, plan_name = _new_policy(line)
        if code:
            if cur:
                policies.append(_resolve_ids(cur))
            cur = {"code": code, "name": plan_name, "group": "", "member_id": "", "status": "",
                   "_ids": [], "expecting": "ids"}
        elif cur:
            exp = cur.get("expecting", "")
            low = line.lower()

            if exp == "ids":
                if not cur["name"]:
                    # Plan name was empty — first non-plan line is the display name
                    cur["name"] = line.strip()
                elif low in ("active", "inactive", "termed"):
                    # End of IDs — resolve: last token = member_id, prev = group
                    ids = cur["_ids"]
                    cur["member_id"] = ids[-1] if ids else ""
                    cur["group"]     = ids[-2] if len(ids) >= 2 else ""
                    cur["status"]    = line.strip()
                    cur["expecting"] = "done"
                elif re.search(r'\d', line) and len(line.split()) <= 2:
                    # ID-like token (short, contains digit)
                    cur["_ids"].append(line.strip())
                else:
                    # Plan name continuation text (e.g. "VETERANS CHOICE", "PROGRAM")
                    cur["name"] = (cur["name"] + " " + line).strip()

    if cur:
        policies.append(_resolve_ids(cur))

    return policies


# ---------------------------------------------------------------------------
# Page parser (PyMuPDF format)
# ---------------------------------------------------------------------------

def _join_addr(parts: list) -> str:
    """Join address fragment lines, skipping any apt/unit fragment whose
    number is already present in the accumulated text (avoids 'Apt 1248 Apt 1248')."""
    result = []
    accumulated = ""
    for part in parts:
        m = re.match(r'^(apt|unit|#|ste|suite|trlr|lot)\s*(\w+)', part, re.IGNORECASE)
        if m:
            unit_num = m.group(2)
            if unit_num.lower() in accumulated.lower():
                continue   # already present — skip duplicate
        result.append(part)
        accumulated += " " + part
    return " ".join(result).strip()


# Lines that are known section headers or column labels — skip them
_SECTION_HEADERS = re.compile(
    r'^(Default Account:|Class:|SSN:|Emp Status:|Chart #:|Employee ID:|Registered:|'
    r'Employer:|First Visit:|Last Visit:|Marital:|Consent:|Referral Src:|'
    r'Accounts:|IDs:|Code|Description|Number|Guarantor|Acct Finance|'
    r'Patient Report|Heart Center|Selections:|Page Break:|Page \d+$)', re.IGNORECASE
)

_KNOWN_KEYS = re.compile(
    r'^(Sex:|DOB:|Lang:|Race\s*/\s*Ethnicity:|Email:|Home:|Work:|Mobile\s*:|'
    r'Assigned:|Referring:|Notes:|Ins\.\s*Policies:)', re.IGNORECASE
)


def parse_page(text: str) -> dict | None:
    if not text:
        return None

    lines = [l.rstrip() for l in text.splitlines()]

    # Find "Patient #" line — data starts 2 lines later (EMR ID)
    pat_idx = None
    for i, line in enumerate(lines):
        if line.strip() == "Patient #":
            pat_idx = i
            break
    if pat_idx is None:
        return None

    # EMR ID is 2 lines after "Patient #" (line after is "Name")
    if pat_idx + 3 >= len(lines):
        return None

    emr_id   = lines[pat_idx + 2].strip()
    raw_name = lines[pat_idx + 3].strip()

    if not emr_id or not raw_name:
        return None

    # EMR ID must be numeric — rejects continuation pages where note class tokens
    # (STICKY, APPT, DIS …) appear in place of a real patient ID
    if not re.match(r'^\d+(\.\d+)?$', emr_id):
        return None
    # Patient name must be in LASTNAME, FIRSTNAME format
    if "," not in raw_name:
        return None

    last, first, middle = parse_name(raw_name)
    full_name = " ".join(p for p in [first, middle, last] if p)

    # --- Walk remaining lines ---
    fields     = {}
    addr_lines = []   # raw address fragments
    notes_parts = []  # note text lines
    ins_lines  = []   # insurance block lines

    section = "header"
    last_key = None   # tracks what key we just saw, for next-line value pickup

    i = pat_idx + 4
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        i += 1

        if not stripped:
            continue

        # --- Section transitions ---
        if stripped == "Notes:" or stripped == "Notes :":
            section = "notes"
            last_key = None
            continue
        if re.match(r'^Accounts\s*:', stripped):
            section = "accounts"
            last_key = None
            continue
        if re.match(r'^Ins\.\s*Policies\s*:', stripped):
            section = "ins"
            last_key = None
            continue
        if re.match(r'^IDs\s*:', stripped):
            section = "ids"
            last_key = None
            continue

        # Skip known table-header / report-header rows
        if re.match(r'^(# Plan|Group|Claim Member|Relation|Accept Assign|'
                    r'Acct Finance|Default|Status|Balance|Number|Guarantor|'
                    r'Code|Description|LEGACYID|PID\d|Class\s+Date)', stripped):
            continue

        # ====== HEADER section ======
        if section == "header":

            # -- Inline key: value lines --
            m = re.match(r'^Sex:\s*(.*)', stripped)
            if m: fields["sex"] = m.group(1).strip(); last_key = None; continue

            m = re.match(r'^DOB:\s*([\d/]+)', stripped)
            if m: fields["dob"] = m.group(1).strip(); last_key = None; continue

            m = re.match(r'^Lang:\s*(\S+)', stripped)
            if m: fields["lang"] = m.group(1).strip(); last_key = None; continue

            m = re.match(r'^Race\s*/\s*Ethnicity:\s*(.*)', stripped, re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if val:
                    fields["race"] = val.split("/")[0].strip()
                    last_key = None
                else:
                    last_key = "race"  # value is on next line
                continue

            m = re.match(r'^Email:\s*(.*)', stripped, re.IGNORECASE)
            if m: fields["email"] = m.group(1).strip(); last_key = None; continue

            m = re.match(r'^Home:\s*(.*)', stripped, re.IGNORECASE)
            if m:
                fields["home_phone"] = m.group(1).strip().rstrip("*").strip()
                last_key = None; continue

            m = re.match(r'^Work:\s*(.*)', stripped, re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if re.match(r'[\(\d]', val):
                    fields["work_phone"] = val.rstrip("*").strip()
                last_key = None; continue

            m = re.match(r'^Mobile\s*:\s*(.*)', stripped, re.IGNORECASE)
            if m:
                fields["mobile_phone"] = m.group(1).strip().rstrip("*").strip()
                last_key = None; continue

            m = re.match(r'^Last Visit:\s*([\d/]+)', stripped, re.IGNORECASE)
            if m: fields["last_visit"] = m.group(1).strip(); last_key = None; continue

            # "Assigned:" — value is on the SAME line after spaces, or NEXT line
            m = re.match(r'^Assigned:\s*(.*)', stripped, re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if val:
                    fields["assigned"] = val
                    last_key = None
                else:
                    last_key = "assigned"
                continue

            # "Referral Src: ... Referring: ..." — referring value on next line
            if re.search(r'Referring:', stripped, re.IGNORECASE):
                m = re.match(r'.*Referring:\s*(.*)', stripped, re.IGNORECASE)
                val = m.group(1).strip() if m else ""
                if val:
                    fields["referring"] = val
                    last_key = None
                else:
                    last_key = "referring"
                continue

            # Consume next-line values for keys that need them
            if last_key == "race":
                fields["race"] = stripped.split("/")[0].strip()
                last_key = None; continue

            if last_key == "assigned":
                if stripped and not re.match(r'^(Work:|Mobile|Referral)', stripped, re.IGNORECASE):
                    fields["assigned"] = stripped
                last_key = None; continue

            if last_key == "referring":
                if stripped:
                    fields["referring"] = stripped
                last_key = None; continue

            # Skip section-header-only lines
            if re.match(r'^(Registered:|First Visit:|Marital:|Consent:|'
                        r'SSN:|Emp Status:|Chart #:|Employee ID:|Employer:|'
                        r'Default Account:|Class:|Referral Src:)', stripped, re.IGNORECASE):
                # Check if this line also signals "next line = address"
                if re.match(r'^(Emp Status:|Employee ID:|Employer:)', stripped, re.IGNORECASE):
                    last_key = "addr_" + re.match(r'^(\w+)', stripped).group(1).lower()
                else:
                    last_key = None
                continue

            # Address lines: lines that follow address-context keys
            if last_key and last_key.startswith("addr_"):
                # Any non-empty line here is an address fragment
                addr_lines.append(stripped)
                last_key = None
                continue

            # Fallback: if line looks like an address fragment (has digits or city pattern)
            # and we're in early header (before Email:)
            if "email" not in fields and "home_phone" not in fields:
                if not _SECTION_HEADERS.match(stripped) and not _KNOWN_KEYS.match(stripped):
                    addr_lines.append(stripped)

        # ====== NOTES section ======
        elif section == "notes":
            # Skip column header row and note class / date-user lines
            # Keep only the actual note text lines:
            #   - Lines starting with " - " are note content
            #   - Continuation lines (plain text following a note text line)
            # Skip: "Class", "Date / User", "Note", class codes (DIS/APPT/GEN/…),
            #        and date-user lines like "03/20/2026 - JOZLYN HERNANDEZ"
            if re.match(r'^(Class|Date\s*/\s*User|Note)$', stripped, re.IGNORECASE):
                continue
            if re.match(r'^(DIS|APPT|GEN|CONSNT|STICKY|PMD|AMEND|LAB|MSG|TODO)$', stripped):
                continue
            if re.match(r'^\d{2}/\d{2}/\d{4}\s+-\s+', stripped):
                # Date - Username line — skip
                continue
            # Actual note text: starts with " - " or is a continuation of previous note
            if line.startswith(" - ") or line.startswith("- "):
                notes_parts.append(stripped.lstrip("- ").strip())
            elif notes_parts and not re.match(
                    r'^(DIS|APPT|GEN|CONSNT|STICKY|PMD|\d{2}/\d{2}/)', stripped):
                # Continuation line of the previous note text
                notes_parts.append(stripped)

        # ====== INS section ======
        elif section == "ins":
            ins_lines.append(stripped)

    # --- Build address ---
    street = city = state = zip_ = ""
    if addr_lines:
        # Last address fragment matching city/state/zip pattern is the city line
        found = False
        for idx in range(len(addr_lines) - 1, -1, -1):
            c, s, z = parse_city_state_zip(addr_lines[idx])
            if s:
                city, state, zip_ = c, s, z
                street = _join_addr(addr_lines[:idx])
                found = True
                break
        if not found:
            street = _join_addr(addr_lines)

    # --- Copay from notes ---
    notes_text = "\n".join(notes_parts) if notes_parts else None
    copay = extract_copay(notes_text)

    # --- Insurance ---
    policies = parse_insurance_section(ins_lines)

    def pol(n):
        return policies[n] if n < len(policies) else {}

    p1, p2, p3 = pol(0), pol(1), pol(2)

    medicare_id = next(
        (p["member_id"] for p in policies
         if "MEDICARE" in p.get("name", "").upper() or re.match(r'^MCR', p.get("code", ""), re.I)),
        ""
    )

    provider = fields.get("assigned", "")

    rec = {
        "EMR ID":               emr_id,
        "PATIENT EMR NAME":     raw_name,
        "FIRST NAME":           first,
        "MIDDLE NAME":          middle,
        "LAST NAME":            last,
        "PATIENT FULL NAME":    full_name,
        "DATE OF BIRTH":        fields.get("dob", ""),
        "GENDER":               fields.get("sex", ""),
        "STREET ADDRESS":       street,
        "CITY":                 city,
        "STATE":                state,
        "ZIP":                  zip_,
        "HOME PHONE":           fields.get("home_phone", ""),
        "MOBILE PHONE":         fields.get("mobile_phone", ""),
        "WORK PHONE":           fields.get("work_phone", ""),
        "EMAIL ADDRESS":        fields.get("email", ""),
        "LANGUAGE":             fields.get("lang", ""),
        "RACE":                 fields.get("race", ""),
        "EMERGENCY CONTACT NAME":        "",
        "EMERGENCY RELATIONSHIP":        "",
        "EMERGENCY CONTACT HOME PHONE":  "",
        "EMERGENCY CONTACT MOBILE PHONE":"",
        "MEDICARE ID":          medicare_id,
        "PRIMARY INSURANCE":    p1.get("name", ""),
        "PRIMARY ID":           p1.get("member_id", ""),
        "PRIMARY GROUP":        p1.get("group", ""),
        "SECONDARY INSURANCE":  p2.get("name", ""),
        "SECONDARY ID":         p2.get("member_id", ""),
        "SECONDARY GROUP":      p2.get("group", ""),
        "TERITARY INSURANCE":   p3.get("name", ""),
        "TERITARY ID":          p3.get("member_id", ""),
        "TERITARY GROUP":       p3.get("group", ""),
        "INSURANCE TYPE":       infer_insurance_type(p1.get("name", "")),
        "CO-PAY":               "" if copay is None else copay,
        **EMPTY_COMORBIDITIES,
        "PRIMARY DX": "", "SECONDARY DX": "", "PRIMARY ICD": "", "SECONDARY ICD": "",
        "LAST SEEN DATE":       fields.get("last_visit", ""),
        "NEXT APPT":            "",
        "PROVIDER DATA":        provider,
        "PROVIDER NAME":        provider,
        "CLINIC FACILITY":      "Heart Center of Nevada",
        "PRIMARY CARE PROVIDER":fields.get("referring", ""),
        "MEDICATIONS":          "",
        "ENCOUNTER NOTES":      notes_text or "",
    }
    return rec


# ---------------------------------------------------------------------------
# CSV -> Excel (write-only = constant memory)
# ---------------------------------------------------------------------------

def csv_to_excel(csv_path: Path, xlsx_path: Path):
    import openpyxl
    print("Converting CSV -> Excel ...")
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("HCN Consolidated")
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(xlsx_path)
    csv_path.unlink()
    print(f"Saved : {xlsx_path}")


# ---------------------------------------------------------------------------
# Appointment PDF pipeline
# ---------------------------------------------------------------------------

_APPT_TODAY = date.today()
_APPT_PATIENT_HDR = re.compile(r'^([A-Z][A-Z ,.\'-]+?)\s{2,}\(\d+y\s+[MF]\)')
_APPT_DATE_LINE   = re.compile(r'^\d{2}/\d{2}/\d{4}$')
_APPT_SKIP = {
    'Appt Date', 'Time', 'Len', 'Prov', 'Location', 'Equipment',
    'Reason', 'Room', 'Encounter #', 'Appointment Report By Patient',
    'Heart Center of Nevada', 'Selections:', 'Appointment Dates:',
}


def _appt_norm_name(raw: str) -> str:
    raw = re.sub(r'\s*\(\d+y\s+[MF]\)', '', raw).strip()
    parts = [p.strip() for p in raw.split(',', 1)]
    if len(parts) == 2:
        last  = parts[0].lower().strip()
        first = parts[1].lower().split()[0] if parts[1].strip() else ''
        return f"{last},{first}"
    return raw.lower().strip()


def _norm_phone(ph: str) -> str:
    digits = re.sub(r'\D', '', ph)
    return digits[-10:] if len(digits) >= 10 else digits


def _parse_appt_date(s: str):
    try:
        return datetime.strptime(s.strip(), '%m/%d/%Y').date()
    except ValueError:
        return None


def extract_appt_to_csv(pdf_path: Path, ts: str) -> Path:
    out_csv = CLEANED_DIR / f"appt_raw_{ts}.csv"
    doc = fitz.open(str(pdf_path))
    total = doc.page_count
    print(f"  Extracting appointment PDF ({total} pages) …")
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["page_num", "raw_text"])
        for i in range(total):
            text = doc[i].get_text()
            writer.writerow([i + 1, text.replace("\n", "\\n")])
    doc.close()
    print(f"  Appointment raw CSV: {out_csv.name}")
    return out_csv


def parse_appt_csv(raw_csv: Path) -> tuple:
    """Returns (patients dict, phone_index dict)."""
    patients    = {}
    phone_index = {}

    cur_name = cur_norm = None
    cur_phones: set = set()
    cur_dates: list = []
    in_phones = False

    def _flush():
        nonlocal cur_name, cur_norm, cur_phones, cur_dates, in_phones
        if cur_norm is None:
            return
        future = [d for d in cur_dates if d >= _APPT_TODAY]
        earliest = min(future) if future else None
        if cur_norm not in patients:
            patients[cur_norm] = {'name': cur_name, 'phones': set(cur_phones), 'earliest': earliest}
        else:
            patients[cur_norm]['phones'].update(cur_phones)
            if earliest and (patients[cur_norm]['earliest'] is None
                             or earliest < patients[cur_norm]['earliest']):
                patients[cur_norm]['earliest'] = earliest
        for ph in cur_phones:
            phone_index.setdefault(ph, cur_norm)
        cur_name = cur_norm = None
        cur_phones = set()
        cur_dates = []
        in_phones = False

    with open(raw_csv, newline='', encoding='utf-8') as fh:
        for row in csv.DictReader(fh):
            for line in row['raw_text'].replace('\\n', '\n').splitlines():
                s = line.strip()
                if not s:
                    continue
                if any(s.startswith(h) for h in _APPT_SKIP):
                    in_phones = False; continue
                if re.match(r'^\d{2}/\d{2}/\d{4}\d{2}:\d{2} [AP]M$', s): continue
                if re.match(r'^Page \d+$', s): continue
                if s.startswith('From:') or s.startswith('To:'): continue

                m = _APPT_PATIENT_HDR.match(s)
                if m:
                    _flush()
                    cur_name = m.group(1).strip()
                    cur_norm = _appt_norm_name(s)
                    in_phones = False
                    continue

                if cur_norm is None:
                    continue

                if s == 'Phone Numbers:':
                    in_phones = True; continue

                if in_phones:
                    for prefix in ('Home:', 'Work:', 'Mobile:'):
                        if s.startswith(prefix):
                            ph_raw = s[len(prefix):].strip().rstrip('*').strip()
                            if ph_raw:
                                ph = _norm_phone(ph_raw)
                                if ph:
                                    cur_phones.add(ph)
                            break
                    else:
                        ph_raw = s.rstrip('*').strip()
                        if re.match(r'^\(?\d', ph_raw):
                            ph = _norm_phone(ph_raw)
                            if ph:
                                cur_phones.add(ph)
                        else:
                            in_phones = False

                if _APPT_DATE_LINE.match(s):
                    d = _parse_appt_date(s)
                    if d:
                        cur_dates.append(d)

    _flush()
    return patients, phone_index


def apply_appointments(xlsx_path: Path, patients: dict, phone_index: dict) -> Path:
    import openpyxl
    print("  Applying NEXT APPT …")
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(name):
        return (headers.index(name) + 1) if name in headers else None

    first_col = col('FIRST NAME')
    last_col  = col('LAST NAME')
    home_col  = col('HOME PHONE')
    cell_col  = col('MOBILE PHONE')
    work_col  = col('WORK PHONE')
    appt_col  = col('NEXT APPT')
    if appt_col is None:
        appt_col = len(headers) + 1
        ws.cell(1, appt_col, 'NEXT APPT')

    matched = phone_matched = name_matched = 0
    for row_idx in range(2, ws.max_row + 1):
        first = (ws.cell(row_idx, first_col).value or '').strip()
        last  = (ws.cell(row_idx, last_col).value  or '').strip()
        if not first and not last:
            continue

        rec = None
        # Primary: phone match
        for ph_col in (home_col, cell_col, work_col):
            if ph_col is None:
                continue
            ph = _norm_phone(str(ws.cell(row_idx, ph_col).value or ''))
            if ph and ph in phone_index:
                rec = patients.get(phone_index[ph])
                if rec:
                    phone_matched += 1
                    break

        # Secondary: exact normalised name
        if rec is None:
            first_token = first.lower().split()[0] if first else ''
            norm = f"{last.lower()},{first_token}"
            rec = patients.get(norm)
            if rec:
                name_matched += 1

        if rec and rec['earliest']:
            ws.cell(row_idx, appt_col).value = rec['earliest'].strftime('%m/%d/%Y')
            matched += 1

    ts2 = datetime.now().strftime('%Y%m%d_%H%M%S')
    out = OUTPUT_DIR / f"HCN_consolidated_{ts2}.xlsx"
    wb.save(str(out))
    wb.close()
    print(f"  NEXT APPT matched={matched} (phone={phone_matched} name={name_matched})")
    print(f"  Saved: {out.name}")
    return out


# ---------------------------------------------------------------------------
# Drugs PDF pipeline  (PatientMedication Detail Report By Patient)
# ---------------------------------------------------------------------------

_DIAG_LINE_RE  = re.compile(r'.+\s+\([A-Z0-9][A-Z0-9.\s]+\)\s*$')
_PATIENT_ID_RE = re.compile(r'^\d+(\.\d+)?$')
_DATE_STATUS   = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(Current|Lapsed|Discontinued|Void)$')
_DRUGS_SKIP    = {
    'Patient Medication Detail Report By Patient',
    'Type', 'Issued', 'Drug Description', 'Provider Name',
    'Quantity', 'Days', 'SIG', 'Pharmacy', 'Status', 'Units', 'Refills',
    'Heart Center of Nevada',
}
_MED_TYPES    = {'Patient Reported', 'Original', 'Renewal', 'Denied', 'POST'}
_MED_STATUSES = {'Current', 'Discontinued', 'Lapsed', 'Denied', 'Void'}
_MED_DATE_RE  = re.compile(r'^\d{2}/\d{2}/\d{2}$')   # MM/DD/YY
_PATIENT_NUM_RE = re.compile(r'^Patient #\s*:\s*(.+)$')

# Drug name keyword → comorbidity for PRIMARY DX inference.
# Only maps to heart-clinic-relevant, NON-DISALLOWED comorbidities.
# Disallowed as primary DX: ASTHMA, ANGINA PECTORIS, CHRONIC KIDNEY DISEASE,
#   COPD, DIABETES, DYSPNEA, OBESITY, SLEEP APNEA
_DRUG_KEYWORDS: dict = {
    # CORONARY ARTERY DISEASE — antiplatelets, nitrates
    'clopidogrel':   'CORONARY ARTERY DISEASE',
    'plavix':        'CORONARY ARTERY DISEASE',
    'ticagrelor':    'CORONARY ARTERY DISEASE',
    'brilinta':      'CORONARY ARTERY DISEASE',
    'prasugrel':     'CORONARY ARTERY DISEASE',
    'effient':       'CORONARY ARTERY DISEASE',
    'isosorbide':    'CORONARY ARTERY DISEASE',
    'nitroglycerin': 'CORONARY ARTERY DISEASE',
    'ranexa':        'CORONARY ARTERY DISEASE',
    'ranolazine':    'CORONARY ARTERY DISEASE',

    # ARRHYTHMIA — antiarrhythmics and anticoagulants (afib)
    'amiodarone':    'ARRHYTHMIA',
    'flecainide':    'ARRHYTHMIA',
    'propafenone':   'ARRHYTHMIA',
    'sotalol':       'ARRHYTHMIA',
    'dronedarone':   'ARRHYTHMIA',
    'multaq':        'ARRHYTHMIA',
    'digoxin':       'ARRHYTHMIA',
    'eliquis':       'ARRHYTHMIA',
    'apixaban':      'ARRHYTHMIA',
    'xarelto':       'ARRHYTHMIA',
    'rivaroxaban':   'ARRHYTHMIA',
    'pradaxa':       'ARRHYTHMIA',
    'dabigatran':    'ARRHYTHMIA',
    'warfarin':      'ARRHYTHMIA',
    'coumadin':      'ARRHYTHMIA',

    # CONGESTIVE HEART FAILURE — loop diuretics, Entresto, aldosterone antagonists
    'furosemide':    'CONGESTIVE HEART FAILURE',
    'lasix':         'CONGESTIVE HEART FAILURE',
    'torsemide':     'CONGESTIVE HEART FAILURE',
    'bumetanide':    'CONGESTIVE HEART FAILURE',
    'spironolactone':'CONGESTIVE HEART FAILURE',
    'eplerenone':    'CONGESTIVE HEART FAILURE',
    'entresto':      'CONGESTIVE HEART FAILURE',
    'sacubitril':    'CONGESTIVE HEART FAILURE',
    'carvedilol':    'CONGESTIVE HEART FAILURE',
    'bisoprolol':    'CONGESTIVE HEART FAILURE',

    # PERIPHERAL VASCULAR — vasodilators specific to PAD
    'cilostazol':    'PERIPHERAL VASCULAR',
    'pletal':        'PERIPHERAL VASCULAR',
    'pentoxifylline':'PERIPHERAL VASCULAR',
    'trental':       'PERIPHERAL VASCULAR',

    # CERBOVASCULAR ACCIDENT — stroke prevention (non-afib anticoagulants)
    'clopidogrel':   'CERBOVASCULAR ACCIDENT',   # also CAD — first occurrence wins; keep CAD above

    # HYPERLIPIDEMIA — statins and lipid agents
    'atorvastatin':  'HYPERLIPIDEMIA',
    'rosuvastatin':  'HYPERLIPIDEMIA',
    'simvastatin':   'HYPERLIPIDEMIA',
    'pravastatin':   'HYPERLIPIDEMIA',
    'lovastatin':    'HYPERLIPIDEMIA',
    'fluvastatin':   'HYPERLIPIDEMIA',
    'pitavastatin':  'HYPERLIPIDEMIA',
    'lipitor':       'HYPERLIPIDEMIA',
    'crestor':       'HYPERLIPIDEMIA',
    'zocor':         'HYPERLIPIDEMIA',
    'fenofibrate':   'HYPERLIPIDEMIA',
    'tricor':        'HYPERLIPIDEMIA',
    'ezetimibe':     'HYPERLIPIDEMIA',
    'zetia':         'HYPERLIPIDEMIA',
    'repatha':       'HYPERLIPIDEMIA',
    'praluent':      'HYPERLIPIDEMIA',

    # HYPERTENSION — ACE inhibitors, ARBs, CCBs, diuretics, beta blockers
    'lisinopril':    'HYPERTENSION',
    'enalapril':     'HYPERTENSION',
    'ramipril':      'HYPERTENSION',
    'benazepril':    'HYPERTENSION',
    'quinapril':     'HYPERTENSION',
    'fosinopril':    'HYPERTENSION',
    'perindopril':   'HYPERTENSION',
    'losartan':      'HYPERTENSION',
    'valsartan':     'HYPERTENSION',
    'olmesartan':    'HYPERTENSION',
    'irbesartan':    'HYPERTENSION',
    'candesartan':   'HYPERTENSION',
    'telmisartan':   'HYPERTENSION',
    'amlodipine':    'HYPERTENSION',
    'diltiazem':     'HYPERTENSION',
    'verapamil':     'HYPERTENSION',
    'hydralazine':   'HYPERTENSION',
    'clonidine':     'HYPERTENSION',
    'doxazosin':     'HYPERTENSION',
    'terazosin':     'HYPERTENSION',
    'hydrochlorothiazide': 'HYPERTENSION',
    'chlorthalidone':'HYPERTENSION',
    'indapamide':    'HYPERTENSION',
    'metoprolol':    'HYPERTENSION',
    'atenolol':      'HYPERTENSION',
    'nadolol':       'HYPERTENSION',
    'propranolol':   'HYPERTENSION',
    'labetalol':     'HYPERTENSION',
    'nebivolol':     'HYPERTENSION',
    'bystolic':      'HYPERTENSION',

    # HYPOTENSION — vasopressors / volume expanders
    'midodrine':     'HYPOTENSION',
    'orvaten':       'HYPOTENSION',
    'fludrocortisone':'HYPOTENSION',
    'florinef':      'HYPOTENSION',
}


def _infer_comorb_from_drug(drug_name: str) -> str:
    d = drug_name.lower()
    for kw, comorb in _DRUG_KEYWORDS.items():
        if kw in d:
            return comorb
    return ''


def extract_drugs_to_csv(pdf_path: Path, ts: str) -> Path:
    out_csv = CLEANED_DIR / f"drugs_raw_{ts}.csv"
    doc = fitz.open(str(pdf_path))
    total = doc.page_count
    print(f"  Extracting medication PDF ({total} pages) …")
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["page_num", "raw_text"])
        for i in range(total):
            text = doc[i].get_text()
            writer.writerow([i + 1, text.replace("\n", "\\n")])
            if (i + 1) % 10000 == 0:
                print(f"    medication PDF: {i+1}/{total} pages", flush=True)
    doc.close()
    print(f"  Medication raw CSV: {out_csv.name}")
    return out_csv


def parse_drugs_csv(raw_csv: Path) -> dict:
    """
    Parses PatientMedication Detail Report By Patient.
    Returns dict: emr_id str → {'current': set, 'all': set}
    Matches by EMR ID (same numbering as patient report).
    """
    patients: dict = {}
    cur_pid:  str  = ''
    state:    str  = 'idle'   # idle | expect_date | expect_drug | in_med
    cur_drug: str  = ''

    with open(raw_csv, newline='', encoding='utf-8') as fh:
        for row in csv.DictReader(fh):
            for line in row['raw_text'].replace('\\n', '\n').splitlines():
                s = line.strip()
                if not s:
                    continue

                # Skip page headers / column headers
                if s in _DRUGS_SKIP:
                    continue
                if re.match(r'^\d{2}/\d{2}/\d{4} \d', s):   # report timestamp
                    continue
                if re.match(r'^Page \d+$', s):
                    continue
                if s.startswith('Direct:'):     # SIG continuation
                    continue

                # Patient header: "Patient # : 222532.0"
                m = _PATIENT_NUM_RE.match(s)
                if m:
                    raw_id = m.group(1).strip()
                    cur_pid = '' if raw_id == '<Restricted>' else raw_id.split('.')[0].strip()
                    state   = 'idle'
                    cur_drug = ''
                    continue

                if not cur_pid:
                    continue

                # Medication type line → start of a new entry
                if s in _MED_TYPES:
                    state    = 'expect_date'
                    cur_drug = ''
                    continue

                if state == 'expect_date':
                    if _MED_DATE_RE.match(s):
                        state = 'expect_drug'
                    # else: unexpected line, stay in expect_date
                    continue

                if state == 'expect_drug':
                    cur_drug = s
                    state    = 'in_med'
                    continue

                if state == 'in_med':
                    if s in _MED_STATUSES:
                        if cur_drug and cur_pid:
                            if cur_pid not in patients:
                                patients[cur_pid] = {'current': set(), 'all': set()}
                            patients[cur_pid]['all'].add(cur_drug)
                            if s == 'Current':
                                patients[cur_pid]['current'].add(cur_drug)
                        state    = 'idle'
                        cur_drug = ''
                    # else: quantity, days, SIG, pharmacy lines — skip

    return patients


def apply_medications(xlsx_path: Path, patients: dict) -> Path:
    import openpyxl
    print("  Applying MEDICATIONS …")
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(name):
        return (headers.index(name) + 1) if name in headers else None

    emr_col  = col('EMR ID')
    meds_col = col('MEDICATIONS')
    if emr_col is None:
        raise ValueError("'EMR ID' column not found in consolidated Excel")
    if meds_col is None:
        meds_col = len(headers) + 1
        ws.cell(1, meds_col, 'MEDICATIONS')

    matched = 0
    for row_idx in range(2, ws.max_row + 1):
        emr_raw = ws.cell(row_idx, emr_col).value
        if not emr_raw:
            continue
        emr_id = str(emr_raw).split('.')[0].strip()
        rec = patients.get(emr_id)
        if rec:
            drugs = rec['current'] if rec['current'] else rec['all']
            if drugs:
                ws.cell(row_idx, meds_col).value = '; '.join(sorted(drugs))
                matched += 1

    ts2 = datetime.now().strftime('%Y%m%d_%H%M%S')
    out = OUTPUT_DIR / f"HCN_consolidated_{ts2}.xlsx"
    wb.save(str(out))
    wb.close()
    print(f"  MEDICATIONS matched={matched}")
    print(f"  Saved: {out.name}")
    return out


# ---------------------------------------------------------------------------
# Diagnosis PDF pipeline — comorbidity mapping (same logic as HCT)
# ---------------------------------------------------------------------------

_COMORBIDITY_COLUMNS = [
    "CORONARY ARTERY DISEASE", "ARRHYTHMIA", "CONGESTIVE HEART FAILURE",
    "PERIPHERAL VASCULAR", "VALVULAR HEART", "CERBOVASCULAR ACCIDENT",
    "HYPERLIPIDEMIA", "ANGINA PECTORIS", "HYPOTENSION", "HYPERTENSION",
    "OBESITY", "DIABETES", "CHRONIC KIDNEY DISEASE", "COPD",
    "RESPIRATORY FAILURE", "ASTHMA", "SLEEP APNEA", "DYSPNEA",
    "EMPHYSEMA", "BRONCHIECTASIS", "HYPOXEMIA",
]

_CAUSE_TO_COMORBIDITY = {
    "Coronary Artery Disease":        "CORONARY ARTERY DISEASE",
    "Arrhythmia":                     "ARRHYTHMIA",
    "CHF (Congestive Heart Failure)": "CONGESTIVE HEART FAILURE",
    "Peripheral Vascular Disease":    "PERIPHERAL VASCULAR",
    "Valvular Heart Disease":         "VALVULAR HEART",
    "Cerebrovascular Accident":       "CERBOVASCULAR ACCIDENT",
    "Hyperlipidemia":                 "HYPERLIPIDEMIA",
    "Angina Pectoris":                "ANGINA PECTORIS",
    "Hypotension":                    "HYPOTENSION",
    "Hypertension":                   "HYPERTENSION",
    "Obesity":                        "OBESITY",
    "Type 2 Diabetes":                "DIABETES",
    "Chronic Kidney":                 "CHRONIC KIDNEY DISEASE",
    "COPD":                           "COPD",
    "Respiratory Failure":            "RESPIRATORY FAILURE",
    "Asthma":                         "ASTHMA",
    "Sleep Apnea":                    "SLEEP APNEA",
    "Dyspnea":                        "DYSPNEA",
    "Emphysema ":                     "EMPHYSEMA",
    "Emphysema":                      "EMPHYSEMA",
    "Bronchiectasis":                 "BRONCHIECTASIS",
    "Hypoxemia":                      "HYPOXEMIA",
}

_PRIMARY_DX_DISALLOWED = {
    "ASTHMA", "ANGINA PECTORIS", "CHRONIC KIDNEY DISEASE", "COPD",
    "DIABETES", "DYSPNEA", "OBESITY", "SLEEP APNEA",
}

_DESC_KEYWORDS = {
    "CORONARY ARTERY DISEASE":  ["coronary artery disease", "athscl heart disease",
                                  "atherosclerotic heart disease", "ischemic heart",
                                  "aortocoronary bypass"],
    "ARRHYTHMIA":               ["arrhythmia", "atrial fibrillation", "atrial flutter",
                                  "tachycardia", "bradycardia", "palpitations"],
    "CONGESTIVE HEART FAILURE": ["heart failure", "congestive heart failure",
                                  "systolic heart failure", "diastolic heart failure"],
    "PERIPHERAL VASCULAR":      ["peripheral vascular", "peripheral angiopath",
                                  "venous insufficiency"],
    "VALVULAR HEART":           ["valve", "valvular", "mitral", "aortic", "tricuspid",
                                  "prosthetic heart valve"],
    "CERBOVASCULAR ACCIDENT":   ["cerebrovascular", "stroke", "transient cerebral ischemic",
                                  "carotid"],
    "HYPERLIPIDEMIA":           ["hyperlipidemia", "hypercholesterolemia",
                                  "mixed hyperlipidemia"],
    "ANGINA PECTORIS":          ["angina", "precordial pain", "chest pain"],
    "HYPOTENSION":              ["hypotension", "orthostatic hypotension"],
    "HYPERTENSION":             ["hypertension", "hypertensive heart disease"],
}

_DIAG_PDF_SKIP = {
    'Diagnosis Analysis Report', 'Diagnosis', 'Serv Date', 'Patient',
    'Procedure', 'Charge', 'Heart Center of Nevada', 'Selections:',
    'Provider Type:', 'Provider Type: ', 'Performing Provider', 'PATIENT',
}
_ICD_CODE_RE   = re.compile(r'^[A-Z]\d[A-Z0-9.]*$')
_PATIENT_ID_RE = re.compile(r'^\d+(\.\d+)?$')


def _load_api_mappings(api_csv: Path) -> tuple:
    """Return (exact_to_comorbidity, prefix_to_comorbidity, comorbidity_to_icd) dicts."""
    exact:        dict = {}
    prefix:       dict = {}
    comorb_to_icd: dict = {}
    with open(api_csv, newline='', encoding='utf-8') as fh:
        for row in csv.DictReader(fh):
            cause = row.get('cause', '').strip()
            icd   = row.get('icd_code', '').strip().upper().replace(' ', '')
            if not cause or not icd:
                continue
            comorbidity = _CAUSE_TO_COMORBIDITY.get(cause) or _CAUSE_TO_COMORBIDITY.get(cause.strip())
            if not comorbidity:
                continue
            exact[icd] = comorbidity
            p3 = icd.replace('.', '')[:3]
            if p3:
                prefix[p3] = comorbidity
            # Store first ICD seen per comorbidity as the representative code
            comorb_to_icd.setdefault(comorbidity, icd)
    return exact, prefix, comorb_to_icd


def _match_icd(icd: str, exact: dict, prefix: dict) -> str:
    norm = icd.strip().upper().replace(' ', '')
    if not norm:
        return ''
    if norm in exact:
        return exact[norm]
    p3 = norm.replace('.', '')[:3]
    return prefix.get(p3, '')


def _desc_comorbidity(desc: str) -> str:
    d = desc.lower()
    for comorbidity, kws in _DESC_KEYWORDS.items():
        if any(kw in d for kw in kws):
            return comorbidity
    return ''


def _pick_primary_secondary(yes_list: list) -> tuple:
    primary = next((c for c in yes_list if c not in _PRIMARY_DX_DISALLOWED), '')
    if not primary and yes_list:
        primary = yes_list[0]
    secondary = next((c for c in yes_list if c != primary), '')
    return primary, secondary


def extract_diag_to_csv(pdf_path: Path, ts: str) -> Path:
    out_csv = CLEANED_DIR / f"diag_raw_{ts}.csv"
    doc = fitz.open(str(pdf_path))
    total = doc.page_count
    print(f"  Extracting diagnosis PDF ({total} pages) …")
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["page_num", "raw_text"])
        for i in range(total):
            text = doc[i].get_text()
            writer.writerow([i + 1, text.replace("\n", "\\n")])
    doc.close()
    print(f"  Diagnosis raw CSV: {out_csv.name}")
    return out_csv


def parse_diag_csv(raw_csv: Path) -> dict:
    """
    Returns dict: emr_id str → {icd_code: description}.
    Uses state machine: 'icd_block' collects ICD codes,
    'visit_block' captures patient IDs under those ICDs.
    Transitions: new ICD code while in visit_block → new group.
    """
    patients: dict = {}   # emr_id → {icd: desc}
    cur_icds: dict = {}   # current ICD group: {code: description}
    next_is_desc = None   # if set, next line is description for this ICD code
    after_date   = False  # if True, next pure-numeric line is a patient ID
    state        = "icd_block"

    with open(raw_csv, newline='', encoding='utf-8') as fh:
        for row in csv.DictReader(fh):
            for line in row['raw_text'].replace('\\n', '\n').splitlines():
                s = line.strip()
                if not s:
                    continue
                if s in _DIAG_PDF_SKIP:
                    continue
                if re.match(r'^\d{2}/\d{2}/\d{4} \d', s):   # report timestamp
                    continue
                if re.match(r'^Page \d+$', s):
                    continue

                # ICD code line
                if _ICD_CODE_RE.match(s):
                    if state == "visit_block":
                        # New ICD block starts — clear previous group
                        cur_icds = {}
                    cur_icds[s] = ''
                    next_is_desc = s
                    state = "icd_block"
                    after_date = False
                    continue

                # Description line immediately after ICD code
                if next_is_desc is not None:
                    cur_icds[next_is_desc] = s
                    next_is_desc = None
                    continue

                # Date line → enter visit block, arm patient-ID capture
                if re.match(r'^\d{2}/\d{2}/\d{4}$', s):
                    state = "visit_block"
                    after_date = True
                    continue

                # Patient ID capture
                if after_date:
                    after_date = False
                    if _PATIENT_ID_RE.match(s):
                        pid = s.split('.')[0].strip()
                        if cur_icds and pid:
                            if pid not in patients:
                                patients[pid] = {}
                            patients[pid].update(cur_icds)
                    continue

                # Everything else (name, proc code, charge, total): ignore
    return patients


def apply_comorbidities(xlsx_path: Path, patients: dict) -> Path:
    import openpyxl
    print("  Applying comorbidities / diagnoses …")

    exact, prefix, _ = _load_api_mappings(API_CAUSE_CSV)

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(name):
        return (headers.index(name) + 1) if name in headers else None

    emr_col = col('EMR ID')
    if emr_col is None:
        raise ValueError("'EMR ID' column not found")

    # Comorbidity column indices
    comorb_cols = {name: col(name) for name in _COMORBIDITY_COLUMNS}
    pri_dx_col   = col('PRIMARY DX')
    sec_dx_col   = col('SECONDARY DX')
    pri_icd_col  = col('PRIMARY ICD')
    sec_icd_col  = col('SECONDARY ICD')

    matched = 0
    for row_idx in range(2, ws.max_row + 1):
        emr_raw = ws.cell(row_idx, emr_col).value
        if not emr_raw:
            continue
        pid = str(emr_raw).split('.')[0].strip()
        icd_map = patients.get(pid)
        if not icd_map:
            continue

        flags: dict = {c: 'NO' for c in _COMORBIDITY_COLUMNS}
        icd_for_comorbidity: dict = {}  # comorbidity → first matching ICD code

        for icd, desc in icd_map.items():
            hit = _match_icd(icd, exact, prefix)
            if not hit and desc:
                hit = _desc_comorbidity(desc)
            if hit:
                flags[hit] = 'YES'
                if hit not in icd_for_comorbidity:
                    icd_for_comorbidity[hit] = icd

        yes_list = [c for c in _COMORBIDITY_COLUMNS if flags[c] == 'YES']
        if not yes_list:
            continue

        # Write comorbidity YES/NO
        for name, cidx in comorb_cols.items():
            if cidx:
                ws.cell(row_idx, cidx).value = flags[name]

        # PRIMARY / SECONDARY DX + ICD
        pri, sec = _pick_primary_secondary(yes_list)
        if pri_dx_col:  ws.cell(row_idx, pri_dx_col).value  = pri
        if sec_dx_col:  ws.cell(row_idx, sec_dx_col).value  = sec
        if pri_icd_col: ws.cell(row_idx, pri_icd_col).value = icd_for_comorbidity.get(pri, '')
        if sec_icd_col: ws.cell(row_idx, sec_icd_col).value = icd_for_comorbidity.get(sec, '')
        matched += 1

    ts2 = datetime.now().strftime('%Y%m%d_%H%M%S')
    out = OUTPUT_DIR / f"HCN_consolidated_{ts2}.xlsx"
    wb.save(str(out))
    wb.close()
    print(f"  Comorbidities matched={matched}")
    print(f"  Saved: {out.name}")
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not PDF_PATH or not PDF_PATH.exists():
        print("ERROR: No PatientReport*.PDF found in src/HCN/.")
        sys.exit(1)

    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path  = CLEANED_DIR / f"HCN_patients_{ts}.csv"
    xlsx_path = OUTPUT_DIR  / f"HCN_consolidated_{ts}.xlsx"

    # ------------------------------------------------------------------
    # Step 1 — Parse PatientReport PDF → CSV → Excel
    # ------------------------------------------------------------------
    print(f"Source : {PDF_PATH.name}")
    doc   = fitz.open(str(PDF_PATH))
    total = doc.page_count
    print(f"Pages  : {total:,}")
    print("Parsing and streaming to CSV ...")

    seen    = set()
    written = skipped = errors = 0

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TEMPLATE_COLS)
        writer.writeheader()

        for i in range(total):
            if i % 10000 == 0 and i > 0:
                pct = i * 100 // total
                print(f"  {pct}%  page {i:,}/{total:,}  written={written:,}")
            try:
                text = doc[i].get_text()
                rec  = parse_page(text)
                if rec is None:
                    skipped += 1; continue
                eid = rec["EMR ID"]
                if eid in seen:
                    skipped += 1; continue
                if not rec["PRIMARY INSURANCE"]:
                    skipped += 1
                    continue
                seen.add(eid)
                writer.writerow(rec)
                written += 1
            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"  Warning page {i+1}: {e}")

    doc.close()
    print(f"\nStep 1 complete — {written:,} patients | skipped={skipped:,} | errors={errors:,}")

    # ------------------------------------------------------------------
    # Step 2 — Extract appt + drugs + diagnosis PDFs to CSV in parallel
    # ------------------------------------------------------------------
    from concurrent.futures import ThreadPoolExecutor

    print("\n--- Step 2: Extracting supplementary PDFs in parallel ---")
    extract_tasks = {}
    with ThreadPoolExecutor(max_workers=3) as pool:
        if APPT_PDF_PATH and APPT_PDF_PATH.exists():
            extract_tasks['appt']  = pool.submit(extract_appt_to_csv,  APPT_PDF_PATH,  ts)
        if DRUGS_PDF_PATH and DRUGS_PDF_PATH.exists():
            extract_tasks['drugs'] = pool.submit(extract_drugs_to_csv, DRUGS_PDF_PATH, ts)
        if DIAG_PDF_PATH and DIAG_PDF_PATH.exists():
            extract_tasks['diag']  = pool.submit(extract_diag_to_csv,  DIAG_PDF_PATH,  ts)

    appt_csv  = extract_tasks['appt'].result()  if 'appt'  in extract_tasks else None
    drugs_csv = extract_tasks['drugs'].result() if 'drugs' in extract_tasks else None
    diag_csv  = extract_tasks['diag'].result()  if 'diag'  in extract_tasks else None

    # ------------------------------------------------------------------
    # Step 3 — Parse all supplementary CSVs into lookup dicts
    # ------------------------------------------------------------------
    print("\n--- Step 3: Parsing supplementary CSVs ---")
    appt_patients, phone_idx = parse_appt_csv(appt_csv)   if appt_csv  else ({}, {})
    drug_patients            = parse_drugs_csv(drugs_csv)  if drugs_csv else {}
    diag_patients            = parse_diag_csv(diag_csv)    if diag_csv  else {}
    exact_icd, prefix_icd, comorb_to_icd = _load_api_mappings(API_CAUSE_CSV)

    print(f"  Appointments:  {len(appt_patients):,} patients")
    print(f"  Medications:   {len(drug_patients):,} patients")
    print(f"  Diagnoses:     {len(diag_patients):,} patients")

    # ------------------------------------------------------------------
    # Step 4 — Single-pass: stream patients CSV → enrich → write Excel
    # ------------------------------------------------------------------
    print(f"\n--- Step 4: Writing enriched Excel ---")
    import openpyxl
    from openpyxl.cell.cell import WriteOnlyCell
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("HCN Consolidated")
    ws.append(TEMPLATE_COLS)   # header row

    appt_written = meds_written = comorb_written = 0
    no_primary_dx_skipped = 0

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            first = row.get('FIRST NAME', '').strip()
            last  = row.get('LAST NAME',  '').strip()
            pid   = row.get('EMR ID', '').strip().split('.')[0]

            # --- NEXT APPT (phone primary, name secondary) ---
            appt_rec = None
            for ph_field in ('HOME PHONE', 'MOBILE PHONE', 'WORK PHONE'):
                ph = _norm_phone(str(row.get(ph_field, '') or ''))
                if ph and ph in phone_idx:
                    appt_rec = appt_patients.get(phone_idx[ph])
                    if appt_rec:
                        break
            if appt_rec is None:
                first_tok = first.lower().split()[0] if first else ''
                appt_rec = appt_patients.get(f"{last.lower()},{first_tok}")
            if appt_rec and appt_rec.get('earliest'):
                row['NEXT APPT'] = appt_rec['earliest'].strftime('%m/%d/%Y')
                appt_written += 1

            # --- MEDICATIONS (EMR ID match) ---
            drug_rec = drug_patients.get(pid)
            if drug_rec:
                drugs = drug_rec['current'] if drug_rec['current'] else drug_rec['all']
                if drugs:
                    row['MEDICATIONS'] = '; '.join(sorted(drugs))
                    meds_written += 1

            # --- COMORBIDITIES / DX ---
            # Primary source: diagnosis PDF (EMR ID match)
            icd_map = diag_patients.get(pid)
            pri = ''
            if icd_map:
                flags = {c: 'NO' for c in _COMORBIDITY_COLUMNS}
                icd_for_comorb = {}
                for icd, desc in icd_map.items():
                    hit = _match_icd(icd, exact_icd, prefix_icd)
                    if not hit and desc:
                        hit = _desc_comorbidity(desc)
                    if hit:
                        flags[hit] = 'YES'
                        icd_for_comorb.setdefault(hit, icd)
                yes_list = [c for c in _COMORBIDITY_COLUMNS if flags[c] == 'YES']
                if yes_list:
                    row.update(flags)
                    pri, sec = _pick_primary_secondary(yes_list)
                    row['PRIMARY DX']    = pri
                    row['SECONDARY DX']  = sec
                    row['PRIMARY ICD']   = icd_for_comorb.get(pri, '')
                    row['SECONDARY ICD'] = icd_for_comorb.get(sec, '')
                    comorb_written += 1

            # Fallback: ICD mapping found only disallowed comorbidities as primary DX
            # → re-infer from medications to find a heart-relevant non-disallowed comorbidity
            if pri in _PRIMARY_DX_DISALLOWED and drug_rec:
                all_drugs = drug_rec['current'] if drug_rec['current'] else drug_rec['all']
                inferred = next(
                    (c for drug in all_drugs
                     for c in [_infer_comorb_from_drug(drug)] if c),
                    ''
                )
                if inferred and inferred not in _PRIMARY_DX_DISALLOWED:
                    row['PRIMARY DX']  = inferred
                    row['PRIMARY ICD'] = comorb_to_icd.get(inferred, '')
                    pri = inferred   # update local for sanity check below

            # --- Filter: skip patients with no PRIMARY DX or a disallowed one ---
            final_pri = row.get('PRIMARY DX', '').strip()
            if not final_pri or final_pri in _PRIMARY_DX_DISALLOWED:
                no_primary_dx_skipped += 1
                continue

            cells = []
            for c in TEMPLATE_COLS:
                val = row.get(c, '')
                if c in _DATE_COLS and val:
                    d = _to_date(val)
                    if isinstance(d, date):
                        cell = WriteOnlyCell(ws, value=d)
                        cell.number_format = 'MM/DD/YYYY'
                        cells.append(cell)
                        continue
                cells.append(val)
            ws.append(cells)

    wb.save(str(xlsx_path))
    csv_path.unlink()

    print(f"  NEXT APPT filled:    {appt_written:,}")
    print(f"  MEDICATIONS filled:  {meds_written:,}")
    print(f"  Comorbidities filled:{comorb_written:,}")
    print(f"  Skipped (no PRIMARY DX): {no_primary_dx_skipped:,}")

    print("  SANITY OK — disallowed PRIMARY DX patients filtered out.")

    print(f"\nFinal output: {xlsx_path}")

    # Clean up supplementary CSVs from cleaned/
    for f in CLEANED_DIR.glob("*.csv"):
        try:
            f.unlink()
        except Exception:
            pass
    print("Cleaned up intermediate CSVs.")


if __name__ == "__main__":
    main()
