"""
HCN Medications updater — CSV-first approach.

Step 1: extract_drugs_csv.py  →  cleaned/drugs_raw_<ts>.csv
Step 2 (this script):
  - Parse raw CSV into per-patient drug records
  - Group by EMR ID (Patient #) → collect distinct current/active drug names
  - Match against latest consolidated Excel by EMR ID
  - Write MEDICATIONS column, save new timestamped Excel to output/

Usage:
    python src/HCN/scripts/update_medications.py
"""

import csv
import re
from datetime import datetime
from pathlib import Path

import openpyxl

MODULE_DIR  = Path(__file__).resolve().parents[1]
CLEANED_DIR = MODULE_DIR / "cleaned"
OUTPUT_DIR  = MODULE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

DIAG_LINE   = re.compile(r'.+\s+\([A-Z0-9][A-Z0-9.\s]+\)\s*$')
PATIENT_ID  = re.compile(r'^\d+(\.\d+)?$')
DATE_STATUS = re.compile(r'^(\d{2}/\d{2}/\d{4})\s+(Current|Lapsed|Discontinued|Void)$')
SKIP_LINES  = {
    'Prescribed Drugs By Diagnosis', 'Diagnosis', 'Drug Description',
    'Patient Name', 'Patient #', 'Total Days', 'Days Left', 'Issued',
    'Status', 'Selections:', 'Report Type:', 'Patient Rx Information',
    'Options:', 'Heart Center of Nevada',
}


def _norm_id(raw: str) -> str:
    """Normalise patient ID: strip .0 suffix."""
    return raw.split('.')[0].strip()


# ---------------------------------------------------------------------------
# Parse raw CSV into per-patient drug records
# ---------------------------------------------------------------------------

def parse_drugs_csv(raw_csv: Path) -> dict:
    """
    Returns dict: emr_id (str) → set of drug description strings (current only)
    We keep ALL statuses but the caller can filter; store as list of dicts.
    Actually store: emr_id → {'current': set, 'all': set}
    """
    patients: dict = {}  # emr_id → {'current': set, 'all': set}

    cur_diag  = ''
    # State machine per drug record: expect drug→name→id→days→daysleft→issued_status
    # We accumulate lines into a small buffer and flush when issued_status detected.
    buf: list = []   # [drug, name, patient_id, total_days, days_left]

    def _flush_buf(status_line: str):
        nonlocal buf
        if len(buf) < 5:
            buf = []
            return
        drug, _name, pid_raw, _tdays, _dleft = buf[:5]
        buf = []
        m = DATE_STATUS.match(status_line)
        if not m:
            return
        status = m.group(2)
        pid = _norm_id(pid_raw)
        if not PATIENT_ID.match(pid_raw.strip()):
            return
        if pid not in patients:
            patients[pid] = {'current': set(), 'all': set()}
        patients[pid]['all'].add(drug)
        if status == 'Current':
            patients[pid]['current'].add(drug)

    with open(raw_csv, newline='', encoding='utf-8') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            raw_text = row['raw_text'].replace('\\n', '\n')
            lines = raw_text.splitlines()
            for line in lines:
                stripped = line.strip()
                if not stripped:
                    continue

                # skip page headers / column headers
                if stripped in SKIP_LINES:
                    continue
                if re.match(r'^\d{2}/\d{2}/\d{4}\d{2}:\d{2} [AP]M$', stripped):
                    continue
                if re.match(r'^Page \d+$', stripped):
                    continue
                if stripped.startswith('PracticeName='):
                    continue

                # diagnosis line (contains ICD code in parens at end)
                if DIAG_LINE.match(stripped) and not DATE_STATUS.match(stripped):
                    cur_diag = stripped
                    buf = []
                    continue

                # issued+status line — flush buffer
                if DATE_STATUS.match(stripped):
                    _flush_buf(stripped)
                    continue

                # accumulate drug record fields into buffer
                buf.append(stripped)
                # only keep last 5 fields (drug, name, id, days, daysleft)
                if len(buf) > 5:
                    buf = buf[-5:]

    return patients


# ---------------------------------------------------------------------------
# Load consolidated Excel and write MEDICATIONS
# ---------------------------------------------------------------------------

def find_latest_consolidated() -> Path:
    files = sorted(OUTPUT_DIR.glob('HCN_consolidated_*.xlsx'), reverse=True)
    if not files:
        raise FileNotFoundError(f"No HCN_consolidated_*.xlsx in {OUTPUT_DIR}")
    return files[0]


def update_excel(patients: dict, src_xlsx: Path) -> Path:
    print(f"Loading {src_xlsx.name} …")
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active

    headers = [c.value for c in ws[1]]

    def col(name):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    emr_id_col   = col('EMR ID')
    meds_col     = col('MEDICATIONS')

    if emr_id_col is None:
        raise ValueError("'EMR ID' column not found in consolidated Excel")

    if meds_col is None:
        meds_col = len(headers) + 1
        ws.cell(row=1, column=meds_col, value='MEDICATIONS')

    matched = unmatched = 0
    total_rows = ws.max_row - 1

    for row_idx in range(2, ws.max_row + 1):
        emr_raw = ws.cell(row_idx, emr_id_col).value
        if not emr_raw:
            continue
        emr_id = str(emr_raw).split('.')[0].strip()

        rec = patients.get(emr_id)
        if rec:
            # prefer current drugs; fall back to all if no current prescriptions
            drugs = rec['current'] if rec['current'] else rec['all']
            if drugs:
                ws.cell(row_idx, meds_col).value = '; '.join(sorted(drugs))
                matched += 1
                continue
        unmatched += 1

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = OUTPUT_DIR / f"HCN_consolidated_{ts}.xlsx"
    wb.save(str(out_path))
    wb.close()

    print(f"Matched: {matched}  |  Unmatched (no drugs): {unmatched}  |  Total: {total_rows}")
    print(f"Saved: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    raw_csvs = sorted(CLEANED_DIR.glob('drugs_raw_*.csv'), reverse=True)
    if not raw_csvs:
        print("No drugs_raw_*.csv found in cleaned/. Run extract_drugs_csv.py first.")
        return

    raw_csv = raw_csvs[0]
    print(f"Parsing raw CSV: {raw_csv.name} …")
    patients = parse_drugs_csv(raw_csv)
    print(f"Parsed {len(patients)} unique patients from drugs PDF")

    src_xlsx = find_latest_consolidated()
    update_excel(patients, src_xlsx)


if __name__ == '__main__':
    main()
