"""
HCN Appointment updater — CSV-first approach.

Step 1: extract_appt_csv.py  →  cleaned/appt_raw_<ts>.csv
Step 2 (this script):
  - Parse the raw CSV into per-patient records (name, phones, earliest appt date)
  - Match against latest consolidated Excel by normalized name (fallback: phone)
  - Write NEXT APPT column, save new timestamped Excel to output/

Usage:
    python src/HCN/scripts/update_appointments.py
"""

import csv
import re
from datetime import datetime, date
from pathlib import Path

import openpyxl

MODULE_DIR = Path(__file__).resolve().parents[1]
CLEANED_DIR = MODULE_DIR / "cleaned"
OUTPUT_DIR  = MODULE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

TODAY = date.today()

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _norm_name(raw: str) -> str:
    """LASTNAME, FIRSTNAME [MIDDLE] → lowercase stripped, drop middle initial."""
    raw = re.sub(r'\s*\(\d+y\s+[MF]\)', '', raw).strip()  # strip (52y M)
    parts = [p.strip() for p in raw.split(',', 1)]
    if len(parts) == 2:
        last = parts[0].lower().strip()
        first_parts = parts[1].lower().split()
        first = first_parts[0] if first_parts else ''
        return f"{last},{first}"
    return raw.lower().strip()


def _norm_phone(ph: str) -> str:
    digits = re.sub(r'\D', '', ph)
    return digits[-10:] if len(digits) >= 10 else digits


def _parse_date(s: str):
    """Return date or None."""
    s = s.strip()
    try:
        return datetime.strptime(s, '%m/%d/%Y').date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Step 1 — parse raw CSV pages into patient appointment records
# ---------------------------------------------------------------------------

PATIENT_HDR = re.compile(r'^([A-Z][A-Z ,.\'-]+?)\s{2,}\(\d+y\s+[MF]\)')
DATE_LINE   = re.compile(r'^\d{2}/\d{2}/\d{4}$')
SKIP_HDRS   = {'Appt Date', 'Time', 'Len', 'Prov', 'Location', 'Equipment',
               'Reason', 'Room', 'Encounter #', 'Appointment Report By Patient',
               'Heart Center of Nevada', 'Selections:', 'Appointment Dates:'}


def parse_raw_csv(raw_csv: Path) -> dict:
    """
    Returns dict: norm_name → {name, phones: set, earliest: date | None}
    """
    patients: dict = {}  # norm_name → record
    phone_index: dict = {}  # norm_phone → norm_name  (for fallback)

    cur_name = None
    cur_norm = None
    cur_phones: set = set()
    cur_dates: list = []
    in_phones = False

    def _flush():
        nonlocal cur_name, cur_norm, cur_phones, cur_dates, in_phones
        if cur_norm is None:
            return
        future = [d for d in cur_dates if d >= TODAY]
        earliest = min(future) if future else None
        if cur_norm not in patients:
            patients[cur_norm] = {
                'name':     cur_name,
                'norm':     cur_norm,
                'phones':   set(cur_phones),
                'earliest': earliest,
            }
        else:
            existing = patients[cur_norm]
            existing['phones'].update(cur_phones)
            if earliest:
                if existing['earliest'] is None or earliest < existing['earliest']:
                    existing['earliest'] = earliest
        for ph in cur_phones:
            phone_index.setdefault(ph, cur_norm)
        cur_name = cur_norm = None
        cur_phones = set()
        cur_dates = []
        in_phones = False

    with open(raw_csv, newline='', encoding='utf-8') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            raw_text = row['raw_text'].replace('\\n', '\n')
            lines = raw_text.splitlines()
            for line in lines:
                stripped = line.strip()
                if not stripped:
                    continue

                # skip page headers
                if any(stripped.startswith(h) for h in SKIP_HDRS):
                    in_phones = False
                    continue
                if re.match(r'^\d{2}/\d{2}/\d{4}\d{2}:\d{2} [AP]M$', stripped):
                    continue  # report timestamp line
                if re.match(r'^Page \d+$', stripped):
                    continue
                if stripped.startswith('From:') or stripped.startswith('To:'):
                    continue

                # patient header?
                m = PATIENT_HDR.match(stripped)
                if m:
                    _flush()
                    cur_name = m.group(1).strip()
                    cur_norm = _norm_name(stripped)
                    in_phones = False
                    continue

                if cur_norm is None:
                    continue

                # phone block
                if stripped == 'Phone Numbers:':
                    in_phones = True
                    continue
                if in_phones:
                    for prefix in ('Home:', 'Work:', 'Mobile:'):
                        if stripped.startswith(prefix):
                            ph_raw = stripped[len(prefix):].strip().rstrip('*').strip()
                            if ph_raw:
                                ph = _norm_phone(ph_raw)
                                if ph:
                                    cur_phones.add(ph)
                            break
                    else:
                        # bare phone line (asterisk default)
                        ph_raw = stripped.rstrip('*').strip()
                        if re.match(r'^\(?\d', ph_raw):
                            ph = _norm_phone(ph_raw)
                            if ph:
                                cur_phones.add(ph)
                        else:
                            in_phones = False

                # appointment date line
                if DATE_LINE.match(stripped):
                    d = _parse_date(stripped)
                    if d:
                        cur_dates.append(d)

    _flush()
    return patients, phone_index


# ---------------------------------------------------------------------------
# Step 2 — load consolidated Excel and write NEXT APPT
# ---------------------------------------------------------------------------

def find_latest_consolidated() -> Path:
    files = sorted(OUTPUT_DIR.glob('HCN_consolidated_*.xlsx'), reverse=True)
    if not files:
        raise FileNotFoundError(f"No HCN_consolidated_*.xlsx in {OUTPUT_DIR}")
    return files[0]


def update_excel(patients: dict, phone_index: dict, src_xlsx: Path) -> Path:
    print(f"Loading {src_xlsx.name} …")
    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb.active

    # find column indices (1-based)
    headers = [c.value for c in ws[1]]
    def col(name):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    first_name_col  = col('FIRST NAME')
    last_name_col   = col('LAST NAME')
    home_phone_col  = col('HOME PHONE')
    cell_phone_col  = col('CELL PHONE')
    work_phone_col  = col('WORK PHONE')
    next_appt_col   = col('NEXT APPT')

    if next_appt_col is None:
        next_appt_col = len(headers) + 1
        ws.cell(row=1, column=next_appt_col, value='NEXT APPT')

    matched = unmatched = phone_matched = name_matched = 0
    total_rows = ws.max_row - 1

    for row_idx in range(2, ws.max_row + 1):
        first = (ws.cell(row_idx, first_name_col).value or '').strip()
        last  = (ws.cell(row_idx, last_name_col).value  or '').strip()
        if not first and not last:
            continue

        rec = None

        # --- PRIMARY: phone match (any of home / cell / work) ---
        for ph_col in (home_phone_col, cell_phone_col, work_phone_col):
            if ph_col is None:
                continue
            ph_raw = ws.cell(row_idx, ph_col).value or ''
            ph = _norm_phone(str(ph_raw))
            if ph and ph in phone_index:
                rec = patients.get(phone_index[ph])
                if rec:
                    phone_matched += 1
                    break

        # --- SECONDARY: exact normalised name match ---
        if rec is None:
            # build norm key: "lastname,firstname_first_token" — exact only
            first_token = first.lower().split()[0] if first else ''
            norm = f"{last.lower()},{first_token}"
            rec = patients.get(norm)
            if rec:
                name_matched += 1

        if rec and rec['earliest']:
            ws.cell(row_idx, next_appt_col).value = rec['earliest'].strftime('%m/%d/%Y')
            matched += 1
        else:
            unmatched += 1

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = OUTPUT_DIR / f"HCN_consolidated_{ts}.xlsx"
    wb.save(str(out_path))
    wb.close()

    print(f"Matched: {matched} (phone={phone_matched}, name={name_matched})  |  Unmatched: {unmatched}  |  Total: {total_rows}")
    print(f"Saved: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    # find latest raw CSV
    raw_csvs = sorted(CLEANED_DIR.glob('appt_raw_*.csv'), reverse=True)
    if not raw_csvs:
        print("No appt_raw_*.csv found in cleaned/. Run extract_appt_csv.py first.")
        return

    raw_csv = raw_csvs[0]
    print(f"Parsing raw CSV: {raw_csv.name} …")
    patients, phone_index = parse_raw_csv(raw_csv)
    print(f"Parsed {len(patients)} unique patients from appointment PDF")

    src_xlsx = find_latest_consolidated()
    update_excel(patients, phone_index, src_xlsx)


if __name__ == '__main__':
    main()
