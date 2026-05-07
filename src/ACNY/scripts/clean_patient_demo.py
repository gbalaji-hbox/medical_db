"""
Cleans PatientDemo_2Line.xlsx (both Sheet1 and Sheet2), merges phone data
from PatientManagementProfileDetail.xlsx, visit/insurance data from all
CustomVisitListing_MM_YYYY.xlsx files, primary care provider from
PatientManagementAnalysisOfVisitsDetail.xlsx, insurance type from
CarrierCode.xlsx, and comorbidity/ICD mapping from api_prescriptioncauselist CSV.
"""

import sys
import os
import csv
import glob
import json
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", ".."))

EXCEL_EPOCH = datetime(1899, 12, 30)
SKIP_ROWS_DEMO     = 4
SKIP_ROWS_PROFILE  = 5
SKIP_ROWS_VISIT    = 10
SKIP_ROWS_ANALYSIS = 8
SKIP_ROWS_CARRIER  = 5
SHEETS_DEMO = ["Sheet1", "Sheet2"]
CLINIC_FACILITY = "ALLERGY AND ASTHMA ASSOC. OF MURRAY HILL"

# CustomVisitListing column indices (0-based)
CV_CHART    = 3
CV_RACE     = 13
CV_DOS      = 16
CV_DX1      = 23
CV_DX2      = 24
CV_DX3      = 25
CV_DX4      = 26
CV_INS      = 32
CV_INS_ID   = 33
CV_COPAY    = 34
CV_GRP_CODE = 38
CV_APT_DT   = 45
CV_LAST_APT = 48

# Clinic-specific primary DX codes (allergy/asthma clinic)
# Each entry: (icd_pattern, description)
CLINIC_PRIMARY_DX = [
    ("J45",             "Asthma"),
    ("J31.0",           "Chronic Rhinitis"),
    ("R05.3",           "Chronic Cough"),
    ("L20",             "Atopic Dermatitis"),
    ("D47.02",          "Systemic Mastocytosis"),
    ("D84.1",           "Complement Deficiency"),
    ("Z91.010/013/018", "Food Allergies"),
    ("L50.1",           "Chronic Idiopathic Urticaria"),
]

# Canonical ICD code per allergy comorbidity column.
# Used as PRIMARY ICD when a description-based match fires but no raw patient ICD is available.
CLINIC_PRIMARY_DX_CANONICAL = {
    "ASTHMA":                       "J45",
    "CHRONIC RHINITIS":             "J31.0",
    "CHRONIC COUGH":                "R05.3",
    "ATOPIC DERMATITIS":            "L20",
    "SYSTEMIC MASTOCYTOSIS":        "D47.02",
    "COMPLEMENT DEFICIENCY":        "D84.1",
    "FOOD ALLERGIES":               "Z91.010",
    "CHRONIC IDIOPATHIC URTICARIA": "L50.1",
}

# Conservative, ICD-10-description-precise keywords for allergy comorbidity fallback.
# Only used when no direct ICD prefix match is found in comorbidity_icd_map.
# Keywords are matched against the ICD description text (lowercased, exact phrase).
ACNY_DESCRIPTION_KEYWORDS = {
    "ASTHMA": [
        "asthma", "bronchial asthma", "allergic asthma",
        "exercise induced asthma", "cough variant asthma", "asthmatic bronchitis",
    ],
    "CHRONIC RHINITIS": [
        "chronic rhinitis", "chronic atrophic rhinitis",
        "chronic hypertrophic rhinitis", "vasomotor rhinitis",
    ],
    "CHRONIC COUGH": [
        "chronic cough",
    ],
    "ATOPIC DERMATITIS": [
        "atopic dermatitis", "atopic neurodermatitis",
        "infantile eczema", "intrinsic (allergic) eczema",
    ],
    "SYSTEMIC MASTOCYTOSIS": [
        "systemic mastocytosis", "aggressive systemic mastocytosis",
        "indolent systemic mastocytosis",
    ],
    "COMPLEMENT DEFICIENCY": [
        "complement deficiency", "deficiency of complement",
        "hereditary angioedema with c1-inhibitor deficiency",
    ],
    "FOOD ALLERGIES": [
        "allergy to peanuts", "allergy to tree nuts",
        "allergy to other foods", "food allergy", "food hypersensitivity",
    ],
    "CHRONIC IDIOPATHIC URTICARIA": [
        "chronic idiopathic urticaria",
    ],
}

# Map CSV cause names → template comorbidity column names
CAUSE_TO_COLUMN = {
    "Asthma":                       "ASTHMA",
    "Bronchiectasis":               "BRONCHIECTASIS",
    "Coronary Artery Disease":      "CORONARY ARTERY DISEASE",
    "Arrhythmia":                   "ARRHYTHMIA",
    "Hypertension":                 "HYPERTENSION",
    "Obesity":                      "OBESITY",
    "COPD":                         "COPD",
    "Sleep Apnea":                  "SLEEP APNEA",
    "Peripheral Vascular Disease":  "PERIPHERAL VASCULAR",
    "Cerebrovascular Accident":     "CERBOVASCULAR ACCIDENT",
    "Valvular Heart Disease":       "VALVULAR HEART",
    "Angina Pectoris":              "ANGINA PECTORIS",
    "Hyperlipidemia":               "HYPERLIPIDEMIA",
    "Emphysema ":                   "EMPHYSEMA",   # trailing space in CSV
    "Emphysema":                    "EMPHYSEMA",
    "Hypoxemia":                    "HYPOXEMIA",
    "CHF (Congestive Heart Failure)": "CONGESTIVE HEART FAILURE",
    "Hypotension":                  "HYPOTENSION",
    "Dyspnea":                      "DYSPNEA",
    "Chronic Kidney":               "CHRONIC KIDNEY DISEASE",
    "Type 2 Diabetes":              "DIABETES",
    "Respiratory Failure":          "RESPIRATORY FAILURE",
}

COMORBIDITY_COLUMNS = [
    # Allergy-specific (clinic focus) — checked first for PRIMARY DX left-to-right
    "ASTHMA",
    "CHRONIC RHINITIS", "CHRONIC COUGH", "ATOPIC DERMATITIS",
    "SYSTEMIC MASTOCYTOSIS", "COMPLEMENT DEFICIENCY", "FOOD ALLERGIES",
    "CHRONIC IDIOPATHIC URTICARIA",
    # General comorbidities
    "CORONARY ARTERY DISEASE", "ARRHYTHMIA", "CONGESTIVE HEART FAILURE",
    "PERIPHERAL VASCULAR", "VALVULAR HEART", "CERBOVASCULAR ACCIDENT",
    "HYPERLIPIDEMIA", "ANGINA PECTORIS", "HYPOTENSION", "HYPERTENSION",
    "OBESITY", "DIABETES", "CHRONIC KIDNEY DISEASE", "COPD",
    "RESPIRATORY FAILURE", "SLEEP APNEA", "DYSPNEA", "EMPHYSEMA",
    "BRONCHIECTASIS", "HYPOXEMIA",
]

DATE_COLUMNS = ["DATE OF BIRTH", "LAST SEEN DATE", "NEXT APPT"]

TEMPLATE_COLUMNS = [
    "EMR ID", "PATIENT EMR NAME", "FIRST NAME", "MIDDLE NAME", "LAST NAME",
    "PATIENT FULL NAME", "DATE OF BIRTH", "GENDER", "STREET ADDRESS", "CITY",
    "STATE", "ZIP", "HOME PHONE", "MOBILE PHONE", "WORK PHONE", "EMAIL ADDRESS",
    "LANGUAGE", "RACE", "EMERGENCY CONTACT NAME", "EMERGENCY RELATIONSHIP",
    "EMERGENCY CONTACT HOME PHONE", "EMERGENCY CONTACT MOBILE PHONE",
    "MEDICARE ID", "PRIMARY INSURANCE", "PRIMARY ID", "PRIMARY GROUP",
    "SECONDARY INSURANCE", "SECONDARY ID", "SECONDARY GROUP",
    "TERITARY INSURANCE", "TERITARY ID", "TERITARY GROUP",
    "INSURANCE TYPE", "CO-PAY",
    # Allergy-specific comorbidities first (clinic focus)
    "ASTHMA",
    "CHRONIC RHINITIS", "CHRONIC COUGH", "ATOPIC DERMATITIS",
    "SYSTEMIC MASTOCYTOSIS", "COMPLEMENT DEFICIENCY", "FOOD ALLERGIES",
    "CHRONIC IDIOPATHIC URTICARIA",
    # General comorbidities
    "CORONARY ARTERY DISEASE", "ARRHYTHMIA", "CONGESTIVE HEART FAILURE",
    "PERIPHERAL VASCULAR", "VALVULAR HEART", "CERBOVASCULAR ACCIDENT",
    "HYPERLIPIDEMIA", "ANGINA PECTORIS", "HYPOTENSION", "HYPERTENSION",
    "OBESITY", "DIABETES", "CHRONIC KIDNEY DISEASE", "COPD",
    "RESPIRATORY FAILURE", "SLEEP APNEA", "DYSPNEA", "EMPHYSEMA",
    "BRONCHIECTASIS", "HYPOXEMIA",
    "PRIMARY DX", "SECONDARY DX", "PRIMARY ICD", "SECONDARY ICD",
    "LAST SEEN DATE", "NEXT APPT", "PROVIDER DATA", "PROVIDER NAME",
    "CLINIC FACILITY", "PRIMARY CARE PROVIDER", "MEDICATIONS", "ENCOUNTER NOTES",
]


# ---------------------------------------------------------------------------
# ICD matching
# ---------------------------------------------------------------------------

def matches_icd_pattern(patient_code, pattern):
    """
    Returns True if patient_code matches an ICD pattern.
      J45.xx / L20.xx  → prefix match stripping .xx
      Z91.010/013/018  → match any of Z91.010, Z91.013, Z91.018
      J44.9            → prefix match (J44.9, J44.90, etc.)
    """
    code = patient_code.upper().strip()
    pat  = pattern.strip()

    if pat.upper().endswith(".XX"):
        return code.startswith(pat[:-3].upper())

    if "/" in pat:
        dot_idx = pat.index(".")
        base    = pat[:dot_idx].upper()
        parts   = pat[dot_idx + 1:].split("/")
        targets = [f"{base}.{p.strip().upper()}" for p in parts]
        return any(code.startswith(t) for t in targets)

    return code.startswith(pat.upper())


def is_clinic_primary_dx(icd_code):
    """Returns True if the ICD code matches any CLINIC_PRIMARY_DX pattern."""
    return any(matches_icd_pattern(icd_code, pat) for pat, _ in CLINIC_PRIMARY_DX)


def _icd_matches_prefix(patient_code, prefix):
    """
    Normalizes both sides (uppercase, strip dots) then checks bidirectionally.
    Bidirectional covers:
      - Patient has more specific code: "I10.0" (norm "I100") vs prefix "I100" → startswith ✓
      - Patient has less specific code: "I10"   (norm "I10")  vs prefix "I100" → reverse ✓
      - Allergy short prefixes:         "J45.20" (norm "J4520") vs prefix "J45" → startswith ✓
    """
    norm_code   = patient_code.upper().replace(".", "")
    norm_prefix = prefix.upper().replace(".", "")
    return norm_code.startswith(norm_prefix) or norm_prefix.startswith(norm_code)


def classify_icd(icd_set, comorbidity_icd_map):
    """
    Returns:
      comorbidities          : dict col → 'YES'/'NO'
      comorbidity_to_raw_icd : dict col → first raw patient ICD that triggered YES
    """
    comorbidities          = {col: "NO" for col in COMORBIDITY_COLUMNS}
    comorbidity_to_raw_icd = {}

    for col, prefixes in comorbidity_icd_map.items():
        for code in sorted(icd_set):           # sorted for determinism
            if any(_icd_matches_prefix(code, p) for p in prefixes):
                comorbidities[col] = "YES"
                comorbidity_to_raw_icd[col] = code
                break

    return comorbidities, comorbidity_to_raw_icd


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def excel_serial_to_date(serial):
    if serial is None or serial == "":
        return ""
    if isinstance(serial, datetime):
        return serial.strftime("%m-%d-%Y")
    try:
        n = int(float(serial))
        if n <= 0:
            return ""
        return (EXCEL_EPOCH + timedelta(days=n)).strftime("%m-%d-%Y")
    except (ValueError, TypeError, OverflowError):
        try:
            return pd.to_datetime(str(serial)).strftime("%m-%d-%Y")
        except Exception:
            return str(serial)


def excel_serial_to_dt(serial):
    if serial is None or serial == "":
        return None
    if isinstance(serial, datetime):
        return serial
    try:
        n = float(serial)
        if n > 0:
            return EXCEL_EPOCH + timedelta(days=n)
    except (ValueError, TypeError):
        pass
    return None


def parse_date_to_dt(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        n = float(s)
        if n > 0:
            return EXCEL_EPOCH + timedelta(days=n)
    except ValueError:
        pass
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def clean_text(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("\n", "\\n") else s


def normalize_chart(val):
    return clean_text(val)


def parse_name(raw):
    raw = raw.strip()
    if "," in raw:
        last, rest = raw.split(",", 1)
        last  = last.strip().title()
        rest  = rest.strip()
        parts = rest.split(None, 1)
        first  = parts[0].title() if parts else ""
        middle = parts[1].title() if len(parts) > 1 else ""
    else:
        last, first, middle = "", raw.title(), ""
    full_name = " ".join(p for p in [first, middle, last] if p)
    return last, first, middle, full_name, raw


def parse_provider_name(raw):
    s = raw.strip().rstrip(",").strip()
    if not s:
        return ""
    if "," in s:
        last, rest = s.split(",", 1)
        last  = last.strip().title()
        given = " ".join(p.title() for p in rest.strip().split())
        return " ".join(p for p in [given, last] if p)
    return s.title()


def safe_col(row, idx):
    return clean_text(row[idx]) if len(row) > idx else ""


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_icd_descriptions(profitability_path, diagnosis_path):
    """
    Builds a dict: icd_code_upper → description
    from Profitability.xlsx (cols 31-38 pairs) and
    DiagnosisCodeProductionDetail.xlsx (cols 0-1).
    Used as a fallback description lookup for any ICD code.
    """
    icd_desc = {}

    # --- Profitability: skip 2 rows, headers on row 3, data from row 4 ---
    wb = load_workbook(profitability_path, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        for code_idx, desc_idx in [(31, 32), (33, 34), (35, 36), (37, 38)]:
            code = clean_text(row[code_idx]) if len(row) > code_idx else ""
            desc = clean_text(row[desc_idx]) if len(row) > desc_idx else ""
            if code and desc and code.upper() not in icd_desc:
                icd_desc[code.upper()] = desc.title()
    wb.close()

    # --- DiagnosisCodeProductionDetail: skip 3 rows, data from row 4 ---
    wb2 = load_workbook(diagnosis_path, read_only=True, data_only=True)
    ws2 = wb2.active
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i < 3:
            continue
        code = clean_text(row[0]).lstrip("*") if len(row) > 0 else ""
        desc = clean_text(row[1]) if len(row) > 1 else ""
        if code and desc and code.upper() not in icd_desc:
            icd_desc[code.upper()] = desc.title()
    wb2.close()

    return icd_desc


def load_allergy_icd_map(template_dir):
    """
    Loads acny_allergy_icd_map.json — the one-time pre-built mapping of allergy
    ICD codes actually present in ACNY source data.
    Returns:
      icd_prefixes  : dict  column → [icd_prefix_strings]
      desc_keywords : dict  column → [lowercase description keyword phrases]
      canonical_icd : dict  column → canonical_icd_string
    """
    json_path = os.path.join(template_dir, "acny_allergy_icd_map.json")
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    icd_prefixes  = {}
    desc_keywords = {}
    canonical_icd = {}

    for col, entry in data.items():
        if col.startswith("_"):
            continue
        icd_prefixes[col]  = entry.get("icd_prefixes", [])
        desc_keywords[col] = entry.get("description_keywords", [])
        canonical_icd[col] = entry.get("canonical_icd", "")

    return icd_prefixes, desc_keywords, canonical_icd


def parse_prescription_csv(fpath, allergy_icd_prefixes):
    """
    Loads api_prescriptioncauselist CSV for general comorbidities.
    Allergy-specific ICD prefixes come from acny_allergy_icd_map.json (pre-built from raw data).
    Returns dict: template_column_name → [icd_prefix_patterns]
    """
    comorbidity_icd_map = {}
    with open(fpath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cause  = row.get("cause", "").strip()
            icd    = row.get("icd_code", "").strip()
            active = row.get("active", "").strip().upper()
            if active != "TRUE" or not icd:
                continue
            col = CAUSE_TO_COLUMN.get(cause) or CAUSE_TO_COLUMN.get(cause.strip())
            if col:
                comorbidity_icd_map.setdefault(col, []).append(icd)

    # Override/add allergy-specific ICD prefixes from JSON (data-driven, not hardcoded)
    for col, prefixes in allergy_icd_prefixes.items():
        if prefixes:
            comorbidity_icd_map[col] = prefixes

    return comorbidity_icd_map


def parse_demo_sheet(ws):
    records = []
    row_num = 0
    pending = None

    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num <= SKIP_ROWS_DEMO:
            continue
        if all((v is None or str(v).strip() == "") for v in row):
            continue

        if pending is None:
            pending = {
                "_raw_patient": clean_text(row[0]),
                "chart_no":     normalize_chart(row[2]),
                "address":      clean_text(row[3]),
                "apt_ste":      clean_text(row[4]),
                "city":         clean_text(row[5]),
                "state":        clean_text(row[6]),
                "zipcode":      clean_text(row[7]),
            }
        else:
            pending.update({
                "provider":        clean_text(row[2]),
                "preferred_phone": clean_text(row[3]),
                "email_address":   clean_text(row[4]),
                "language":        clean_text(row[5]),
                "sex":             clean_text(row[6]),
                "birthdate":       excel_serial_to_dt(row[7]),
            })
            records.append(pending)
            pending = None

    return records


def parse_profile(ws):
    phone_map = {}
    row_num   = 0

    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num <= SKIP_ROWS_PROFILE:
            continue
        patient_name = clean_text(row[1]) if len(row) > 1 else ""
        if not patient_name:
            continue

        chart_no  = normalize_chart(row[0])
        home_ph   = clean_text(row[3]) if len(row) > 3 else ""
        office_ph = clean_text(row[4]) if len(row) > 4 else ""

        if chart_no not in phone_map:
            phone_map[chart_no] = {"home_phone": home_ph, "office_phone": office_ph}
        else:
            if not phone_map[chart_no]["home_phone"] and home_ph:
                phone_map[chart_no]["home_phone"] = home_ph
            if not phone_map[chart_no]["office_phone"] and office_ph:
                phone_map[chart_no]["office_phone"] = office_ph

    return phone_map


def parse_visit_files(base_dir):
    pattern = os.path.join(base_dir, "CustomVisitListing_*.xlsx")
    files   = sorted(glob.glob(pattern))
    print(f"  Found {len(files)} CustomVisitListing files.")

    visit_map = {}

    for fpath in files:
        fname = os.path.basename(fpath)
        wb    = load_workbook(fpath, read_only=True, data_only=True)
        ws    = wb.active
        row_num   = 0
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            row_num += 1
            if row_num <= SKIP_ROWS_VISIT:
                continue
            if len(row) <= CV_CHART:
                continue

            chart_no = normalize_chart(row[CV_CHART])
            if not chart_no:
                continue

            dos_dt      = parse_date_to_dt(safe_col(row, CV_DOS))
            last_apt_dt = parse_date_to_dt(safe_col(row, CV_LAST_APT))
            appt_dt     = excel_serial_to_dt(row[CV_APT_DT] if len(row) > CV_APT_DT else None)

            candidates    = [d for d in [dos_dt, last_apt_dt] if d is not None]
            best_visit_dt = max(candidates) if candidates else None

            ins      = safe_col(row, CV_INS)
            ins_id   = safe_col(row, CV_INS_ID)
            copay    = safe_col(row, CV_COPAY)
            grp_code = safe_col(row, CV_GRP_CODE)
            race     = safe_col(row, CV_RACE)

            dx_codes = {safe_col(row, CV_DX1), safe_col(row, CV_DX2),
                        safe_col(row, CV_DX3), safe_col(row, CV_DX4)}
            dx_codes.discard("")
            row_count += 1

            if chart_no not in visit_map:
                visit_map[chart_no] = {
                    "visit_dt":  best_visit_dt,
                    "ins":       ins,
                    "ins_id":    ins_id,
                    "copay":     copay,
                    "grp_code":  grp_code,
                    "appt_dt":   appt_dt if (appt_dt and appt_dt.date() >= datetime.now().date()) else None,
                    "race":      race,
                    "icd_codes": dx_codes,   # kept as set for matching
                }
            else:
                entry = visit_map[chart_no]
                if best_visit_dt and (entry["visit_dt"] is None or best_visit_dt > entry["visit_dt"]):
                    entry["visit_dt"] = best_visit_dt
                    if ins:
                        entry["ins"]      = ins
                        entry["ins_id"]   = ins_id
                        entry["copay"]    = copay
                        entry["grp_code"] = grp_code
                # Only consider future or today appointments
                if appt_dt and appt_dt.date() >= datetime.now().date():
                    if entry["appt_dt"] is None or appt_dt > entry["appt_dt"]:
                        entry["appt_dt"] = appt_dt
                if not entry["race"] and race:
                    entry["race"] = race
                entry["icd_codes"].update(dx_codes)

        wb.close()
        print(f"    {fname}: {row_count} data rows")

    result = {}
    for chart_no, entry in visit_map.items():
        result[chart_no] = {
            "last_seen_date": entry["visit_dt"] if entry["visit_dt"] else None,
            "next_appt":      entry["appt_dt"] if entry["appt_dt"] else None,
            "insurance":      entry["ins"],
            "insurance_id":   entry["ins_id"],
            "copay":          entry["copay"],
            "group_code":     entry["grp_code"],
            "race":           entry["race"],
            "icd_codes":      entry["icd_codes"],   # set — used for matching
        }
    return result


def parse_carrier_codes(fpath):
    carrier_map  = {}
    wb           = load_workbook(fpath, read_only=True, data_only=True)
    ws           = wb.active
    row_num      = 0
    pending_name = None

    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num <= SKIP_ROWS_CARRIER:
            continue
        if all((v is None or str(v).strip() == "") for v in row):
            continue

        carrier_name = clean_text(row[3]) if len(row) > 3 else ""
        active       = clean_text(row[9]) if len(row) > 9 else ""
        ins_type     = clean_text(row[6]) if len(row) > 6 else ""

        if carrier_name and active.lower() == "yes":
            pending_name = carrier_name.upper()
        elif pending_name is not None:
            if ins_type:
                carrier_map[pending_name] = ins_type
            pending_name = None

    wb.close()
    return carrier_map


def parse_qualified_registry(fpath):
    """
    Reads QualifiedRegistry.xlsx (skip 2 header rows, data from row 3).
    Columns: TIN(0), NPI(1), FirstName(2), LastName(3), PatientID(4),
             Gender(5), DOB(6), DOS(7), PrimaryInsurer(8), SecInsurer(9),
             CPT(10), Modifier(11), ICD(12), PlaceOfService(13).
    Returns dict: chart_no → {"sec_insurer": str, "icd_codes": set}
    """
    registry = {}
    wb = load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:          # rows 0-2 are title / clinic / headers
            continue
        if not row or len(row) <= 12:
            continue

        chart_no   = normalize_chart(row[4])
        sec_insurer = clean_text(row[9])
        icd         = clean_text(row[12]).lstrip("*").strip()

        if not chart_no:
            continue

        if chart_no not in registry:
            registry[chart_no] = {"sec_insurer": "", "icd_codes": set()}

        entry = registry[chart_no]
        if sec_insurer and not entry["sec_insurer"]:
            entry["sec_insurer"] = sec_insurer
        if icd:
            entry["icd_codes"].add(icd)

    wb.close()
    return registry


def parse_future_appointments(fpath):
    """
    Reads AppointmentConfirmationDetail _Future.xlsx (119 sheets, one per day).
    Per sheet:
      row 3 (0-indexed): col[0] = "Appointment Date: MM/DD/YYYY"
      row 4            : column headers
      row 5+           : data — col[2] = Chart#
    Returns dict: chart_no → earliest future appointment date string MM-DD-YYYY.
    """
    future_map = {}
    today = datetime.now().date()

    wb = load_workbook(fpath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 5:
            continue

        # Row index 3 = "Appointment Date: MM/DD/YYYY" in col[0]
        date_cell = clean_text(rows[3][0]) if rows[3] else ""
        appt_date = None
        if date_cell.lower().startswith("appointment date:"):
            date_str = date_cell.split(":", 1)[1].strip()
            try:
                appt_date = datetime.strptime(date_str, "%m/%d/%Y").date()
            except ValueError:
                pass

        if appt_date is None or appt_date < today:
            continue

        # Rows 5+ are data (index 5 onward)
        for row in rows[5:]:
            if not row or len(row) <= 2:
                continue
            chart_no = normalize_chart(row[2])
            if not chart_no:
                continue

            if chart_no not in future_map or appt_date < future_map[chart_no]:
                future_map[chart_no] = appt_date

    wb.close()
    return {k: datetime(v.year, v.month, v.day) for k, v in future_map.items()}


def parse_analysis_of_visits(fpath):
    pcp_map       = {}
    wb            = load_workbook(fpath, read_only=True, data_only=True)
    ws            = wb.active
    row_num       = 0
    current_chart = None

    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num <= SKIP_ROWS_ANALYSIS:
            continue
        if all((v is None or str(v).strip() == "") for v in row):
            continue

        col1 = clean_text(row[1]) if len(row) > 1 else ""
        col2 = clean_text(row[2]) if len(row) > 2 else ""
        col6 = clean_text(row[6]) if len(row) > 6 else ""

        if col1 and "year" in col2.lower():
            current_chart = normalize_chart(row[1])
            continue
        if not col1:
            continue
        if current_chart and col6 and current_chart not in pcp_map:
            pcp_map[current_chart] = parse_provider_name(col6)

    wb.close()
    return pcp_map


# ---------------------------------------------------------------------------
# Consolidation
# ---------------------------------------------------------------------------

def build_consolidated(all_records, phone_map, visit_map, carrier_map, pcp_map,
                       comorbidity_icd_map, icd_desc, future_appt_map=None,
                       registry_map=None, allergy_desc_keywords=None,
                       allergy_canonical_icd=None):
    rows = []
    for r in all_records:
        last, first, middle, full_name, emr_name = parse_name(r["_raw_patient"])

        street = (f"{r['apt_ste']}, {r['address']}" if r["apt_ste"] else r["address"])

        existing_home  = r["preferred_phone"]
        profile        = phone_map.get(r["chart_no"], {})
        profile_home   = profile.get("home_phone", "")
        profile_office = profile.get("office_phone", "")

        if not existing_home:
            home_phone, mobile_phone = profile_home, ""
        elif not profile_home or profile_home == existing_home:
            home_phone, mobile_phone = existing_home, ""
        else:
            home_phone, mobile_phone = existing_home, profile_home

        visit    = visit_map.get(r["chart_no"], {})
        icd_set  = set(visit.get("icd_codes", set()))

        registry = (registry_map or {}).get(r["chart_no"], {})
        sec_insurer_name = registry.get("sec_insurer", "")
        # Supplement ICD codes from registry if patient has none from visits
        if not icd_set and registry.get("icd_codes"):
            icd_set = set(registry["icd_codes"])

        # --- ICD classification ---
        if icd_set:
            comorbidities, comorbidity_to_raw_icd = classify_icd(icd_set, comorbidity_icd_map)
        else:
            comorbidities          = {col: "" for col in COMORBIDITY_COLUMNS}
            comorbidity_to_raw_icd = {}

        # YES comorbidities in COMORBIDITY_COLUMNS order (allergy columns are first,
        # so left-to-right naturally prioritises allergy diagnoses for this clinic)
        yes_conditions = [col for col in COMORBIDITY_COLUMNS if comorbidities.get(col) == "YES"]

        # PRIMARY DX / ICD — three-tier:
        #   Tier 1: First YES comorbidity whose triggering raw ICD matches CLINIC_PRIMARY_DX
        #   Tier 2: First YES comorbidity left-to-right (allergy columns lead the list)
        #   Tier 3: ICD description keyword match against ACNY_DESCRIPTION_KEYWORDS
        #           (only fires when no ICD code mapped to any comorbidity column)
        primary_dx = primary_icd = ""

        # Tier 1 — prefer a YES whose raw ICD is an allergy/asthma code
        for col in yes_conditions:
            raw_icd = comorbidity_to_raw_icd.get(col, "")
            if raw_icd and is_clinic_primary_dx(raw_icd):
                primary_dx  = col
                primary_icd = raw_icd
                break

        _canonical = allergy_canonical_icd or CLINIC_PRIMARY_DX_CANONICAL
        _desc_kws  = allergy_desc_keywords  or ACNY_DESCRIPTION_KEYWORDS

        # Tier 2 — first YES in column order (ASTHMA and allergy columns are leftmost)
        if not primary_dx and yes_conditions:
            primary_dx  = yes_conditions[0]
            primary_icd = comorbidity_to_raw_icd.get(primary_dx, "")
            if not primary_icd:
                primary_icd = _canonical.get(primary_dx, "")

        # Tier 3 — description keyword fallback when no comorbidity was flagged via ICD matching
        if not primary_dx and icd_set:
            for code in sorted(icd_set):
                desc = icd_desc.get(code.upper(), "")
                if not desc:
                    continue
                desc_lower = desc.lower()
                for col in COMORBIDITY_COLUMNS:
                    kws = _desc_kws.get(col)
                    if kws and any(kw in desc_lower for kw in kws):
                        primary_dx  = col
                        primary_icd = code
                        break
                if primary_dx:
                    break

        # SECONDARY DX / ICD — next YES comorbidity after primary
        secondary_dx = secondary_icd = ""
        for col in yes_conditions:
            if col == primary_dx:
                continue
            secondary_dx  = col
            secondary_icd = comorbidity_to_raw_icd.get(col, "")
            if not secondary_icd:
                secondary_icd = _canonical.get(secondary_dx, "")
            break

        comorbidity_vals = comorbidities
        raw_icd_str = ", ".join(sorted(icd_set))

        ins_name = visit.get("insurance", "")
        ins_type = carrier_map.get(ins_name.upper(), "")
        medicare_id = visit.get("insurance_id", "") if "medicare" in ins_type.lower() else ""

        last_seen = visit.get("last_seen_date", None)   # datetime or None
        raw_provider = r["provider"]

        row = {col: "" for col in TEMPLATE_COLUMNS}
        row["_RAW_ICD"]            = raw_icd_str          # temp — dropped before output
        row["EMR ID"]              = r["chart_no"]
        row["PATIENT EMR NAME"]    = emr_name
        row["FIRST NAME"]          = first
        row["MIDDLE NAME"]         = middle
        row["LAST NAME"]           = last
        row["PATIENT FULL NAME"]   = full_name
        row["DATE OF BIRTH"]       = r["birthdate"]   # datetime object
        _sex = r["sex"].strip().upper()
        row["GENDER"]              = "Male" if _sex in ("M", "MALE") else "Female" if _sex in ("F", "FEMALE") else r["sex"]
        row["STREET ADDRESS"]      = street
        row["CITY"]                = r["city"]
        row["STATE"]               = r["state"]
        row["ZIP"]                 = r["zipcode"]
        row["HOME PHONE"]          = home_phone
        row["MOBILE PHONE"]        = mobile_phone
        row["WORK PHONE"]          = profile_office
        row["EMAIL ADDRESS"]       = r["email_address"]
        row["LANGUAGE"]            = r["language"]
        row["RACE"]                = visit.get("race", "")
        row["MEDICARE ID"]         = medicare_id
        row["PRIMARY INSURANCE"]   = ins_name
        row["PRIMARY ID"]          = visit.get("insurance_id", "")
        row["PRIMARY GROUP"]       = visit.get("group_code", "")
        row["SECONDARY INSURANCE"] = sec_insurer_name
        row["INSURANCE TYPE"]      = ins_type
        row["CO-PAY"]              = visit.get("copay", "")
        row.update(comorbidity_vals)
        row["PRIMARY DX"]          = primary_dx
        row["SECONDARY DX"]        = secondary_dx
        row["PRIMARY ICD"]         = primary_icd
        row["SECONDARY ICD"]       = secondary_icd
        row["LAST SEEN DATE"]      = last_seen         # datetime object or None
        chart_no_key = r["chart_no"]
        if future_appt_map is not None:
            next_appt = future_appt_map.get(chart_no_key, None)
        else:
            next_appt = visit.get("next_appt", None) if last_seen else None
        row["NEXT APPT"]           = next_appt
        row["PROVIDER DATA"]       = raw_provider
        row["PROVIDER NAME"]       = parse_provider_name(raw_provider)
        row["CLINIC FACILITY"]       = CLINIC_FACILITY
        row["PRIMARY CARE PROVIDER"] = pcp_map.get(r["chart_no"], "")

        # Skip patients for whom no primary diagnosis could be determined
        if not primary_dx:
            continue

        rows.append(row)

    # Build df with temp column, then drop it
    all_cols = ["_RAW_ICD"] + TEMPLATE_COLUMNS
    df = pd.DataFrame(rows, columns=all_cols)
    df.drop(columns=["_RAW_ICD"], inplace=True)
    return df


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    base              = os.path.join(os.path.dirname(__file__), "..")
    demo_path         = os.path.join(base, "PatientDemo_2Line.xlsx")
    profile_path      = os.path.join(base, "PatientManagementProfileDetail.xlsx")
    analysis_path     = os.path.join(base, "PatientManagementAnalysisOfVisitsDetail.xlsx")
    carrier_path      = os.path.join(base, "CarrierCode.xlsx")
    profitability_path = os.path.join(base, "Profitability.xlsx")
    diagnosis_path    = os.path.join(base, "DiagnosisCodeProductionDetail.xlsx")
    template_dir      = os.path.join(base, "template")
    output_dir    = os.path.join(base, "output")
    os.makedirs(output_dir, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"ACNY_consolidated_{timestamp}.xlsx")

    # Load allergy ICD map from pre-built JSON (one-time scan of raw diagnosis files)
    print(f"\nLoading allergy ICD map from acny_allergy_icd_map.json ...")
    allergy_icd_prefixes, allergy_desc_keywords, allergy_canonical_icd = load_allergy_icd_map(template_dir)
    print(f"  Allergy ICD map loaded for {len(allergy_icd_prefixes)} allergy columns.")

    # Locate prescription CSV
    csv_files = glob.glob(os.path.join(template_dir, "api_prescription*.csv"))
    if not csv_files:
        print("WARNING: api_prescription*.csv not found in template/ — only allergy ICD patterns will be active.")
        comorbidity_icd_map = {col: prefixes for col, prefixes in allergy_icd_prefixes.items() if prefixes}
    else:
        csv_path = csv_files[0]
        print(f"Reading {csv_path} ...")
        comorbidity_icd_map = parse_prescription_csv(csv_path, allergy_icd_prefixes)
        print(f"  Comorbidity ICD map loaded for {len(comorbidity_icd_map)} conditions.")

    # Parse PatientDemo
    print(f"\nReading {demo_path} ...")
    wb = load_workbook(demo_path, read_only=True, data_only=True)
    all_records = []
    for sheet_name in SHEETS_DEMO:
        if sheet_name not in wb.sheetnames:
            continue
        records = parse_demo_sheet(wb[sheet_name])
        print(f"  {sheet_name}: {len(records)} patient records parsed.")
        all_records.extend(records)
    wb.close()
    print(f"  Total: {len(all_records)} records")

    # Parse phone profile
    print(f"\nReading {profile_path} ...")
    wb2 = load_workbook(profile_path, read_only=True, data_only=True)
    phone_map = parse_profile(wb2.active)
    wb2.close()
    print(f"  Phone data loaded for {len(phone_map)} unique chart numbers.")

    # Parse visit files
    print(f"\nReading CustomVisitListing files ...")
    visit_map = parse_visit_files(base)
    print(f"  Visit data loaded for {len(visit_map)} unique chart numbers.")

    # Parse carrier codes
    print(f"\nReading {carrier_path} ...")
    carrier_map = parse_carrier_codes(carrier_path)
    print(f"  Carrier type map loaded for {len(carrier_map)} carriers.")

    # Parse analysis of visits
    print(f"\nReading {analysis_path} ...")
    pcp_map = parse_analysis_of_visits(analysis_path)
    print(f"  Primary care provider loaded for {len(pcp_map)} chart numbers.")

    # Parse qualified registry (secondary insurance + ICD supplement)
    registry_path = os.path.join(base, "QualifiedRegistry.xlsx")
    if os.path.exists(registry_path):
        print(f"\nReading {registry_path} ...")
        registry_map = parse_qualified_registry(registry_path)
        has_sec = sum(1 for v in registry_map.values() if v["sec_insurer"])
        has_icd = sum(1 for v in registry_map.values() if v["icd_codes"])
        print(f"  Registry loaded for {len(registry_map)} patients "
              f"({has_sec} with secondary insurer, {has_icd} with ICD codes).")
    else:
        print("\nWARNING: QualifiedRegistry.xlsx not found.")
        registry_map = None

    # Parse future appointments
    future_appt_path = os.path.join(base, "AppointmentConfirmationDetail _Future.xlsx")
    if os.path.exists(future_appt_path):
        print(f"\nReading {future_appt_path} ...")
        future_appt_map = parse_future_appointments(future_appt_path)
        print(f"  Future appointments loaded for {len(future_appt_map)} chart numbers.")
    else:
        print("\nWARNING: AppointmentConfirmationDetail _Future.xlsx not found — NEXT APPT will use visit data.")
        future_appt_map = None

    # Build ICD description lookup
    print(f"\nBuilding ICD description lookup from Profitability and DiagnosisCodeProductionDetail ...")
    icd_desc = parse_icd_descriptions(profitability_path, diagnosis_path)
    print(f"  ICD descriptions loaded for {len(icd_desc)} unique codes.")

    # Build consolidated output
    print("\nBuilding consolidated output ...")
    df = build_consolidated(all_records, phone_map, visit_map, carrier_map, pcp_map,
                            comorbidity_icd_map, icd_desc, future_appt_map, registry_map,
                            allergy_desc_keywords=allergy_desc_keywords,
                            allergy_canonical_icd=allergy_canonical_icd)

    before = len(df)
    df.drop_duplicates(inplace=True)
    after = len(df)
    print(f"  Removed {before - after} duplicates. {after} records remain.")

    # Filter: remove only patients with no ICD codes at all in visit data
    before_filter = len(df)
    has_icd = df["EMR ID"].map(
        lambda chart: bool(visit_map.get(chart, {}).get("icd_codes"))
                      or bool((registry_map or {}).get(chart, {}).get("icd_codes"))
    )
    df = df[has_icd].reset_index(drop=True)
    after_filter = len(df)
    print(f"  Removed {before_filter - after_filter} rows with no ICD data. Final record count: {after_filter}")

    has_primary_icd = (df["PRIMARY ICD"] != "").sum()
    has_secondary   = (df["SECONDARY ICD"] != "").sum()
    has_asthma      = (df["ASTHMA"] == "YES").sum()
    has_insurance   = (df["INSURANCE TYPE"] != "").sum()
    has_medicare    = (df["MEDICARE ID"] != "").sum()
    has_pcp         = (df["PRIMARY CARE PROVIDER"] != "").sum()
    has_next_appt   = df["NEXT APPT"].notna().sum()
    print(f"  PRIMARY ICD populated         : {has_primary_icd}")
    print(f"  SECONDARY ICD populated       : {has_secondary}")
    print(f"  ASTHMA = YES                  : {has_asthma}")
    print(f"  INSURANCE TYPE populated      : {has_insurance}")
    print(f"  MEDICARE ID populated         : {has_medicare}")
    print(f"  PRIMARY CARE PROVIDER populated: {has_pcp}")
    print(f"  NEXT APPT populated            : {has_next_appt}")

    df.to_excel(output_path, index=False)

    # Apply MM-DD-YYYY date format to date columns
    from openpyxl import load_workbook as _lw
    from openpyxl.styles import numbers as _nums
    _wb = _lw(output_path)
    _ws = _wb.active
    header = [cell.value for cell in _ws[1]]
    date_fmt = "MM-DD-YYYY"
    for col_name in DATE_COLUMNS:
        if col_name in header:
            col_idx = header.index(col_name) + 1   # 1-based
            for row_cells in _ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row_cells[0]
                if cell.value is not None:
                    cell.number_format = date_fmt
    _wb.save(output_path)
    _wb.close()

    print(f"\nOutput written to: {output_path}")


if __name__ == "__main__":
    main()
