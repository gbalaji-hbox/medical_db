#!/usr/bin/env python3
"""
HCT consolidation pipeline.

Workflow:
1) Generate cleaned demographics, insurance, and ICD CSV files.
2) Merge them by EMR ID (Per Nbr as text key) into template column order.
3) Export final CSV and formatted XLSX in src/HCT/output.
4) Delete intermediate CSV files, keep only final combined CSV.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

from clean_demographics import transform_demographics
from clean_insurance import transform_insurance
from clean_icd_codes import build_grouped_icd


INSURANCE_COLUMNS = [
    "MEDICARE ID",
    "PRIMARY INSURANCE",
    "PRIMARY ID",
    "PRIMARY GROUP",
    "SECONDARY INSURANCE",
    "SECONDARY ID",
    "SECONDARY GROUP",
    "TERITARY INSURANCE",
    "TERITARY ID",
    "TERITARY GROUP",
    "INSURANCE TYPE",
    "CO-PAY",
]

ICD_COLUMNS = [
    "CORONARY ARTERY DISEASE",
    "ARRHYTHMIA",
    "CONGESTIVE HEART FAILURE",
    "PERIPHERAL VASCULAR",
    "VALVULAR HEART",
    "CERBOVASCULAR ACCIDENT",
    "HYPERLIPIDEMIA",
    "ANGINA PECTORIS",
    "HYPOTENSION",
    "HYPERTENSION",
    "OBESITY",
    "DIABETES",
    "CHRONIC KIDNEY DISEASE",
    "COPD",
    "RESPIRATORY FAILURE",
    "ASTHMA",
    "SLEEP APNEA",
    "DYSPNEA",
    "EMPHYSEMA",
    "BRONCHIECTASIS",
    "HYPOXEMIA",
    "PRIMARY DX",
    "SECONDARY DX",
    "PRIMARY ICD",
    "SECONDARY ICD",
]


def clean_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def normalize_key(value: object) -> str:
    return clean_text(value)


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}")


def is_date_header(header: str) -> bool:
    h = clean_text(header).upper()
    return ("DATE" in h) or h in {"DOB"}


def build_excel_dataframe_with_dates(df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep CSV data unchanged, but convert date-like columns to real datetimes
    for Excel so filtering/sorting works as date fields.
    """
    excel_df = df.copy()
    for col in excel_df.columns:
        if not is_date_header(col):
            continue

        parsed = pd.to_datetime(excel_df[col], format="%m-%d-%Y", errors="coerce")
        if parsed.notna().any():
            excel_df[col] = parsed

    return excel_df


def fill_without_overwrite(base: pd.DataFrame, incoming: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """
    Fill only blank cells in base from incoming, keyed by EMR ID index.
    """
    result = base.copy()
    for col in columns:
        if col not in result.columns or col not in incoming.columns:
            continue

        base_series = result[col].map(clean_text)
        incoming_series = incoming[col].map(clean_text).reindex(result.index)
        fill_mask = base_series == ""
        result.loc[fill_mask, col] = incoming_series.loc[fill_mask]

    return result


def format_excel_file(xlsx_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    date_col_indexes: set[int] = set()

    for column_cells in ws.columns:
        max_length = 0
        header = clean_text(column_cells[0].value)
        if is_date_header(header):
            date_col_indexes.add(column_cells[0].column)

        col_letter = column_cells[0].column_letter
        for cell in column_cells[:1000]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in date_col_indexes and isinstance(cell.value, (datetime, date)):
                cell.number_format = "MM-DD-YYYY"

    wb.save(xlsx_path)


def remove_intermediate_csvs(cleaned_dir: Path, output_dir: Path, final_csv: Path) -> None:
    for csv_file in cleaned_dir.glob("*.csv"):
        if csv_file.exists():
            csv_file.unlink()

    for csv_file in output_dir.glob("*.csv"):
        if csv_file.resolve() != final_csv.resolve() and csv_file.exists():
            csv_file.unlink()


def run_pipeline(base_dir: Path) -> tuple[Path, Path]:
    template_file = base_dir / "template" / "consolidated_view-template - new.xlsx"
    demographics_input = base_dir / "patient-demographics.xlsx"
    insurance_input = base_dir / "patient-insurance.xlsx"
    icd_input = base_dir / "HCT_Location Wise_Provider Wise_Patient Wise_ICD Codes.xlsx"
    api_file = base_dir / "template" / "api_prescriptioncauselist_202603101243.csv"

    cleaned_dir = base_dir / "cleaned"
    output_dir = base_dir / "output"
    cleaned_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    log("Starting HCT consolidation pipeline")
    log("Generating cleaned demographics")

    demographics_csv = transform_demographics(demographics_input, template_file, cleaned_dir)
    log(f"Demographics complete: {demographics_csv.name}")

    log("Generating cleaned insurance")
    insurance_csv = transform_insurance(insurance_input, template_file, cleaned_dir)
    log(f"Insurance complete: {insurance_csv.name}")

    log("Generating cleaned ICD")
    icd_csv = build_grouped_icd(icd_input, api_file, template_file, cleaned_dir)
    log(f"ICD complete: {icd_csv.name}")

    log("Loading cleaned outputs")
    template_headers = list(pd.read_excel(template_file, nrows=0).columns)

    dem_df = pd.read_csv(demographics_csv, dtype=str).fillna("")
    ins_df = pd.read_csv(insurance_csv, dtype=str).fillna("")
    icd_df = pd.read_csv(icd_csv, dtype=str).fillna("")

    dem_df["EMR ID"] = dem_df["EMR ID"].map(normalize_key)
    ins_df["EMR ID"] = ins_df["EMR ID"].map(normalize_key)
    icd_df["EMR ID"] = icd_df["EMR ID"].map(normalize_key)

    dem_df = dem_df[dem_df["EMR ID"] != ""].drop_duplicates(subset=["EMR ID"], keep="first")
    ins_df = ins_df[ins_df["EMR ID"] != ""].drop_duplicates(subset=["EMR ID"], keep="first")
    icd_df = icd_df[icd_df["EMR ID"] != ""].drop_duplicates(subset=["EMR ID"], keep="first")

    log(f"Rows after dedupe - demographics: {len(dem_df)}, insurance: {len(ins_df)}, icd: {len(icd_df)}")

    dem_df = dem_df.set_index("EMR ID", drop=False)
    ins_df = ins_df.set_index("EMR ID", drop=False)
    icd_df = icd_df.set_index("EMR ID", drop=False)

    insurance_with_primary = ins_df
    if "PRIMARY INSURANCE" in insurance_with_primary.columns:
        insurance_with_primary = insurance_with_primary[
            insurance_with_primary["PRIMARY INSURANCE"].map(clean_text) != ""
        ]

    icd_with_primary_dx = icd_df
    if "PRIMARY DX" in icd_with_primary_dx.columns:
        icd_with_primary_dx = icd_with_primary_dx[
            icd_with_primary_dx["PRIMARY DX"].map(clean_text) != ""
        ]

    matching_ids = dem_df.index.intersection(insurance_with_primary.index).intersection(icd_with_primary_dx.index)
    final_df = dem_df.loc[matching_ids].copy()

    log(
        "Matched rows after strict key filter "
        f"(demographics + primary insurance + primary dx): {len(final_df)}"
    )

    log("Merging insurance into demographics")
    final_df = fill_without_overwrite(final_df, ins_df, INSURANCE_COLUMNS)
    log("Merging ICD into demographics")
    final_df = fill_without_overwrite(final_df, icd_df, ICD_COLUMNS)

    # Safety filter: keep only complete rows even if source format changes.
    if "PRIMARY INSURANCE" in final_df.columns:
        final_df = final_df[final_df["PRIMARY INSURANCE"].map(clean_text) != ""]
    if "PRIMARY DX" in final_df.columns:
        final_df = final_df[final_df["PRIMARY DX"].map(clean_text) != ""]

    final_df = final_df.reset_index(drop=True)

    # Keep strict template order only in final output.
    final_df = final_df.reindex(columns=template_headers, fill_value="")

    log(f"Final merged row count: {len(final_df)}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_csv = output_dir / f"HCT_Consolidated_{timestamp}.csv"
    final_xlsx = output_dir / f"HCT_Consolidated_{timestamp}.xlsx"

    log(f"Writing final CSV: {final_csv.name}")
    final_df.to_csv(final_csv, index=False, encoding="utf-8-sig")

    excel_df = build_excel_dataframe_with_dates(final_df)
    log(f"Writing final XLSX: {final_xlsx.name}")
    excel_df.to_excel(final_xlsx, index=False)
    log("Formatting final XLSX")
    format_excel_file(final_xlsx)

    log("Removing intermediate CSV files")
    remove_intermediate_csvs(cleaned_dir, output_dir, final_csv)
    log("Pipeline complete")

    return final_csv, final_xlsx


def main() -> None:
    parser = argparse.ArgumentParser(description="Run full HCT consolidation pipeline.")
    parser.add_argument(
        "--base-dir",
        default=str(Path(__file__).resolve().parents[1]),
        help="HCT base directory",
    )
    args = parser.parse_args()

    final_csv, final_xlsx = run_pipeline(Path(args.base_dir))
    print(f"Created CSV: {final_csv}")
    print(f"Created XLSX: {final_xlsx}")


if __name__ == "__main__":
    main()
