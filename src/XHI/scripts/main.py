#!/usr/bin/env python3
"""
XHI Data Consolidator

Consolidates raw CSV files (EMR, medication, problem reports) into the template format
(consolidated_view-template - new.xlsx) with:
- Patient data merging by Chart ID
- Name parsing from Full Name
- Insurance type classification
- Comorbidity mapping from Problem Report
- Medication selection (most recent)
- Missing column handling
"""

import os
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import re
from datetime import datetime
import openpyxl

# Paths
ROOT = os.environ.get('MEDICAL_DB_ROOT', r'D:\Work_Folder\medical_db')
EMR_FILE = os.path.join(ROOT, 'src', 'XHI', 'EMR_Final_Report_final_part.csv')
MED_FILE = os.path.join(ROOT, 'src', 'XHI', 'medication_report.csv')
PROB_FILE = os.path.join(ROOT, 'src', 'XHI', 'problem_report.csv')
PRESCRIPTION_FILE = os.path.join(ROOT, 'src', 'XHI', 'template', 'api_prescriptioncauselist_202603101243.csv')
TEMPLATE_FILE = os.path.join(ROOT, 'src', 'XHI', 'template', 'consolidated_view-template - new.xlsx')
# Output file with date_time
now = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT_FILE = os.path.join(ROOT, 'src', 'XHI', 'output', f'XHI_consolidated_{now}.xlsx')

# Insurance classification (from MCA scripts)
INSURANCE_TYPE_MAP = [
    (['medicare advantage', 'med adv', 'med adv', 'medicareadvantage', 'medicare adv', 'medicare plus'], 'medicare advantage'),
    (['medicare part', 'medicare/medicare', 'medicare rail', 'medicare'], 'medicare'),
    (['medicaid', 'health link', 'mi health link', 'healthy mi', 'medicaid hmo'], 'medicaid'),
    (['dual', 'd-snp', 'duals', 'dual complete', 'd-snp', 'snp'], 'medicare advantage'),
    (['advantage', 'med adv', 'med adv ppo', 'med adv hmo', 'med adv ppo'], 'medicare advantage'),
    (['hap', 'blue', 'blue cross', 'blue care', 'bcn', 'aetna', 'cigna', 'humana', 'priority health', 'united', 'molina', 'meridian', 'wellcare', 'geha', 'align', 'aco', 'ascension', 'trinity', 'mcclaren', 'gravie', 'meritain', 'umr', 'pa i', 'pai', 'cofinity', 'cofinity', 'core', 'core', 'mutual of omaha', 'etna'], 'commercial'),
    (['self pay', 'self-pay', 'selfpay'], 'self pay'),
    (['workers compensation', 'wc ', 'wc '], 'workers comp'),
    (['motor vehicle', 'auto '], 'auto'),
    (['veterans', 'va', 'triwest', 'veterans administration', 'optum va'], 'veterans'),
    (['tricare'], 'government'),
]

# Comorbidity columns in order (for left-to-right primary/secondary dx)
COMORBIDITY_COLUMNS = [
    'CORONARY ARTERY DISEASE', 'ARRHYTHMIA', 'CONGESTIVE HEART FAILURE', 'PERIPHERAL VASCULAR',
    'VALVULAR HEART', 'CERBOVASCULAR ACCIDENT', 'HYPERLIPIDEMIA', 'ANGINA PECTORIS',
    'HYPOTENSION', 'HYPERTENSION', 'OBESITY', 'DIABETES', 'CHRONIC KIDNEY DISEASE',
    'COPD', 'RESPIRATORY FAILURE', 'ASTHMA', 'SLEEP APNEA', 'DYSPNEA', 'EMPHYSEMA',
    'BRONCHIECTASIS', 'HYPOXEMIA'
]

# Cause to comorbidity mapping (from MCA scripts)
CAUSE_TO_COMORBIDITY = {
    'Coronary Artery Disease': 'CORONARY ARTERY DISEASE',
    'Arrhythmia': 'ARRHYTHMIA',
    'CHF (Congestive Heart Failure)': 'CONGESTIVE HEART FAILURE',
    'Peripheral Vascular Disease': 'PERIPHERAL VASCULAR',
    'Valvular Heart Disease': 'VALVULAR HEART',
    'Cerebrovascular Accident': 'CERBOVASCULAR ACCIDENT',
    'Hyperlipidemia': 'HYPERLIPIDEMIA',
    'Angina Pectoris': 'ANGINA PECTORIS',
    'Hypotension': 'HYPOTENSION',
    'Hypertension': 'HYPERTENSION',
    'Obesity': 'OBESITY',
    'Type 2 Diabetes': 'DIABETES',
    'Chronic Kidney': 'CHRONIC KIDNEY DISEASE',
    'COPD': 'COPD',
    'Chronic Bronchitis': 'COPD',
    'Respiratory Failure': 'RESPIRATORY FAILURE',
    'Asthma': 'ASTHMA',
    'Sleep Apnea': 'SLEEP APNEA',
    'Dyspnea': 'DYSPNEA',
    'Emphysema': 'EMPHYSEMA',
    'Emphysema ': 'EMPHYSEMA',
    'Bronchiectasis': 'BRONCHIECTASIS',
    'Hypoxemia': 'HYPOXEMIA',
    'Chronic Hypoxia': 'HYPOXEMIA'
}

COMORBIDITY_TO_CAUSE = {v: k for k, v in CAUSE_TO_COMORBIDITY.items()}

def load_icd_to_cause():
    """Load ICD to cause mapping from prescription file."""
    df = pd.read_csv(PRESCRIPTION_FILE)
    icd_to_cause = {}
    prefix_to_cause = {}
    for _, row in df.iterrows():
        icd = str(row.get('icd_code', '')).strip()
        cause = str(row.get('cause', '')).strip()
        if icd and cause:
            icd_to_cause[icd] = cause
            # Create prefix mapping (first char + next two digits)
            if len(icd) >= 3:
                prefix = icd[:3]
                prefix_to_cause[prefix] = cause
            else:
                prefix_to_cause[icd] = cause
    return icd_to_cause, prefix_to_cause

def classify_insurance(ins_text: str) -> str:
    if not ins_text or pd.isna(ins_text):
        return ''
    s = str(ins_text).lower()
    for patterns, label in INSURANCE_TYPE_MAP:
        if any(p in s for p in patterns):
            return label
    return 'other'

def parse_name(full_name: str) -> Dict[str, str]:
    """Parse first, middle, last from full name."""
    if pd.isna(full_name):
        return {'first': '', 'middle': '', 'last': ''}
    parts = str(full_name).strip().split()
    if len(parts) == 1:
        return {'first': parts[0], 'middle': '', 'last': ''}
    elif len(parts) == 2:
        return {'first': parts[0], 'middle': '', 'last': parts[1]}
    else:
        return {'first': parts[0], 'middle': ' '.join(parts[1:-1]), 'last': parts[-1]}

def format_date(date_str: str):
    """Convert date string to datetime object for Excel date format."""
    if pd.isna(date_str) or not date_str:
        return None
    try:
        dt = pd.to_datetime(date_str, format='%d-%m-%Y')
        return dt.to_pydatetime()
    except:
        return None

def map_comorbidities(icds: List[str], prefix_to_cause: Dict[str, str]) -> tuple[Dict[str, str], bool, Dict[str, str]]:
    """Map ICD codes to causes using prefix matching, then to comorbidity flags.
    
    Returns:
    - comorbidity flags dict
    - has_match boolean
    - comorbidity_to_raw_icd dict mapping comorbidity names to raw ICD codes
    """
    causes = set()
    has_match = False
    comorbidity_to_raw_icd = {}
    
    for icd in icds:
        icd = str(icd).strip()
        if not icd:
            continue
        icd_prefix = icd[:3] if len(icd) >= 3 else icd
        if icd_prefix in prefix_to_cause:
            cause = prefix_to_cause[icd_prefix]
            causes.add(cause)
            has_match = True
            # Map cause to comorbidity and store the raw ICD
            if cause in CAUSE_TO_COMORBIDITY:
                comorbidity = CAUSE_TO_COMORBIDITY[cause]
                if comorbidity not in comorbidity_to_raw_icd:
                    comorbidity_to_raw_icd[comorbidity] = icd
    
    flags = {col: 'NO' for col in COMORBIDITY_COLUMNS}
    for cause in causes:
        if cause in CAUSE_TO_COMORBIDITY:
            flags[CAUSE_TO_COMORBIDITY[cause]] = 'YES'
    
    return flags, has_match, comorbidity_to_raw_icd

def get_recent_medication(med_df: pd.DataFrame, chart_id: str) -> str:
    """Get most recent active medication for a patient."""
    patient_meds = med_df[(med_df['Chart ID'] == chart_id) & (med_df['Status'].str.lower() == 'active')].copy()
    if patient_meds.empty:
        return ''
    # Use Prescribed Datetime, fallback to Start Taking Datetime
    patient_meds.loc[:, 'datetime'] = pd.to_datetime(patient_meds['Prescribed Datetime'], format='%m/%d/%Y %I:%M %p', errors='coerce')
    patient_meds.loc[patient_meds['datetime'].isna(), 'datetime'] = pd.to_datetime(patient_meds['Start Taking Datetime'], format='%m/%d/%Y %I:%M %p', errors='coerce')
    recent = patient_meds.sort_values('datetime', na_position='last').iloc[-1] if not patient_meds['datetime'].isna().all() else patient_meds.iloc[0]
    return str(recent['Medication'])

def format_excel_file(file_path: str):
    """Format Excel file to set EMR ID as text and dates as date format."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find column indices
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    emr_id_col = header_row.index('EMR ID') + 1  # 1-based
    date_cols = []
    for col_name in ['DATE OF BIRTH', 'LAST SEEN DATE', 'NEXT APPT']:
        if col_name in header_row:
            date_cols.append(header_row.index(col_name) + 1)  # 1-based

    # Set EMR ID column to text format
    for row in range(2, ws.max_row + 1):  # Skip header
        cell = ws.cell(row=row, column=emr_id_col)
        cell.number_format = '@'  # Text format

    # Set date columns to date format
    for col in date_cols:
        for row in range(2, ws.max_row + 1):  # Skip header
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.number_format = 'mm/dd/yyyy'

    wb.save(file_path)

def main():
    print("Starting main")
    # Load mappings
    icd_to_cause, prefix_to_cause = load_icd_to_cause()
    print("Loaded mappings")

    # Load data — use python engine for prob/med to handle unquoted fields with
    # embedded commas/newlines in DrChrono exports (avoids pandas ParserError)
    emr_df  = pd.read_csv(EMR_FILE)
    med_df  = pd.read_csv(MED_FILE,  engine='python', on_bad_lines='skip')
    prob_df = pd.read_csv(PROB_FILE, engine='python', on_bad_lines='skip')
    print(f"Loaded data — emr:{len(emr_df)} med:{len(med_df)} prob:{len(prob_df)}")

    # Group EMR by Chart ID, taking most recent appointment data
    emr_df['Date of Last Appointment'] = pd.to_datetime(emr_df['Date of Last Appointment'], errors='coerce')
    emr_grouped = emr_df.sort_values(['Chart ID', 'Date of Last Appointment'], na_position='last').groupby('Chart ID').last().reset_index()
    print(f"EMR grouped from {len(emr_df)} to {len(emr_grouped)} rows")

    # Group medications and problems by Chart ID
    med_grouped = med_df.groupby('Chart ID')['Medication'].apply(list).reset_index()
    prob_grouped = prob_df.groupby('Chart ID').agg({
        'Problem': list,
        'ICD10 Code': list
    }).reset_index()

    # Merge all
    df = emr_grouped.merge(med_grouped, on='Chart ID', how='left')
    df = df.merge(prob_grouped, on='Chart ID', how='left')
    print("Merged, shape:", df.shape)
    print(f"Input records: {len(df)} records")

    # Fill NaN
    df = df.fillna('')

    # Apply mappings
    output_rows = []
    print("Starting processing", len(df), "rows")
    for i, row in df.iterrows():
        if i % 1000 == 0:
            print(f"Processed {i} rows")
        parsed_name = parse_name(row['Full Name'])
        
        # Skip deceased patients
        name_parts = [parsed_name['first'], parsed_name['middle'], parsed_name['last']]
        if any('deceased' in str(part).lower() for part in name_parts if part):
            continue
        
        icds = row.get('ICD10 Code', [])
        comorbidities, has_icd_match, comorbidity_to_raw_icd = map_comorbidities(icds, prefix_to_cause)
        recent_med = get_recent_medication(med_df, row['Chart ID'])

        # Primary and secondary dx: left-to-right comorbidity strategy
        primary_dx = ''
        secondary_dx = ''
        primary_icd = ''
        secondary_icd = ''
        true_comorbidities = [col for col in COMORBIDITY_COLUMNS if comorbidities[col] == 'YES']
        if true_comorbidities:
            primary_comorb = true_comorbidities[0]
            primary_dx = primary_comorb  # Use the column header as DX
            # Use raw ICD code that matched to this comorbidity
            primary_icd = comorbidity_to_raw_icd.get(primary_comorb, '')
            
            if len(true_comorbidities) > 1:
                secondary_comorb = true_comorbidities[1]
                secondary_dx = secondary_comorb  # Use the column header as DX
                # Use raw ICD code that matched to this comorbidity
                secondary_icd = comorbidity_to_raw_icd.get(secondary_comorb, '')

        # Handle clinic facility name - take only the part before the first comma
        clinic_facility = str(row['Practice Official Name']).split(',')[0].strip()

        output_row = {
            'EMR ID': row['Chart ID'],
            'PATIENT EMR NAME': row['Full Name'],
            'FIRST NAME': parsed_name['first'],
            'MIDDLE NAME': parsed_name['middle'],
            'LAST NAME': parsed_name['last'],
            'PATIENT FULL NAME': row['Full Name'],
            'DATE OF BIRTH': format_date(row['Date of Birth']),
            'GENDER': row['Gender'],
            'STREET ADDRESS': row['Address'],
            'CITY': row['City'],
            'STATE': row['State'],
            'ZIP': row['Zip Code'],
            'HOME PHONE': row['Home Phone'],
            'MOBILE PHONE': row['Cell Phone'],
            'WORK PHONE': row['Office Phone'],
            'EMAIL ADDRESS': row['Email'].lower() if pd.notna(row['Email']) else '',
            'LANGUAGE': 'English',
            'RACE': row['Race'],
            'EMERGENCY CONTACT NAME': row['Emerg Contact Name'],
            'EMERGENCY RELATIONSHIP': row['Emerg Contact Relation'],
            'EMERGENCY CONTACT HOME PHONE': row['Emerg Contact Phone'],
            'EMERGENCY CONTACT MOBILE PHONE': '',
            'MEDICARE ID': row['Primary Member ID'] if classify_insurance(row['Primary Ins Payer']) == 'medicare' else '',
            'PRIMARY INSURANCE': row['Primary Ins Payer'],
            'PRIMARY ID': row['Primary Member ID'],
            'PRIMARY GROUP': row['Primary Ins Group #'],
            'SECONDARY INSURANCE': row['Secondary Ins Payer'],
            'SECONDARY ID': row['Secondary Member ID'],
            'SECONDARY GROUP': row['Secondary Ins Group #'],
            'TERITARY INSURANCE': '',
            'TERITARY ID': '',
            'TERITARY GROUP': '',
            'INSURANCE TYPE': classify_insurance(row['Primary Ins Payer']),
            'CO-PAY': row['Expected Copay'],
            **comorbidities,
            'PRIMARY DX': primary_dx,
            'SECONDARY DX': secondary_dx,
            'PRIMARY ICD': primary_icd,
            'SECONDARY ICD': secondary_icd,
            'LAST SEEN DATE': format_date(row['Date of Last Appointment']),
            'NEXT APPT': format_date(row['Date of Next Appointment']),
            'PROVIDER DATA': row['Provider'],
            'PROVIDER NAME': row['Primary Provider'],
            'CLINIC FACILITY': clinic_facility,
            'PRIMARY CARE PROVIDER': row['Primary Care Physician'],
            'MEDICATIONS': recent_med,
            'ENCOUNTER NOTES': row['Appointment Notes'],
        }
        output_rows.append((output_row, has_icd_match))

    print("Processing done, creating dataframe")
    # Filter: only patients with primary insurance AND primary DX
    filtered_rows = [row for row, has_match in output_rows if row['PRIMARY INSURANCE'] and row['PRIMARY DX']]
    print(f"Filtered from {len(output_rows)} to {len(filtered_rows)} rows")
    print(f"Records output: {len(filtered_rows)} records")
    
    # Create output DataFrame
    output_df = pd.DataFrame(filtered_rows)

    # Load template to get column order
    template_df = pd.read_excel(TEMPLATE_FILE, nrows=0)
    template_cols = list(template_df.columns)

    # Reorder columns to match template
    output_df = output_df[template_cols]

    print("Saving")
    # Save to Excel
    output_df.to_excel(OUTPUT_FILE, index=False)
    print(f"Consolidated data saved to {OUTPUT_FILE}")

    # Format Excel file
    format_excel_file(OUTPUT_FILE)
    print(f"Excel formatting applied to {OUTPUT_FILE}")

if __name__ == '__main__':
    main()